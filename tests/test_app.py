import json
import os
import tempfile
import threading
import unittest
from decimal import Decimal
from io import BytesIO
from http.client import HTTPConnection
from http.server import ThreadingHTTPServer
from pathlib import Path
from urllib.parse import quote

import app
from openpyxl import load_workbook


class ParserTests(unittest.TestCase):
    def test_parse_amount_supports_russian_and_machine_formats(self):
        self.assertEqual(app.parse_amount("1 234 567,89"), 1234567.89)
        self.assertEqual(app.parse_amount("44 622 636,12"), 44622636.12)
        self.assertEqual(app.parse_amount("-37 206,75"), -37206.75)
        self.assertEqual(app.parse_amount("10000000.00"), 10000000.0)
        self.assertEqual(app.parse_amount(""), 0.0)
        self.assertEqual(app.parse_amount(None), 0.0)

    def test_json_safe_converts_decimal(self):
        payload = app.json_safe({"amount": Decimal("10.125"), "rows": [Decimal("0.10")]})
        self.assertEqual(payload, {"amount": 10.13, "rows": [0.1]})
        self.assertIsInstance(payload["amount"], float)

    def test_decimal_money_sum_avoids_float_artifact(self):
        self.assertEqual(app.money_sum([Decimal("0.10"), Decimal("0.20")]), Decimal("0.30"))

    def test_parse_date_normalizes_known_input_formats(self):
        self.assertEqual(app.parse_date("20.08.2025"), "2025-08-20")
        self.assertEqual(app.parse_date("2025-03-07 00:00:00.000"), "2025-03-07")
        self.assertEqual(app.parse_date("2026-04-01"), "2026-04-01")

    def test_normalize_code_removes_separators_and_keeps_cyrillic_letters(self):
        self.assertEqual(app.normalize_code("08.3.02.97070"), "0830297070")
        self.assertEqual(app.normalize_code("13.2.01.97003"), "1320197003")
        self.assertEqual(app.normalize_code("101016105Б"), "101016105Б")

    def test_find_header_value_matches_year_independent_prefix(self):
        row = {"Лимиты ПБС 2026 год": "123"}
        self.assertEqual(app.find_header_value(row, "Лимиты ПБС"), "123")

    def test_object_group_key_uses_code_and_budget_or_normalized_name(self):
        self.assertEqual(app.object_group_key({"object_code_norm": "001", "budget": "Областной бюджет"}), "001|областной бюджет")
        self.assertEqual(app.object_group_key({"object_name": "  Объект, СКК! ", "budget": ""}), "name:объект скк|")


class DataLoadTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.store = app.STORE

    def test_fastapi_app_is_importable(self):
        from analytics.api import app as fastapi_app

        self.assertIsNotNone(fastapi_app)

    def test_loads_expected_dataset_volume_and_sources(self):
        self.assertEqual(self.store.meta["records"], 4037)
        self.assertEqual(
            set(self.store.meta["sources"]),
            {"РЧБ", "Соглашения", "ГЗ: контракты", "ГЗ: платежи", "БУАУ"},
        )

    def test_detects_expected_budgets_and_snapshots(self):
        self.assertIn("Областной бюджет Амурской области", self.store.meta["budgets"])
        self.assertIn("Бюджет г. Тынды", self.store.meta["budgets"])
        self.assertIn("2025-02-01", self.store.meta["snapshots"])
        self.assertIn("2026-04-01", self.store.meta["snapshots"])

    def test_code_filter_finds_rcb_and_agreement_records(self):
        result = app.aggregate(
            app.apply_filters(
                self.store.records,
                {"code": ["970"], "start": ["2025-01-01"], "end": ["2026-04-01"]},
            )
        )
        self.assertEqual(result["count"], 997)
        self.assertGreaterEqual(len(result["rows"]), 10)
        top_sources = result["rows"][0]["sources"]
        self.assertIn("РЧБ", top_sources)
        self.assertIn("Соглашения", top_sources)
        self.assertGreater(result["totals"]["limit"], 0)
        self.assertGreater(result["totals"]["agreement"], 0)

    def test_text_budget_and_source_filters_are_combined(self):
        filtered = app.apply_filters(
            self.store.records,
            {
                "q": ["Тынды"],
                "budget": ["Бюджет г. Тынды"],
                "source": ["БУАУ"],
                "start": ["2025-08-01"],
                "end": ["2025-12-31"],
            },
        )
        self.assertGreater(len(filtered), 0)
        self.assertTrue(all(record["source"] == "БУАУ" for record in filtered))
        self.assertTrue(all("Тынды".lower() in " ".join(str(v) for v in record.values()).lower() for record in filtered))

    def test_aggregate_has_stable_shape_for_frontend(self):
        result = app.aggregate(self.store.records)
        self.assertEqual(
            set(result.keys()),
            {"totals", "rows", "details", "timeline", "count"},
        )
        self.assertLessEqual(len(result["rows"]), 300)
        self.assertLessEqual(len(result["details"]), 500)
        self.assertTrue(result["timeline"])
        for metric in ["limit", "obligation", "cash", "agreement", "contract", "payment", "buau"]:
            self.assertIn(metric, result["totals"])

    def test_records_have_trace_fields(self):
        record = self.store.records[0]
        self.assertIn("id", record)
        self.assertIn("source_file", record)
        self.assertIn("source_row", record)
        self.assertIn("raw", record)
        self.assertTrue(record["source_file"])
        self.assertTrue(record["source_row"])

    def test_metric_selection_limits_totals(self):
        result = app.aggregate(self.store.records, ["limit", "cash"])
        self.assertEqual(set(result["totals"]), {"limit", "cash"})
        self.assertGreater(result["totals"]["limit"], 0)
        self.assertGreater(result["totals"]["cash"], 0)

    def test_templates_match_expected_code_fragments(self):
        self.assertTrue(app.matches_template({"object_code_norm": "000006105Б"}, "skk"))
        self.assertTrue(app.matches_template({"object_code_norm": "00000975"}, "kik"))
        self.assertTrue(app.matches_template({"object_code_norm": "00000978"}, "kik"))
        self.assertTrue(app.matches_template({"object_code_norm": "00000970"}, "two_thirds"))
        self.assertTrue(app.matches_template({"object_code_norm": "", "kvr": "414"}, "okv"))

    def test_kik_template_matches_current_case_data(self):
        rows = app.apply_filters(
            self.store.records,
            {"template": ["kik"], "start": ["2025-02-01"], "end": ["2026-04-02"]},
        )
        result = app.aggregate(rows)
        self.assertGreater(len(rows), 0)
        self.assertGreater(len(result["rows"]), 0)
        self.assertGreater(result["totals"]["limit"], 0)
        self.assertTrue(all(record["object_code_norm"][5:8] in {"975", "978"} for record in rows if record["object_code_norm"]))

    def test_quick_actions_include_core_scenarios(self):
        self.assertGreaterEqual(len(app.quick_actions_payload()), 6)
        for code in ("show_skk", "show_kik", "show_two_thirds", "show_okv"):
            self.assertIn(code, app.QUICK_ACTIONS)
            self.assertIn("metrics", app.QUICK_ACTIONS[code])

    def test_demo_quick_action_is_first_and_uses_skk_problem_flow(self):
        first = app.quick_actions_payload()[0]
        self.assertEqual(first["code"], "demo_60s")
        self.assertEqual(first["template"], "skk")
        self.assertEqual(first["post_filter"], "execution_problems")
        self.assertEqual(first["open_view"], "problems")

    def test_low_execution_post_filter_keeps_problem_rows_only(self):
        metrics = ["limit", "cash", "payment", "buau"]
        result = app.aggregate(self.store.records, metrics)
        filtered = app.apply_aggregate_post_filter(result, "low_execution", metrics)
        self.assertGreater(len(filtered["rows"]), 0)
        self.assertLessEqual(len(filtered["rows"]), len(result["rows"]))
        for row in filtered["rows"]:
            plan = float(row.get("limit") or 0) + float(row.get("obligation") or 0)
            execution = float(row.get("cash") or 0) + float(row.get("payment") or 0) + float(row.get("buau") or 0)
            self.assertGreater(plan, 0)
            self.assertTrue(execution == 0 or execution / plan < 0.25)

    def test_query_as_of_does_not_sum_monthly_rcb_snapshots(self):
        records = [
            {"source": "РЧБ", "record_kind": "rcb_snapshot", "snapshot": "2025-01-01", "object_code_norm": "1", "object_code": "1", "object_name": "obj", "budget": "", "limit": 10, "obligation": 0, "cash": 0, "agreement": 0, "contract": 0, "payment": 0, "buau": 0},
            {"source": "РЧБ", "record_kind": "rcb_snapshot", "snapshot": "2025-02-01", "object_code_norm": "1", "object_code": "1", "object_name": "obj", "budget": "", "limit": 20, "obligation": 0, "cash": 0, "agreement": 0, "contract": 0, "payment": 0, "buau": 0},
        ]
        result = app.aggregate(app.select_as_of(records, "2025-02-15", {}))
        self.assertEqual(result["totals"]["limit"], 20)

    def test_rcb_2026_limit_and_obligation_columns_are_loaded(self):
        result = app.query_as_of({"date": ["2026-04-01"], "template": ["all"]})
        self.assertGreater(result["totals"]["limit"], 0)
        self.assertGreater(result["totals"]["obligation"], 0)

    def test_agreements_are_not_duplicated_across_monthly_snapshots(self):
        records = [
            {"source": "Соглашения", "record_kind": "agreement_snapshot", "snapshot": "2025-01-01", "document_id": "a1", "object_code_norm": "1", "object_code": "1", "object_name": "obj", "budget": "", "limit": 0, "obligation": 0, "cash": 0, "agreement": 50, "contract": 0, "payment": 0, "buau": 0},
            {"source": "Соглашения", "record_kind": "agreement_snapshot", "snapshot": "2025-02-01", "document_id": "a1", "object_code_norm": "1", "object_code": "1", "object_name": "obj", "budget": "", "limit": 0, "obligation": 0, "cash": 0, "agreement": 60, "contract": 0, "payment": 0, "buau": 0},
        ]
        result = app.aggregate(app.select_as_of(records, "2025-02-15", {}))
        self.assertEqual(result["totals"]["agreement"], 60)

    def test_reporting_dates_exclude_payment_only_dates(self):
        payload = app.reporting_dates_payload()
        dates = [item["date"] for item in payload]
        self.assertIn("2026-04-01", dates)
        self.assertNotIn("2026-04-02", dates)

    def test_compare_uses_as_of_semantics(self):
        result = app.compare_periods({"base": ["2025-02-01"], "target": ["2026-04-01"], "template": ["skk"], "metrics": ["limit,cash"]})
        self.assertEqual(result["view"], "as_of")
        self.assertEqual(result["available_dates"], app.STORE.meta["reporting_dates"])
        self.assertTrue(result["rows"])

    def test_problem_filters_no_cash_no_payments_no_documents(self):
        records = [
            {"source": "РЧБ", "record_kind": "rcb_snapshot", "snapshot": "2026-04-01", "object_code_norm": "1", "object_code": "1", "object_name": "no cash", "budget": "", "limit": 100, "obligation": 0, "cash": 0, "agreement": 10, "contract": 0, "payment": 1, "buau": 0},
            {"source": "РЧБ", "record_kind": "rcb_snapshot", "snapshot": "2026-04-01", "object_code_norm": "2", "object_code": "2", "object_name": "no docs", "budget": "", "limit": 100, "obligation": 0, "cash": 1, "agreement": 0, "contract": 0, "payment": 0, "buau": 0},
            {"source": "Соглашения", "record_kind": "agreement_snapshot", "snapshot": "2026-04-01", "object_code_norm": "3", "object_code": "3", "object_name": "no pay", "budget": "", "limit": 0, "obligation": 0, "cash": 0, "agreement": 10, "contract": 0, "payment": 0, "buau": 0},
        ]
        result = app.aggregate(records)
        self.assertEqual(len(app.apply_aggregate_post_filter(result, "no_cash", app.METRIC_KEYS)["rows"]), 1)
        self.assertEqual(len(app.apply_aggregate_post_filter(result, "no_documents", app.METRIC_KEYS)["rows"]), 1)
        self.assertEqual(len(app.apply_aggregate_post_filter(result, "no_payments", app.METRIC_KEYS)["rows"]), 1)

    def test_pipeline_fields_are_returned_for_rows(self):
        result = app.query_as_of({"date": ["2026-04-01"], "template": ["skk"]})
        row = result["rows"][0]
        self.assertIn("pipeline", row)
        self.assertIn("plan", row["pipeline"])
        self.assertIn("problem_reasons", row)
        self.assertIn("object_key", row)
        self.assertIn("match_confidence", row)

    def test_attention_summary_includes_next_actions(self):
        result = app.query_as_of({"date": ["2026-04-01"], "template": ["skk"]})
        actions = result["attention_summary"]["next_actions"]
        self.assertTrue(actions)
        self.assertIn("Скачать Excel", [item["label"] for item in actions])

    def test_next_actions_include_pdf_download(self):
        result = app.query_as_of({"date": ["2026-04-01"], "template": ["skk"]})
        actions = [item["action"] for item in result["attention_summary"]["next_actions"]]
        self.assertIn({"download": "pdf"}, actions)

    def test_top_risks_include_short_fields(self):
        result = app.query_as_of({"date": ["2026-04-01"], "template": ["skk"], "post_filter": ["execution_problems"]})
        top = result["attention_summary"]["top_risks"][0]
        self.assertIn("short_name", top)
        self.assertIn("object_code", top)
        self.assertIn("budget", top)

    def test_risk_score_and_level_are_added_to_rows(self):
        result = app.query_as_of({"date": ["2026-04-01"], "template": ["skk"]})
        row = result["rows"][0]
        self.assertIn("risk_score", row)
        self.assertIn("risk_level", row)
        self.assertIn("risk_label", row)
        self.assertIn("risk_explanation", row)

    def test_risk_distribution_source_fields_stay_human_readable(self):
        result = app.query_as_of({"date": ["2026-04-01"], "template": ["skk"]})
        row = result["rows"][0]
        self.assertIsInstance(row.get("risk_label"), str)
        self.assertIsInstance(row.get("risk_explanation"), list)
        self.assertNotIn("problem_reasons", row["risk_label"])
        self.assertNotIn("pipeline", " ".join(row["risk_explanation"]))

    def test_attention_summary_returns_human_bullets(self):
        result = app.query_as_of({"date": ["2026-04-01"], "template": ["skk"]})
        summary = result["attention_summary"]
        self.assertEqual(summary["title"], "Что требует внимания")
        self.assertTrue(summary["bullets"])
        text = " ".join(summary["bullets"])
        self.assertNotIn("problem_reasons", text)
        self.assertNotIn("pipeline", text)

    def test_problem_rows_are_sorted_by_risk(self):
        result = app.query_as_of({
            "date": ["2026-04-01"],
            "template": ["skk"],
            "post_filter": ["execution_problems"],
        })
        scores = [row["risk_score"] for row in result["rows"]]
        self.assertEqual(scores, sorted(scores, reverse=True))

    def test_risk_breakdown_matches_score(self):
        result = app.query_as_of({"date": ["2026-04-01"], "template": ["skk"]})
        row = result["rows"][0]
        breakdown = app.risk_breakdown(row)
        self.assertEqual(breakdown["version"], app.RISK_MODEL_VERSION)
        self.assertEqual(breakdown["score"], app.risk_score(row))
        self.assertEqual(row["risk_breakdown"]["score"], row["risk_score"])
        self.assertEqual(row["risk_breakdown"]["label"], row["risk_label"])

    def test_compare_insights_are_returned(self):
        result = app.compare_periods({
            "base": ["2025-02-01"],
            "target": ["2026-04-01"],
            "template": ["skk"],
        })
        self.assertIn("compare_insights", result)
        self.assertIn("bullets", result["compare_insights"])

    def test_object_detail_returns_first_query_row(self):
        result = app.query_as_of({"date": ["2026-04-01"], "template": ["skk"]})
        row = result["rows"][0]
        payload = app.object_detail({"date": ["2026-04-01"], "template": ["skk"], "object_key": [row["object_key"]]})
        self.assertEqual(payload["object_key"], row["object_key"])
        self.assertIn("pipeline", payload)
        self.assertIn("documents", payload)

    def test_object_detail_includes_risk(self):
        query = app.query_as_of({"date": ["2026-04-01"], "template": ["skk"]})
        key = query["rows"][0]["object_key"]
        detail = app.object_detail({"date": ["2026-04-01"], "template": ["skk"], "object_key": [key]})
        self.assertIn("risk_score", detail)
        self.assertIn("risk_label", detail)

    def test_object_detail_includes_risk_breakdown(self):
        query = app.query_as_of({"date": ["2026-04-01"], "template": ["skk"]})
        key = query["rows"][0]["object_key"]
        detail = app.object_detail({"date": ["2026-04-01"], "template": ["skk"], "object_key": [key]})
        self.assertEqual(detail["risk_breakdown"]["version"], app.RISK_MODEL_VERSION)
        self.assertEqual(detail["risk_breakdown"]["score"], detail["risk_score"])
        self.assertTrue(detail["risk_breakdown"]["factors"])

    def test_object_detail_includes_review(self):
        original_path = app.REVIEWS_PATH
        with tempfile.TemporaryDirectory() as temp_dir:
            app.REVIEWS_PATH = Path(temp_dir) / "reviews.json"
            try:
                query = app.query_as_of({"date": ["2026-04-01"], "template": ["skk"]})
                key = query["rows"][0]["object_key"]
                detail = app.object_detail({"date": ["2026-04-01"], "template": ["skk"], "object_key": [key]})
                self.assertEqual(detail["review"]["status"], "new")
                self.assertEqual(detail["review"]["label"], "Новый")
            finally:
                app.REVIEWS_PATH = original_path

    def test_unreviewed_post_filter(self):
        original_path = app.REVIEWS_PATH
        with tempfile.TemporaryDirectory() as temp_dir:
            app.REVIEWS_PATH = Path(temp_dir) / "reviews.json"
            try:
                query = app.query_as_of({"date": ["2026-04-01"], "template": ["skk"], "post_filter": ["execution_problems"]})
                key = query["rows"][0]["object_key"]
                app.update_review(key, {"status": "checked"})
                filtered = app.query_as_of({"date": ["2026-04-01"], "template": ["skk"], "post_filter": ["unreviewed"]})
                self.assertNotIn(key, {row["object_key"] for row in filtered["rows"]})
                self.assertTrue(all(row["review"]["status"] in {"new", "in_progress"} for row in filtered["rows"]))
            finally:
                app.REVIEWS_PATH = original_path

    def test_old_scenarios_work_without_reviews_file(self):
        original_path = app.REVIEWS_PATH
        with tempfile.TemporaryDirectory() as temp_dir:
            app.REVIEWS_PATH = Path(temp_dir) / "missing" / "reviews.json"
            try:
                result = app.query_as_of({"date": ["2026-04-01"], "template": ["skk"]})
                self.assertTrue(result["rows"])
                self.assertEqual(result["rows"][0]["review"]["status"], "new")
            finally:
                app.REVIEWS_PATH = original_path

    def test_object_detail_includes_diagnosis(self):
        query = app.query_as_of({"date": ["2026-04-01"], "template": ["skk"], "post_filter": ["execution_problems"]})
        key = query["rows"][0]["object_key"]
        detail = app.object_detail({"date": ["2026-04-01"], "template": ["skk"], "object_key": [key]})
        self.assertEqual(detail["diagnosis"]["title"], "Что проверить")
        self.assertTrue(detail["diagnosis"]["bullets"])

    def test_readiness_summary_returns_checks(self):
        payload = app.readiness_response({"date": ["2026-04-01"], "template": ["skk"]})
        self.assertEqual(payload["date"], "2026-04-01")
        self.assertTrue(payload["checks"])
        self.assertIn("summary", payload)

    def test_as_of_timeline_uses_reporting_dates_and_matches_totals(self):
        result = app.query_as_of({"date": ["2026-04-01"], "template": ["skk"], "metrics": ["limit,cash"]})
        dates = [point["date"] for point in result["timeline"]]
        self.assertTrue(dates)
        self.assertTrue(all(date in app.STORE.meta["reporting_dates"] for date in dates))
        self.assertEqual(dates[-1], "2026-04-01")
        self.assertEqual(result["timeline"][-1]["limit"], result["totals"]["limit"])
        self.assertEqual(result["timeline"][-1]["cash"], result["totals"]["cash"])

    def test_empty_readiness_marks_empty_result(self):
        payload = app.readiness_response({"date": ["2026-04-01"], "template": ["skk"], "q": ["zzzz-no-data"]})
        empty = next(check for check in payload["checks"] if check["code"] == "empty_result")
        self.assertIn(empty["status"], {"warn", "bad"})

    def test_quick_actions_use_as_of_date(self):
        self.assertEqual(app.QUICK_ACTIONS["execution_problems"]["post_filter"], "execution_problems")
        self.assertEqual(app.default_date_range()[1], app.STORE.meta["reporting_dates"][-1])

    def test_assistant_rule_based_core_intents(self):
        skk = app.assistant_rule_based("Покажи СКК", {})
        self.assertEqual(skk["action"]["template"], "skk")
        self.assertIn("followups", skk)
        self.assertEqual(skk["alternatives"][0]["label"], "Показать все данные")
        self.assertEqual(skk["alternatives"][0]["action"]["q"], "")

        city = app.assistant_rule_based("6105 Благовещенск", {})
        self.assertEqual(city["action"]["template"], "skk")
        self.assertIn("Благовещенск", city["action"]["q"])
        self.assertEqual(city["alternatives"][0]["label"], "Искать во всех данных")
        self.assertTrue(city["alternatives"][0]["action"]["reset_scope"])
        self.assertEqual(city["alternatives"][0]["action"]["code"], "")
        self.assertEqual(city["alternatives"][0]["action"]["budget"], "")
        self.assertEqual(city["alternatives"][0]["action"]["source"], "")

        compare = app.assistant_rule_based("сравни СКК", {})
        self.assertEqual(compare["intent"], "run_compare")
        self.assertEqual(compare["action"]["mode"], "compare")

        explain = app.assistant_rule_based("что такое БО", {})
        self.assertEqual(explain["intent"], "explain_metric")

    def test_env_loader_reads_local_env_file(self):
        with tempfile.NamedTemporaryFile("w", encoding="utf-8", delete=False) as handle:
            handle.write("CODEX_TEST_ENV_LOADER='ok'\n")
            path = Path(handle.name)
        try:
            os.environ.pop("CODEX_TEST_ENV_LOADER", None)
            app.load_local_env(path)
            self.assertEqual(os.environ.get("CODEX_TEST_ENV_LOADER"), "ok")
        finally:
            os.environ.pop("CODEX_TEST_ENV_LOADER", None)
            path.unlink(missing_ok=True)

    def test_env_loader_missing_file_is_noop(self):
        app.load_local_env(app.ROOT / "missing-test-env-file")

    def test_validate_assistant_action_rejects_unknown_fields(self):
        fallback = app.assistant_rule_based("Покажи СКК", {})["action"]
        action = app.validate_assistant_action(
            {
                "mode": "slice",
                "template": "skk",
                "date": "2026-04-01",
                "post_filter": "no_cash",
                "endpoint": "/api/trace",
                "q": "x" * 250,
            },
            fallback,
        )
        self.assertNotIn("endpoint", action)
        self.assertEqual(action["q"], "x" * 200)
        self.assertEqual(action["post_filter"], "no_cash")

    def test_assistant_schema_contains_only_allowed_actions(self):
        schema = app.assistant_json_schema()
        action_properties = schema["properties"]["action"]["properties"]
        self.assertEqual(set(action_properties), app.ASSISTANT_ACTION_FIELDS)
        self.assertNotIn("endpoint", action_properties)

    def test_assistant_complex_problem_query_maps_to_problem_filter(self):
        old_key = os.environ.pop("GROQ_API_KEY", None)
        self.addCleanup(lambda: os.environ.__setitem__("GROQ_API_KEY", old_key) if old_key is not None else os.environ.pop("GROQ_API_KEY", None))
        payload = app.assistant_response("покажи проблемные СКК без кассы", {"mode": "slice", "template": "all"})
        self.assertEqual(payload["action"]["template"], "skk")
        self.assertIn(payload["action"]["post_filter"], {"execution_problems", "no_cash"})
        self.assertEqual(payload["action"]["open_view"], "problems")

    def test_assistant_rule_based_understands_dates_and_service_actions(self):
        payload = app.assistant_rule_based("покажи СКК за март 2026", {"mode": "slice", "template": "all"})
        self.assertEqual(payload["action"]["template"], "skk")
        self.assertEqual(payload["action"]["date"], "2026-03-01")
        self.assertEqual(payload["action"]["q"], "")

        compare = app.assistant_rule_based("сравни СКК февраль 2025 и апрель 2026", {"mode": "slice", "template": "all"})
        self.assertEqual(compare["action"]["mode"], "compare")
        self.assertEqual(compare["action"]["base"], "2025-02-01")
        self.assertEqual(compare["action"]["target"], "2026-04-01")
        self.assertEqual(compare["action"]["open_view"], "changes")
        self.assertEqual(compare["action"]["q"], "")

        no_docs = app.assistant_rule_based("покажи объекты без договоров по ОКВ", {"mode": "slice", "template": "all"})
        self.assertEqual(no_docs["action"]["template"], "okv")
        self.assertEqual(no_docs["action"]["post_filter"], "no_documents")
        self.assertEqual(no_docs["action"]["open_view"], "problems")

        unreviewed = app.assistant_rule_based("показать непроверенные проблемы СКК", {"mode": "slice", "template": "all"})
        self.assertEqual(unreviewed["action"]["post_filter"], "unreviewed")
        self.assertEqual(unreviewed["action"]["open_view"], "problems")

        control = app.assistant_rule_based("покажи контроль загрузки", {"mode": "slice", "template": "all"})
        self.assertEqual(control["action"]["open"], "control")
        self.assertEqual(control["action"]["q"], "")

        export = app.assistant_rule_based("скачать excel по СКК", {"mode": "slice", "template": "all"})
        self.assertEqual(export["intent"], "export_excel")
        self.assertEqual(export["action"]["download"], "excel")
        self.assertEqual(export["action"]["q"], "")

    def test_assistant_message_overrides_clean_bad_llm_like_action(self):
        fallback = app.assistant_rule_based("сравни СКК февраль 2025 и апрель 2026", {})["action"]
        dirty = {
            "mode": "slice",
            "template": "all",
            "date": "2026-04-01",
            "q": "сравни СКК февраль 2025 и апрель 2026",
            "metrics": ["limit"],
            "open_view": "overview",
        }
        action = app.validate_assistant_action(app.apply_message_overrides("сравни СКК февраль 2025 и апрель 2026", dirty), fallback)
        self.assertEqual(action["mode"], "compare")
        self.assertEqual(action["template"], "skk")
        self.assertEqual(action["base"], "2025-02-01")
        self.assertEqual(action["target"], "2026-04-01")
        self.assertEqual(action["q"], "")
        self.assertEqual(action["open_view"], "changes")

    def test_assistant_message_overrides_fix_object_date_query(self):
        message = "скк благовещенск март 2025"
        fallback = app.assistant_rule_based(message, {"mode": "slice", "template": "all", "date": "2026-04-01"})["action"]
        dirty = {
            "mode": "compare",
            "template": "all",
            "date": "2026-04-01",
            "q": "благовещенск март 2025",
            "open_view": "changes",
        }
        action = app.validate_assistant_action(app.apply_message_overrides(message, dirty), fallback)
        intent = app.normalize_assistant_intent("run_compare", message, action, "run_query")
        self.assertEqual(intent, "find_object")
        self.assertEqual(action["mode"], "slice")
        self.assertEqual(action["template"], "skk")
        self.assertEqual(action["date"], "2025-03-01")
        self.assertEqual(action["q"], "благовещенск")
        self.assertEqual(action["open_view"], "overview")

    def test_malformed_llm_response_falls_back_rule_based(self):
        old_key = os.environ.get("GROQ_API_KEY")
        original = app.assistant_llm
        os.environ["GROQ_API_KEY"] = "test-key"
        app.assistant_llm = lambda *_args, **_kwargs: (_ for _ in ()).throw(ValueError("bad llm"))
        try:
            payload = app.assistant_response("Покажи СКК", {})
            self.assertEqual(payload["mode"], "rule_based")
            self.assertEqual(payload["action"]["template"], "skk")
        finally:
            app.assistant_llm = original
            if old_key is None:
                os.environ.pop("GROQ_API_KEY", None)
            else:
                os.environ["GROQ_API_KEY"] = old_key

    def test_rag_loader_reads_markdown_documents(self):
        documents = app.load_rag_documents()
        self.assertGreaterEqual(len(documents), 6)
        self.assertTrue(any("БО" in item["content"] for item in documents))


class HttpTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.server = ThreadingHTTPServer(("127.0.0.1", 0), app.Handler)
        cls.port = cls.server.server_address[1]
        cls.thread = threading.Thread(target=cls.server.serve_forever, daemon=True)
        cls.thread.start()

    @classmethod
    def tearDownClass(cls):
        cls.server.shutdown()
        cls.server.server_close()
        cls.thread.join(timeout=5)

    def request(self, path, method="GET", payload=None):
        conn = HTTPConnection("127.0.0.1", self.port, timeout=10)
        try:
            body = None
            headers = {}
            if payload is not None:
                body = json.dumps(payload).encode("utf-8")
                headers["Content-Type"] = "application/json"
            conn.request(method, path, body=body, headers=headers)
            response = conn.getresponse()
            body = response.read()
            return response.status, response.getheader("Content-Type"), body
        finally:
            conn.close()

    def request_multipart(self, path, fields, file_field=None):
        boundary = "----codex-test-boundary"
        parts = []
        for name, value in fields.items():
            parts.append(
                f"--{boundary}\r\n"
                f'Content-Disposition: form-data; name="{name}"\r\n\r\n'
                f"{value}\r\n".encode("utf-8")
            )
        if file_field:
            name, filename, content_type, content = file_field
            parts.append(
                f"--{boundary}\r\n"
                f'Content-Disposition: form-data; name="{name}"; filename="{filename}"\r\n'
                f"Content-Type: {content_type}\r\n\r\n".encode("utf-8")
                + content
                + b"\r\n"
            )
        parts.append(f"--{boundary}--\r\n".encode("utf-8"))
        body = b"".join(parts)
        conn = HTTPConnection("127.0.0.1", self.port, timeout=20)
        try:
            conn.request(
                "POST",
                path,
                body=body,
                headers={"Content-Type": f"multipart/form-data; boundary={boundary}", "Content-Length": str(len(body))},
            )
            response = conn.getresponse()
            return response.status, response.getheader("Content-Type"), response.read()
        finally:
            conn.close()

    def test_meta_endpoint_returns_json(self):
        status, content_type, body = self.request("/api/meta")
        self.assertEqual(status, 200)
        self.assertIn("application/json", content_type)
        payload = json.loads(body.decode("utf-8"))
        self.assertEqual(payload["records"], 4037)
        self.assertIn("РЧБ", payload["sources"])

    def test_query_endpoint_returns_filtered_result(self):
        status, content_type, body = self.request("/api/query?code=6105&source=%D0%A0%D0%A7%D0%91")
        self.assertEqual(status, 200)
        self.assertIn("application/json", content_type)
        payload = json.loads(body.decode("utf-8"))
        self.assertGreater(payload["count"], 0)
        self.assertTrue(payload["rows"])
        self.assertTrue(all("РЧБ" in row["sources"] for row in payload["rows"]))
        self.assertTrue(all(isinstance(value, (int, float)) for value in payload["totals"].values()))
        self.assertTrue(all(not isinstance(value, str) for value in payload["totals"].values()))

    def test_static_index_is_served(self):
        status, content_type, body = self.request("/")
        self.assertEqual(status, 200)
        self.assertIn("text/html", content_type)
        self.assertIn("Простая аналитика расходов", body.decode("utf-8"))


    def test_catalog_quality_trace_and_compare_endpoints_return_json(self):
        for path in (
            "/api/catalog/templates",
            "/api/catalog/metrics",
            "/api/catalog/quick-actions",
            "/api/catalog/dates",
            "/api/catalog/sources",
            "/api/catalog/budgets",
            "/api/catalog/objects?q=6105",
            "/api/quality",
            "/api/readiness?view=as_of&date=2026-04-01&template=skk",
            "/api/compare?base=2025-02-01&target=2026-04-01&template=skk&metrics=limit,cash",
        ):
            status, content_type, body = self.request(path)
            self.assertEqual(status, 200, path)
            self.assertIn("application/json", content_type)
            self.assertIsNotNone(json.loads(body.decode("utf-8")))

        record_id = app.STORE.records[0]["id"]
        status, content_type, body = self.request(f"/api/trace?id={record_id}")
        self.assertEqual(status, 200)
        payload = json.loads(body.decode("utf-8"))
        self.assertEqual(payload["id"], record_id)
        self.assertIn("raw", payload)
        self.assertIn("human_summary", payload)

    def test_control_endpoint_returns_load_summary(self):
        status, content_type, body = self.request("/api/control?date=2026-04-01&template=skk")
        self.assertEqual(status, 200)
        self.assertIn("application/json", content_type)
        payload = json.loads(body.decode("utf-8"))
        self.assertGreater(payload["summary"]["records"], 0)
        self.assertGreater(payload["summary"]["sources"], 0)
        self.assertIn("sources", payload)
        self.assertIn("files", payload)
        self.assertIn("object_linkage", payload)
        self.assertIn("issues", payload)
        self.assertIn("by_name", payload["object_linkage"])

    def test_review_post_persists_status(self):
        original_path = app.REVIEWS_PATH
        with tempfile.TemporaryDirectory() as temp_dir:
            app.REVIEWS_PATH = Path(temp_dir) / "reviews.json"
            try:
                object_key = app.query_as_of({"date": ["2026-04-01"], "template": ["skk"]})["rows"][0]["object_key"]
                status, content_type, body = self.request(
                    "/api/review",
                    method="POST",
                    payload={"object_key": object_key, "status": "in_progress", "assignee": "Анна", "comment": "Проверить платежи"},
                )
                self.assertEqual(status, 200)
                self.assertIn("application/json", content_type)
                payload = json.loads(body.decode("utf-8"))
                self.assertEqual(payload["status"], "in_progress")
                self.assertEqual(payload["label"], "В работе")
                self.assertTrue(app.REVIEWS_PATH.exists())
                self.assertEqual(app.review_for_object(object_key)["assignee"], "Анна")
            finally:
                app.REVIEWS_PATH = original_path

    def test_review_invalid_status_returns_400(self):
        original_path = app.REVIEWS_PATH
        with tempfile.TemporaryDirectory() as temp_dir:
            app.REVIEWS_PATH = Path(temp_dir) / "reviews.json"
            try:
                status, _, body = self.request(
                    "/api/review",
                    method="POST",
                    payload={"object_key": "x", "status": "bad"},
                )
                self.assertEqual(status, 400)
                self.assertEqual(json.loads(body.decode("utf-8"))["error"], "invalid_status")
            finally:
                app.REVIEWS_PATH = original_path

    def test_import_rejects_wrong_extension(self):
        status, _, body = self.request_multipart(
            "/api/import",
            {"source_type": "rcb"},
            ("file", "bad.txt", "text/plain", b"not csv"),
        )
        self.assertEqual(status, 400)
        self.assertEqual(json.loads(body.decode("utf-8"))["error"], "invalid_extension")

    def test_import_requires_source_type(self):
        status, _, body = self.request_multipart(
            "/api/import",
            {},
            ("file", "data.csv", "text/csv", b"a,b\r\n1,2\r\n"),
        )
        self.assertEqual(status, 400)
        self.assertEqual(json.loads(body.decode("utf-8"))["error"], "source_type_required")

    def test_successful_import_updates_meta_records(self):
        original_uploads = app.DATA_UPLOADS_DIR
        original_store = app.STORE
        with tempfile.TemporaryDirectory() as temp_dir:
            app.DATA_UPLOADS_DIR = Path(temp_dir) / "uploads"
            try:
                before = app.STORE.meta["records"]
                content = (
                    "Отчет на 01.04.2026\r\n"
                    "Бюджет;КЦСР;Наименование КЦСР;КФСР;КВР;КОСГУ;Наименование КВСР;Наименование КВР;Лимиты ПБС;Подтв. лимитов по БО;Всего выбытий;Дата проводки\r\n"
                    "Тестовый бюджет;999996105;Импорт тест;01;244;226;Ведомство;Услуга;100,00;0;10,00;01.04.2026\r\n"
                ).encode("utf-8-sig")
                status, content_type, body = self.request_multipart(
                    "/api/import?date=2026-04-01&template=all",
                    {"source_type": "rcb"},
                    ("file", "import_april.csv", "text/csv", content),
                )
                self.assertEqual(status, 200)
                self.assertIn("application/json", content_type)
                payload = json.loads(body.decode("utf-8"))
                self.assertGreater(payload["summary"]["records"], 0)
                self.assertGreater(app.STORE.meta["records"], before)
                self.assertTrue((app.DATA_UPLOADS_DIR / "rcb" / "import_april.csv").exists())
            finally:
                app.DATA_UPLOADS_DIR = original_uploads
                app.STORE = original_store

    def test_object_and_excel_export_endpoints(self):
        query = app.query_as_of({"date": ["2026-04-01"], "template": ["skk"]})
        object_key = query["rows"][0]["object_key"]
        status, content_type, body = self.request(f"/api/object?date=2026-04-01&template=skk&object_key={quote(object_key)}")
        self.assertEqual(status, 200)
        self.assertIn("application/json", content_type)
        payload = json.loads(body.decode("utf-8"))
        self.assertEqual(payload["object_key"], object_key)

        status, content_type, body = self.request("/api/export.xlsx?date=2026-04-01&template=skk")
        self.assertEqual(status, 200)
        self.assertIn("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", content_type)
        workbook = load_workbook(BytesIO(body), read_only=True)
        self.assertEqual(workbook.sheetnames, ["Выводы", "Итоги", "Объекты", "Проблемы", "Исходные строки", "Контроль загрузки", "Методика"])
        self.assertIn("Выводы", workbook.sheetnames)
        method_text = "\n".join(str(row[0] or "") for row in workbook["Методика"].iter_rows(values_only=True))
        self.assertIn("последний месячный срез", method_text)
        control_text = "\n".join(str(row[0] or "") for row in workbook["Контроль загрузки"].iter_rows(values_only=True))
        self.assertIn("Итоги загрузки", control_text)

        try:
            import reportlab  # noqa: F401
        except ImportError:
            return
        status, content_type, body = self.request("/api/export.pdf?date=2026-04-01&template=skk")
        self.assertEqual(status, 200)
        self.assertIn("application/pdf", content_type)
        self.assertTrue(body.startswith(b"%PDF"))
        self.assertGreater(len(body), 5000)

    def test_excel_export_has_formatting(self):
        content, _ = app.export_excel({"date": ["2026-04-01"], "template": ["skk"]})
        workbook = load_workbook(BytesIO(content))
        self.assertEqual(workbook.sheetnames, ["Выводы", "Итоги", "Объекты", "Проблемы", "Исходные строки", "Контроль загрузки", "Методика"])
        self.assertEqual(workbook["Выводы"].freeze_panes, "A4")
        self.assertEqual(workbook["Итоги"].freeze_panes, "A2")
        self.assertEqual(workbook["Контроль загрузки"].freeze_panes, "A2")
        self.assertTrue(workbook["Объекты"].auto_filter.ref)
        self.assertTrue(workbook["Итоги"]["A1"].font.bold)
        self.assertEqual(workbook["Итоги"]["A1"].font.color.rgb, "00FFFFFF")
        выводы = "\n".join(str(row[0] or "") for row in workbook["Выводы"].iter_rows(values_only=True))
        self.assertIn("Следующие действия", выводы)

    def test_excel_methodology_includes_risk_rules(self):
        content, _ = app.export_excel({"date": ["2026-04-01"], "template": ["skk"]})
        workbook = load_workbook(BytesIO(content), read_only=True)
        object_headers = [cell for cell in next(workbook["Объекты"].iter_rows(values_only=True))]
        self.assertIn("Факторы риска", object_headers)
        self.assertIn("Статус проверки", object_headers)
        self.assertIn("Ответственный", object_headers)
        self.assertIn("Комментарий", object_headers)
        method_rows = list(workbook["Методика"].iter_rows(values_only=True))
        method_text = "\n".join(" ".join(str(cell or "") for cell in row) for row in method_rows)
        self.assertIn(app.RISK_MODEL_VERSION, method_text)
        self.assertIn("Код Фактор Баллы", method_text)
        self.assertIn("no_cash", method_text)

    def test_pdf_export_endpoint_or_function_returns_pdf(self):
        try:
            import reportlab  # noqa: F401
        except ImportError:
            self.skipTest("reportlab is not installed")
        body, filename = app.export_pdf({"date": ["2026-04-01"], "template": ["skk"]})
        self.assertTrue(body.startswith(b"%PDF"))
        self.assertTrue(filename.endswith(".pdf"))
        self.assertGreater(len(body), 5000)

    def test_pdf_export_compare_returns_pdf(self):
        try:
            import reportlab  # noqa: F401
        except ImportError:
            self.skipTest("reportlab is not installed")
        body, filename = app.export_pdf({"mode": ["compare"], "base": ["2025-02-01"], "target": ["2026-04-01"], "template": ["skk"]})
        self.assertTrue(body.startswith(b"%PDF"))
        self.assertTrue(filename.endswith(".pdf"))
        self.assertGreater(len(body), 1000)

    def test_assistant_endpoint_returns_rule_based_json_without_groq(self):
        old_key = os.environ.pop("GROQ_API_KEY", None)
        self.addCleanup(lambda: os.environ.__setitem__("GROQ_API_KEY", old_key) if old_key is not None else os.environ.pop("GROQ_API_KEY", None))
        status, content_type, body = self.request(
            "/api/assistant",
            method="POST",
            payload={"message": "Покажи СКК", "context": {"mode": "slice", "template": "all"}},
        )
        self.assertEqual(status, 200)
        self.assertIn("application/json", content_type)
        payload = json.loads(body.decode("utf-8"))
        self.assertEqual(payload["mode"], "rule_based")
        self.assertEqual(payload["action"]["template"], "skk")

    def test_explain_endpoint_returns_rule_based_without_groq(self):
        old_key = os.environ.pop("GROQ_API_KEY", None)
        self.addCleanup(lambda: os.environ.__setitem__("GROQ_API_KEY", old_key) if old_key is not None else os.environ.pop("GROQ_API_KEY", None))
        status, content_type, body = self.request(
            "/api/explain",
            method="POST",
            payload={"kind": "query", "payload": {"attention_summary": {"bullets": ["Главная проблема - нет кассы."]}}},
        )
        self.assertEqual(status, 200)
        self.assertIn("application/json", content_type)
        payload = json.loads(body.decode("utf-8"))
        self.assertEqual(payload["mode"], "rule_based")
        self.assertTrue(payload["bullets"])

    def test_quick_actions_run_against_public_api(self):
        status, _, body = self.request("/api/catalog/quick-actions")
        self.assertEqual(status, 200)
        actions = json.loads(body.decode("utf-8"))
        self.assertGreaterEqual(len(actions), 6)
        for action in actions:
            metrics = ",".join(action.get("metrics", []))
            if action["mode"] == "compare":
                path = f"/api/compare?template={action['template']}&base=2025-02-01&target=2026-04-01&metrics={metrics}"
                status, content_type, body = self.request(path)
                self.assertEqual(status, 200, action["code"])
                payload = json.loads(body.decode("utf-8"))
                self.assertIn("rows", payload)
            else:
                post_filter = f"&post_filter={action['post_filter']}" if action.get("post_filter") else ""
                path = f"/api/query?template={action['template']}&start=2025-02-01&end=2026-04-01&metrics={metrics}{post_filter}"
                status, content_type, body = self.request(path)
                self.assertEqual(status, 200, action["code"])
                payload = json.loads(body.decode("utf-8"))
                self.assertIn("rows", payload)
                self.assertIn("totals", payload)


class StaticFilesTests(unittest.TestCase):
    def test_gitignore_mentions_env(self):
        gitignore = (app.ROOT / ".gitignore").read_text(encoding="utf-8")
        self.assertIn(".env", gitignore)
        self.assertIn("!.env.example", gitignore)

    def test_frontend_assets_exist_and_reference_api(self):
        index = (app.STATIC_DIR / "index.html").read_text(encoding="utf-8")
        script = (app.STATIC_DIR / "app.js").read_text(encoding="utf-8")
        styles = (app.STATIC_DIR / "styles.css").read_text(encoding="utf-8")

        self.assertIn("/static/app.js", index)
        self.assertIn("/api/meta", script)
        self.assertIn("/api/query", script)
        self.assertIn("/api/compare", script)
        self.assertIn("/api/trace", script)
        self.assertIn("/api/readiness", script)
        self.assertIn("/api/object", script)
        self.assertIn("/api/review", script)
        self.assertIn("/api/import", script)
        self.assertIn("/api/export.xlsx", script)
        self.assertIn("/api/catalog/quick-actions", script)
        self.assertIn("/api/assistant", script)
        self.assertIn("/api/explain", script)
        self.assertIn("grid-template-columns", styles)

    def test_local_vue_vendor_exists(self):
        vendor = app.STATIC_DIR / "vendor" / "vue.global.prod.js"
        self.assertTrue(vendor.exists())
        self.assertIn("Vue", vendor.read_text(encoding="utf-8"))

    def test_index_uses_local_vue_runtime(self):
        index = (app.STATIC_DIR / "index.html").read_text(encoding="utf-8")
        self.assertNotIn("unpkg.com/vue", index)
        self.assertIn("/static/vendor/vue.global.prod.js", index)

    def test_vue_frontend_is_declarative(self):
        index = (app.STATIC_DIR / "index.html").read_text(encoding="utf-8")
        script = (app.STATIC_DIR / "app.js").read_text(encoding="utf-8")

        self.assertIn('id="app"', index)
        self.assertIn("vue.global.prod.js", index)
        self.assertIn("/static/app.js", index)
        for marker in (
            "createApp",
            'mount("#app")',
            "loadInitialData",
            "loadData",
            "loadCompare",
            "openTrace",
            "exportCsv",
            "exportExcel",
            "openObject",
            "saveReview",
            "quickActions",
            "command",
            "applyQuickAction",
            "buildCommandSuggestions",
            "runCommand",
            "explainResult",
            "resultNarrative",
            "simpleRows",
            "problemRows",
        ):
            self.assertIn(marker, script)
        for marker in (
            "Напишите, что нужно получить",
            "Короткий вывод",
            "Объекты",
            "Проблемы",
        ):
            self.assertIn(marker, index)
        self.assertNotIn("Спросить помощника", index)
        for legacy_marker in (
            "document.querySelector",
            "innerHTML =",
            "addEventListener",
        ):
            self.assertNotIn(legacy_marker, script)


if __name__ == "__main__":
    unittest.main()
