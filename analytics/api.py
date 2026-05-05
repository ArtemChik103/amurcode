"""Локальный backend аналитики расходов.

Файл совмещает загрузку исходных CSV, нормализацию записей, расчет витрин,
HTTP API и экспорт Excel. Данные держатся в памяти, поэтому функции ниже
стараются возвращать готовые для UI структуры без отдельного слоя БД.
"""

from __future__ import annotations

import csv
import cgi
import json
import os
import re
import sys
from io import BytesIO
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse
from urllib.error import HTTPError
from urllib.request import Request as UrlRequest, urlopen

import requests
from fastapi import FastAPI, File, Form, Request, UploadFile
from fastapi.responses import FileResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles


ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "case"
DATA_RUNTIME_DIR = ROOT / "data"
DATA_UPLOADS_DIR = DATA_RUNTIME_DIR / "uploads"
REVIEWS_PATH = DATA_RUNTIME_DIR / "reviews.json"
STATIC_DIR = ROOT / "static"
RAG_DIR = ROOT / "docs" / "rag"
QUALITY_ISSUES: list[dict] = []
LOAD_STATS: dict[str, dict[str, int]] = {}
MONEY_ZERO = Decimal("0.00")
IMPORT_SOURCE_TYPES = {
    "rcb": "РЧБ",
    "agreements": "Соглашения",
    "state_task_contracts": "ГЗ: контракты",
    "state_task_payments": "ГЗ: платежи",
    "buau": "БУАУ",
}
IMPORT_MAX_BYTES = 50 * 1024 * 1024


def load_local_env(path: Path = ROOT / ".env") -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


load_local_env()

app = FastAPI(title="Expense Analytics")
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

# Метрики хранятся в едином порядке, чтобы backend, UI и Excel одинаково
# трактовали выбранные суммы.
METRIC_KEYS = ["limit", "obligation", "cash", "agreement", "contract", "payment", "buau"]
METRICS = {
    "limit": "Лимиты",
    "obligation": "БО",
    "cash": "Касса",
    "agreement": "Соглашения",
    "contract": "Контракты",
    "payment": "Платежи",
    "buau": "БУ/АУ",
}
TEMPLATES = {
    "all": {"label": "Все данные", "description": "Без предметного шаблона"},
    "kik": {"label": "КИК", "description": "КЦСР содержит 975 или 978 с 6-й позиции"},
    "skk": {"label": "СКК", "description": "КЦСР содержит 6105 с 6-й позиции"},
    "two_thirds": {"label": "2/3", "description": "КЦСР содержит 970 с 6-й позиции"},
    "okv": {"label": "ОКВ", "description": "Капитальные вложения по КВР"},
}

# Быстрые действия являются контрактом между API и первым экраном UI:
# каждое действие разворачивается в обычные query/compare параметры.
QUICK_ACTIONS = {
    "demo_60s": {
        "label": "Демо за 60 секунд",
        "description": "Показать проблемные СКК, главный риск и отчет",
        "mode": "slice",
        "template": "skk",
        "post_filter": "execution_problems",
        "open_view": "problems",
        "highlight": "top_risks",
        "demo_mode": "skk_risks",
        "metrics": ["limit", "obligation", "cash", "agreement", "contract", "payment", "buau"],
    },
    "show_skk": {
        "label": "Собрать отчет СКК",
        "description": "Готовая выборка по СКК на выбранную дату",
        "mode": "slice",
        "template": "skk",
        "metrics": ["limit", "obligation", "cash", "agreement", "contract", "payment", "buau"],
    },
    "show_kik": {
        "label": "Собрать отчет КИК",
        "description": "Готовая выборка по КИК на выбранную дату",
        "mode": "slice",
        "template": "kik",
        "metrics": ["limit", "obligation", "cash", "agreement", "contract", "payment"],
    },
    "show_two_thirds": {
        "label": "Собрать отчет 2/3",
        "description": "Готовая выборка по высвобождаемым средствам",
        "mode": "slice",
        "template": "two_thirds",
        "metrics": ["limit", "obligation", "cash", "agreement"],
    },
    "show_okv": {
        "label": "Собрать отчет ОКВ",
        "description": "Готовая выборка по капитальным вложениям",
        "mode": "slice",
        "template": "okv",
        "metrics": ["limit", "obligation", "cash", "contract", "payment", "buau"],
    },
    "compare_skk": {
        "label": "Сравнить две даты",
        "description": "Показать изменения между отчетными датами",
        "mode": "compare",
        "template": "skk",
        "metrics": ["limit", "obligation", "cash"],
    },
    "execution_problems": {
        "label": "Найти проблемные объекты",
        "description": "Нет документов, оплат, кассы или есть разрывы данных",
        "mode": "slice",
        "template": "all",
        "metrics": ["limit", "cash", "payment", "buau"],
        "post_filter": "execution_problems",
    },
    "demo_skk_problems": {
        "label": "Проблемные СКК",
        "description": "Показать проблемные объекты СКК и источник цифр",
        "mode": "slice",
        "template": "skk",
        "post_filter": "execution_problems",
        "highlight": "top_risks",
        "open_view": "problems",
        "metrics": ["limit", "obligation", "cash", "agreement", "contract", "payment", "buau"],
    },
    "find_object": {
        "label": "Найти объект",
        "description": "Поиск по названию, коду, бюджету или документу",
        "mode": "slice",
        "template": "all",
        "metrics": ["limit", "obligation", "cash", "agreement", "contract", "payment", "buau"],
    },
}
ASSISTANT_INTENTS = {
    "run_query",
    "run_compare",
    "show_execution_problems",
    "find_object",
    "open_object",
    "export_excel",
    "explain_metric",
    "explain_template",
    "explain_result",
    "help",
}
ASSISTANT_POST_FILTERS = {
    "",
    "execution_problems",
    "no_documents",
    "no_payments",
    "no_cash",
    "low_cash",
    "data_gap",
    "unreviewed",
}
ASSISTANT_OPEN_VIEWS = {"overview", "objects", "problems", "records", "changes"}
ASSISTANT_ACTION_FIELDS = {
    "mode",
    "template",
    "date",
    "base",
    "target",
    "open",
    "download",
    "q",
    "code",
    "budget",
    "source",
    "post_filter",
    "metrics",
    "open_view",
    "reset_scope",
}
ASSISTANT_FOLLOWUP_FIELDS = {"label", "action"}
ASSISTANT_FOLLOWUP_ACTION_FIELDS = {"open", "download", "post_filter", "open_view"}
ASSISTANT_OPEN_ACTIONS = {"", "top_risk", "control"}
ASSISTANT_DOWNLOAD_ACTIONS = {"", "excel", "pdf", "csv"}
CAPITAL_KVR = {"400", "410", "411", "412", "413", "414", "415", "416", "417"}

SNAPSHOT_RE = re.compile(r"на\s+(\d{2}\.\d{2}\.\d{4})")
MONTHS = {
    "январь": 1,
    "января": 1,
    "февраль": 2,
    "февраля": 2,
    "март": 3,
    "марта": 3,
    "апрель": 4,
    "апреля": 4,
    "май": 5,
    "мая": 5,
    "июнь": 6,
    "июня": 6,
    "июль": 7,
    "июля": 7,
    "август": 8,
    "августа": 8,
    "сентябрь": 9,
    "сентября": 9,
    "октябрь": 10,
    "октября": 10,
    "ноябрь": 11,
    "ноября": 11,
    "декабрь": 12,
    "декабря": 12,
}


def relative_source(path: Path | str) -> str:
    try:
        return str(Path(path).resolve().relative_to(ROOT)).replace("\\", "/")
    except ValueError:
        return str(path).replace("\\", "/")


def uploaded_paths(source_type: str) -> list[Path]:
    folder = DATA_UPLOADS_DIR / source_type
    if not folder.exists():
        return []
    return sorted([*folder.glob("*.csv"), *folder.glob("*.xlsx")])


def csv_dict_rows(path: Path, delimiter: str = ",") -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle, delimiter=delimiter))


def xlsx_dict_rows(path: Path) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    result = []
    for values in rows[1:]:
        if not any(value not in (None, "") for value in values):
            continue
        result.append({header: "" if value is None else str(value) for header, value in zip(headers, values) if header})
    return result


def dict_rows(path: Path, delimiter: str = ",") -> list[dict[str, str]]:
    if path.suffix.lower() == ".xlsx":
        return xlsx_dict_rows(path)
    return csv_dict_rows(path, delimiter)


def csv_raw_rows(path: Path, delimiter: str = ";") -> list[list[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.reader(handle, delimiter=delimiter))


def xlsx_raw_rows(path: Path) -> list[list[str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        return [["" if value is None else str(value) for value in row] for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def raw_rows(path: Path, delimiter: str = ";") -> list[list[str]]:
    if path.suffix.lower() == ".xlsx":
        return xlsx_raw_rows(path)
    return csv_raw_rows(path, delimiter)


def sanitize_upload_filename(filename: str) -> str:
    name = Path(filename or "upload").name
    return re.sub(r"[^0-9A-Za-zА-Яа-яЁё._-]+", "_", name).strip("._") or "upload"


def unique_upload_path(source_type: str, filename: str) -> Path:
    folder = DATA_UPLOADS_DIR / source_type
    folder.mkdir(parents=True, exist_ok=True)
    safe_name = sanitize_upload_filename(filename)
    target = folder / safe_name
    if target.exists():
        stamp = datetime.now().strftime("%Y%m%d%H%M%S")
        target = folder / f"{target.stem}_{stamp}{target.suffix}"
    return target


def import_payload(params: dict[str, list[str]] | None = None) -> dict:
    return control_summary(params or {})


def save_import_upload(source_type: str, filename: str, content: bytes) -> Path:
    target = unique_upload_path(source_type, filename)
    target.write_bytes(content)
    return target


def add_quality_issue(
    source_file: str,
    source_row: int | str,
    severity: str,
    code: str,
    message: str,
    field: str = "",
    value: object = "",
) -> None:
    QUALITY_ISSUES.append(
        {
            "source_file": source_file,
            "source_row": source_row,
            "severity": severity,
            "code": code,
            "message": message,
            "field": field,
            "value": "" if value is None else str(value),
        }
    )


REVIEW_STATUSES = {
    "new": "Новый",
    "in_progress": "В работе",
    "checked": "Проверен",
    "not_issue": "Не проблема",
}


def load_reviews() -> dict:
    if not REVIEWS_PATH.exists():
        return {}
    try:
        payload = json.loads(REVIEWS_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        add_quality_issue(
            relative_source(REVIEWS_PATH),
            "",
            "warn",
            "reviews_parse_failed",
            "Не удалось прочитать локальные статусы проверки.",
        )
        return {}
    return payload if isinstance(payload, dict) else {}


def save_reviews(payload: dict) -> None:
    REVIEWS_PATH.parent.mkdir(parents=True, exist_ok=True)
    temp_path = REVIEWS_PATH.with_name(f"{REVIEWS_PATH.name}.tmp")
    temp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_path.replace(REVIEWS_PATH)


def default_review() -> dict:
    return {
        "status": "new",
        "label": REVIEW_STATUSES["new"],
        "assignee": "",
        "comment": "",
        "updated_at": "",
        "updated_by": "",
    }


def review_for_object(object_key: str) -> dict:
    saved = load_reviews().get(object_key, {})
    review = default_review()
    if isinstance(saved, dict):
        review.update({key: value for key, value in saved.items() if key in review or key in {"updated_by"}})
    if review.get("status") not in REVIEW_STATUSES:
        review["status"] = "new"
    review["label"] = REVIEW_STATUSES[review["status"]]
    return review


def update_review(object_key: str, payload: dict) -> dict:
    if not object_key:
        return {"error": "object_key_required"}
    status = str(payload.get("status") or "new").strip()
    if status not in REVIEW_STATUSES:
        return {"error": "invalid_status"}
    reviews = load_reviews()
    review = {
        "status": status,
        "assignee": str(payload.get("assignee") or "")[:120],
        "comment": str(payload.get("comment") or "")[:1000],
        "updated_at": datetime.now().replace(microsecond=0).isoformat(),
        "updated_by": "local",
    }
    reviews[object_key] = review
    save_reviews(reviews)
    return review_for_object(object_key)


def reviews_payload() -> dict:
    return {key: review_for_object(key) for key in load_reviews()}


def parse_money(value: object, source_file: str = "", source_row: int | str = "", field: str = "") -> Decimal:
    """Разбирает суммы из русских CSV и машинных выгрузок в Decimal."""
    if value is None:
        return MONEY_ZERO
    if isinstance(value, Decimal):
        return value
    text = str(value).strip().replace("\xa0", " ")
    if not text:
        return MONEY_ZERO
    text = text.replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        if source_file or field:
            add_quality_issue(
                source_file,
                source_row,
                "warning",
                "amount_parse_failed",
                "Не удалось распарсить сумму",
                field,
                value,
            )
        return MONEY_ZERO


def money_to_json(value: Decimal) -> float:
    return float(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def money_sum(values) -> Decimal:
    total = MONEY_ZERO
    for value in values:
        total += parse_money(value)
    return total


def json_safe(value: object) -> object:
    if isinstance(value, Decimal):
        return money_to_json(value)
    if isinstance(value, dict):
        return {key: json_safe(item) for key, item in value.items()}
    if isinstance(value, list):
        return [json_safe(item) for item in value]
    if isinstance(value, tuple):
        return [json_safe(item) for item in value]
    return value


def request_params(request: Request) -> dict[str, list[str]]:
    params: dict[str, list[str]] = {}
    for key, value in request.query_params.multi_items():
        params.setdefault(key, []).append(value)
    return params


def fastapi_json(payload: object, status_code: int = 200) -> JSONResponse:
    return JSONResponse(content=json_safe(payload), status_code=status_code)


def fastapi_binary(body: bytes, content_type: str, filename: str) -> Response:
    return Response(
        content=body,
        media_type=content_type,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-store",
        },
    )


def parse_amount(value: object, source_file: str = "", source_row: int | str = "", field: str = "") -> float:
    """Compatibility wrapper: старый API возвращает float."""
    return float(parse_money(value, source_file, source_row, field))


def parse_date(value: object, source_file: str = "", source_row: int | str = "", field: str = "") -> str:
    """Нормализует известные форматы дат к ISO-строке YYYY-MM-DD."""
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    for fmt in ("%d.%m.%Y", "%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    if source_file or field:
        add_quality_issue(
            source_file,
            source_row,
            "warning",
            "date_parse_failed",
            "Не удалось распарсить дату",
            field,
            value,
        )
    return ""


def normalize_code(value: object) -> str:
    return re.sub(r"[^0-9A-Za-zА-Яа-я]", "", str(value or "")).upper()


def display_code(value: object) -> str:
    return str(value or "").strip()


def find_header_value(row: dict[str, str], *prefixes: str) -> str:
    for prefix in prefixes:
        for key, value in row.items():
            if str(key).strip().lower().startswith(prefix.lower()):
                return value
    return ""


def normalize_name(value: object) -> str:
    return " ".join(re.findall(r"[0-9A-Za-zА-Яа-яЁё]+", str(value or "").lower()))


def object_group_key(record: dict) -> str:
    """Строит стабильный ключ объекта для склейки строк из разных источников."""
    budget_norm = normalize_name(record.get("budget"))
    code_norm = normalize_code(record.get("object_code_norm") or record.get("object_code"))
    if code_norm:
        return f"{code_norm}|{budget_norm}"
    return f"name:{normalize_name(record.get('object_name'))}|{budget_norm}"


def make_record(records: list[dict], source_file: str, source_row: int, raw: dict, **fields: object) -> dict:
    record = {
        "id": f"r{len(records) + 1}",
        "source_file": source_file,
        "source_row": source_row,
        "raw": dict(raw),
    }
    record.update(fields)
    for metric in METRIC_KEYS:
        record[metric] = parse_money(record.get(metric))
    return record


def selected_metrics(params: dict[str, list[str]] | None = None) -> list[str]:
    if not params:
        return list(METRIC_KEYS)
    raw = params.get("metrics", [""])[0].strip()
    if not raw:
        return list(METRIC_KEYS)
    result = [metric for metric in raw.split(",") if metric in METRIC_KEYS]
    return result or list(METRIC_KEYS)


def default_date_range() -> tuple[str, str]:
    snapshots = STORE.meta.get("reporting_dates", []) if "STORE" in globals() else []
    if not snapshots:
        return "", ""
    return snapshots[0], snapshots[-1]


def quick_actions_payload() -> list[dict]:
    return [{"code": code, **action} for code, action in QUICK_ACTIONS.items()]


def matches_template(record: dict, template: str) -> bool:
    """Проверяет предметные шаблоны по фрагментам КЦСР или КВР."""
    if not template or template == "all":
        return True
    code = record.get("object_code_norm", "")
    if template == "kik":
        return code[5:8] in {"975", "978"}
    if template == "skk":
        return code[5:9] == "6105"
    if template == "two_thirds":
        return code[5:8] == "970"
    if template == "okv":
        return normalize_code(record.get("kvr")) in CAPITAL_KVR
    return True


def rcb_snapshot_from_rows(rows: list[list[str]], fallback_name: str) -> str:
    for row in rows[:8]:
        joined = " ".join(row)
        match = SNAPSHOT_RE.search(joined)
        if match:
            return parse_date(match.group(1))
    lower = fallback_name.lower()
    for name, month in MONTHS.items():
        if name in lower:
            year_match = re.search(r"(20\d{2})", lower)
            year = int(year_match.group(1)) if year_match else 2025
            next_month = month + 1
            next_year = year
            if next_month == 13:
                next_month = 1
                next_year += 1
            return f"{next_year:04d}-{next_month:02d}-01"
    return ""


def agreement_snapshot(row: dict[str, str], filename: str) -> str:
    period = row.get("period_of_date", "")
    if " - " in period:
        return parse_date(period.split(" - ")[-1])
    digits = re.search(r"(\d{2})(\d{2})(20\d{2})", filename)
    if digits:
        day, month, year = digits.groups()
        return f"{year}-{month}-{day}"
    return ""


def buau_snapshot(filename: str) -> str:
    lower = filename.lower()
    year_match = re.search(r"(20\d{2})", lower)
    year = int(year_match.group(1)) if year_match else 2025
    for name, month in MONTHS.items():
        if name in lower:
            return f"{year:04d}-{month:02d}-01"
    return ""


@dataclass
class DataStore:
    records: list[dict]
    meta: dict


def load_data() -> DataStore:
    """Загружает все источники и собирает справочники для UI."""
    QUALITY_ISSUES.clear()
    LOAD_STATS.clear()
    records: list[dict] = []
    load_rcb(records)
    load_agreements(records)
    load_state_task(records)
    load_buau(records)
    enrich_records(records)

    budgets = sorted({r["budget"] for r in records if r.get("budget")})
    sources = sorted({r["source"] for r in records})
    snapshots = sorted({r["snapshot"] for r in records if r.get("snapshot")})
    reporting_dates = sorted(
        {
            r["snapshot"]
            for r in records
            if r.get("snapshot") and r.get("record_kind") in {"rcb_snapshot", "agreement_snapshot"}
        }
    )
    objects = sorted(
        {r["object_name"] for r in records if r.get("object_name")},
        key=lambda x: x.lower(),
    )
    meta = {
        "records": len(records),
        "budgets": budgets,
        "sources": sources,
        "snapshots": snapshots,
        "reporting_dates": reporting_dates,
        "objects": objects[:500],
        "load_stats": LOAD_STATS,
        "quality": quality_summary(),
    }
    return DataStore(records=records, meta=meta)


def enrich_records(records: list[dict]) -> None:
    """Дополняет записи техническими полями, нужными для trace и агрегации."""
    source_counts: dict[str, int] = defaultdict(int)
    source_folders = {
        "РЧБ": "case/1_RCB",
        "Соглашения": "case/2_Agreements",
        "ГЗ: контракты": "case/3_StateTask",
        "ГЗ: платежи": "case/3_StateTask",
        "БУАУ": "case/4_BUAU_Export",
    }
    for index, record in enumerate(records, start=1):
        source = record.get("source", "")
        source_counts[source] += 1
        record.setdefault("id", f"r{index}")
        record.setdefault("source_file", source_folders.get(source, ""))
        record.setdefault("source_row", source_counts[source])
        record.setdefault("record_kind", "event")
        record.setdefault("document_id", record.get("document_number", ""))
        record.setdefault("document_date", record.get("event_date", ""))
        record.setdefault("snapshot_source", "event")
        record.setdefault("amount_semantics", "event_amount")
        record.setdefault("raw", {key: value for key, value in record.items() if key not in {"raw"}})
        record["object_key"] = object_group_key(record)
    for source, count in source_counts.items():
        LOAD_STATS[source or "unknown"] = {"read_rows": count, "records": count, "warnings": 0, "errors": 0}


def quality_summary() -> dict[str, int]:
    return {
        "warnings": sum(1 for issue in QUALITY_ISSUES if issue["severity"] == "warning"),
        "errors": sum(1 for issue in QUALITY_ISSUES if issue["severity"] == "error"),
    }


def control_records(params: dict[str, list[str]] | None = None) -> list[dict]:
    """Возвращает записи для контрольных итогов в семантике текущей выборки."""
    params = params or {}
    meaningful = {"date", "template", "q", "code", "budget", "source", "start", "end", "view", "post_filter"}
    has_filters = any(params.get(key, [""])[0].strip() for key in meaningful if key in params)
    if not has_filters:
        return list(STORE.records)
    if params.get("date", [""])[0].strip() or params.get("view", [""])[0].strip() == "as_of":
        return select_as_of(STORE.records, params.get("date", [""])[0].strip(), params)
    return apply_filters(STORE.records, params)


def control_summary(params: dict[str, list[str]] | None = None) -> dict:
    """Собирает контроль загрузки: источники, файлы, связку объектов и предупреждения."""
    records = control_records(params)
    source_groups: dict[str, dict] = {}
    file_groups: dict[str, dict] = {}
    object_sources: dict[str, set[str]] = defaultdict(set)

    for record in records:
        source = record.get("source") or "unknown"
        source_file = record.get("source_file") or ""
        key = record.get("object_key") or object_group_key(record)
        object_sources[key].add(source)

        amount = money_sum(record.get(metric) for metric in METRIC_KEYS)
        source_item = source_groups.setdefault(
            source,
            {"source": source, "read_rows": 0, "records": 0, "warnings": 0, "errors": 0, "total_amount": MONEY_ZERO, "files": set()},
        )
        source_item["records"] += 1
        source_item["read_rows"] += 1
        source_item["total_amount"] += amount
        if source_file:
            source_item["files"].add(source_file)

        file_item = file_groups.setdefault(
            source_file or "unknown",
            {"source_file": source_file or "unknown", "read_rows": 0, "records": 0, "warnings": 0, "errors": 0},
        )
        file_item["records"] += 1
        file_item["read_rows"] += 1

    for issue in QUALITY_ISSUES:
        source_file = issue.get("source_file") or "unknown"
        if source_file not in file_groups and not params:
            file_groups[source_file] = {"source_file": source_file, "read_rows": 0, "records": 0, "warnings": 0, "errors": 0}
        file_item = file_groups.get(source_file)
        if file_item is not None:
            if issue.get("severity") == "error":
                file_item["errors"] += 1
            else:
                file_item["warnings"] += 1
        for source_item in source_groups.values():
            if source_file in source_item["files"]:
                if issue.get("severity") == "error":
                    source_item["errors"] += 1
                else:
                    source_item["warnings"] += 1

    load_files = []
    for key, stats in LOAD_STATS.items():
        if "/" not in key and "\\" not in key:
            continue
        item = dict(file_groups.get(key, {"source_file": key, "read_rows": 0, "records": 0, "warnings": 0, "errors": 0}))
        if not params:
            item.update({field: stats.get(field, item.get(field, 0)) for field in ("read_rows", "records", "warnings", "errors")})
        load_files.append(item)
    for key, item in file_groups.items():
        if key != "unknown" and all(existing["source_file"] != key for existing in load_files):
            load_files.append(item)

    sources = []
    for item in source_groups.values():
        files = sorted(item.pop("files"))
        item["files"] = files
        item["total_amount"] = item["total_amount"].quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        sources.append(item)
    sources.sort(key=lambda item: item["source"])
    load_files.sort(key=lambda item: item["source_file"])

    by_name = sum(1 for key in object_sources if str(key).startswith("name:"))
    with_code = len(object_sources) - by_name
    single_source = sum(1 for sources_for_object in object_sources.values() if len(sources_for_object) == 1)
    multi_source = sum(1 for sources_for_object in object_sources.values() if len(sources_for_object) > 1)
    warnings = quality_summary()["warnings"]
    errors = quality_summary()["errors"]
    return {
        "summary": {
            "records": len(records),
            "sources": len(source_groups),
            "warnings": warnings,
            "errors": errors,
            "unmatched_name_keys": by_name,
            "single_source_objects": single_source,
        },
        "sources": sources,
        "files": load_files,
        "object_linkage": {
            "with_code": with_code,
            "by_name": by_name,
            "single_source": single_source,
            "multi_source": multi_source,
        },
        "issues": list(QUALITY_ISSUES),
    }


def start_load_stats(path: Path) -> dict[str, int]:
    key = relative_source(path)
    stats = {"read_rows": 0, "records": 0, "warnings": 0, "errors": 0}
    LOAD_STATS[key] = stats
    return stats


def finish_load_stats(stats: dict[str, int], issue_start: int) -> None:
    issues = QUALITY_ISSUES[issue_start:]
    stats["warnings"] = sum(1 for issue in issues if issue["severity"] == "warning")
    stats["errors"] = sum(1 for issue in issues if issue["severity"] == "error")


def load_rcb(records: list[dict]) -> None:
    """Читает РЧБ как месячные срезы лимитов, БО и кассового исполнения."""
    folder = DATA_DIR / "1_RCB"
    for path in sorted(folder.glob("*.csv")) + uploaded_paths("rcb"):
        rows = raw_rows(path, delimiter=";")
        header_index = next(
            (i for i, row in enumerate(rows) if row and row[0].strip() == "Бюджет"),
            None,
        )
        if header_index is None:
            continue
        snapshot = rcb_snapshot_from_rows(rows, path.name)
        header = rows[header_index]
        source_file = relative_source(path)
        for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            if not any(cell.strip() for cell in row):
                continue
            item = dict(zip(header, row))
            kcsr = display_code(item.get("КЦСР"))
            object_name = item.get("Наименование КЦСР", "").strip() or kcsr
            budget = item.get("Бюджет", "").strip()
            records.append(
                make_record(
                    records,
                    source_file,
                    row_number,
                    item,
                    source="РЧБ",
                    snapshot=snapshot,
                    event_date=parse_date(item.get("Дата проводки")),
                    record_kind="rcb_snapshot",
                    document_id="",
                    document_date="",
                    snapshot_source="monthly_snapshot",
                    amount_semantics="balance_as_of",
                    budget=budget,
                    object_code=kcsr,
                    object_code_norm=normalize_code(kcsr),
                    object_name=object_name,
                    kfsr=display_code(item.get("КФСР")),
                    kvr=display_code(item.get("КВР")),
                    kosgu=display_code(item.get("КОСГУ")),
                    counterparty=item.get("Наименование КВСР", "").strip(),
                    document_number="",
                    description=item.get("Наименование КВР", "").strip(),
                    limit=parse_money(find_header_value(item, "Лимиты ПБС"), source_file, row_number, "Лимиты ПБС"),
                    obligation=parse_money(find_header_value(item, "Подтв. лимитов по БО"), source_file, row_number, "Подтв. лимитов по БО"),
                    cash=parse_money(find_header_value(item, "Всего выбытий"), source_file, row_number, "Всего выбытий"),
                    agreement=MONEY_ZERO,
                    contract=MONEY_ZERO,
                    payment=MONEY_ZERO,
                    buau=MONEY_ZERO,
                )
            )


def load_agreements(records: list[dict]) -> None:
    """Читает соглашения как срезы на отчетные даты."""
    folder = DATA_DIR / "2_Agreements"
    class_names = {
        "273": "МБТ",
        "278": "Иные цели БУ/АУ",
        "272": "Госзадание",
        "313": "ЮЛ/ИП/ФЛ",
    }
    for path in sorted(folder.glob("*.csv")) + uploaded_paths("agreements"):
        source_file = relative_source(path)
        for row_number, row in enumerate(dict_rows(path), start=2):
                kcsr = display_code(row.get("kcsr_code"))
                recipient = (row.get("dd_recipient_caption") or row.get("dd_estimate_caption") or "").strip()
                document_id = (row.get("document_id") or row.get("id") or row.get("reg_number") or "").strip()
                document_date = parse_date(row.get("document_date") or row.get("date") or row.get("agreement_date") or row.get("close_date"))
                records.append(
                    make_record(
                        records,
                        source_file,
                        row_number,
                        row,
                        source="Соглашения",
                        snapshot=agreement_snapshot(row, path.name),
                        event_date=parse_date(row.get("close_date")),
                        record_kind="agreement_snapshot",
                        document_id=document_id,
                        document_date=document_date,
                        snapshot_source="monthly_snapshot",
                        amount_semantics="balance_as_of",
                        budget=(row.get("caption") or "").replace("!!! НЕ РАБОТАТЬ !!!", "").strip(),
                        object_code=kcsr,
                        object_code_norm=normalize_code(kcsr),
                        object_name=recipient or kcsr,
                        kfsr=display_code(row.get("kfsr_code")),
                        kvr=display_code(row.get("kvr_code")),
                        kosgu=display_code(row.get("kesr_code")),
                        counterparty=recipient,
                        document_number=(row.get("reg_number") or "").strip(),
                        description=class_names.get(str(row.get("documentclass_id")), "Соглашение"),
                        limit=MONEY_ZERO,
                        obligation=MONEY_ZERO,
                        cash=MONEY_ZERO,
                        agreement=parse_money(row.get("amount_1year")),
                        contract=MONEY_ZERO,
                        payment=MONEY_ZERO,
                        buau=MONEY_ZERO,
                    )
                )


def load_state_task(records: list[dict]) -> None:
    """Читает выгрузки госзадания: контракты и платежки как события."""
    folder = DATA_DIR / "3_StateTask"
    budget_lines: dict[str, list[dict[str, str]]] = defaultdict(list)
    lines_path = folder / "Бюджетные строки.csv"
    contracts_path = folder / "Контракты и договора.csv"
    payments_path = folder / "Платежки.csv"

    if lines_path.exists():
        for row in dict_rows(lines_path):
            budget_lines[row["con_document_id"]].append(row)

    contracts: dict[str, dict[str, str]] = {}
    for path in ([contracts_path] if contracts_path.exists() else []) + uploaded_paths("state_task_contracts"):
        source_file = relative_source(path)
        for row_number, row in enumerate(dict_rows(path), start=2):
                contracts[row["con_document_id"]] = row
                document_date = parse_date(row.get("con_date"))
                for line in budget_lines.get(row["con_document_id"], [{}]):
                    kcsr = display_code(line.get("kcsr_code"))
                    records.append(
                        make_record(
                            records,
                            source_file,
                            row_number,
                            row,
                            source="ГЗ: контракты",
                            snapshot=document_date,
                            event_date=document_date,
                            record_kind="contract_document",
                            document_id=(row.get("con_document_id") or row.get("con_number") or "").strip(),
                            document_date=document_date,
                            snapshot_source="event",
                            amount_semantics="event_amount",
                            budget="",
                            object_code=kcsr,
                            object_code_norm=normalize_code(kcsr),
                            object_name=kcsr or row.get("con_number", "").strip(),
                            kfsr=display_code(line.get("kfsr_code")),
                            kvr=display_code(line.get("kvr_code")),
                            kosgu=display_code(line.get("kesr_code")),
                            counterparty=(row.get("zakazchik_key") or "").strip(),
                            document_number=(row.get("con_number") or "").strip(),
                            description="Контракт/договор",
                            limit=MONEY_ZERO,
                            obligation=MONEY_ZERO,
                            cash=MONEY_ZERO,
                            agreement=MONEY_ZERO,
                            contract=parse_money(row.get("con_amount")),
                            payment=MONEY_ZERO,
                            buau=MONEY_ZERO,
                        )
                    )

    for path in ([payments_path] if payments_path.exists() else []) + uploaded_paths("state_task_payments"):
        source_file = relative_source(path)
        for row_number, row in enumerate(dict_rows(path), start=2):
                contract = contracts.get(row["con_document_id"], {})
                related_lines = budget_lines.get(row["con_document_id"], [{}])
                document_date = parse_date(row.get("platezhka_paydate"))
                for line in related_lines:
                    kcsr = display_code(line.get("kcsr_code"))
                    records.append(
                        make_record(
                            records,
                            source_file,
                            row_number,
                            row,
                            source="ГЗ: платежи",
                            snapshot=document_date,
                            event_date=document_date,
                            record_kind="payment_event",
                            document_id=(row.get("platezhka_id") or row.get("platezhka_num") or "").strip(),
                            document_date=document_date,
                            snapshot_source="event",
                            amount_semantics="event_amount",
                            budget="",
                            object_code=kcsr,
                            object_code_norm=normalize_code(kcsr),
                            object_name=kcsr or contract.get("con_number", "").strip(),
                            kfsr=display_code(line.get("kfsr_code")),
                            kvr=display_code(line.get("kvr_code")),
                            kosgu=display_code(line.get("kesr_code")),
                            counterparty=(contract.get("zakazchik_key") or "").strip(),
                            document_number=(row.get("platezhka_num") or "").strip(),
                            description="Оплата по контракту",
                            limit=MONEY_ZERO,
                            obligation=MONEY_ZERO,
                            cash=MONEY_ZERO,
                            agreement=MONEY_ZERO,
                            contract=MONEY_ZERO,
                            payment=parse_money(row.get("platezhka_amount")),
                            buau=MONEY_ZERO,
                        )
                    )


def load_buau(records: list[dict]) -> None:
    """Читает выплаты БУ/АУ как события, накопительные к выбранной дате."""
    folder = DATA_DIR / "4_BUAU_Export"
    for path in sorted(folder.glob("*.csv")) + uploaded_paths("buau"):
        snapshot = buau_snapshot(path.name)
        source_file = relative_source(path)
        for row_number, row in enumerate(dict_rows(path, delimiter=";"), start=2):
                kcsr = display_code(row.get("КЦСР"))
                organization = (row.get("Организация") or "").strip()
                event_date = parse_date(row.get("Дата проводки")) or snapshot
                records.append(
                    make_record(
                        records,
                        source_file,
                        row_number,
                        row,
                        source="БУАУ",
                        snapshot=snapshot,
                        event_date=event_date,
                        record_kind="buau_event",
                        document_id="",
                        document_date=event_date,
                        snapshot_source="event",
                        amount_semantics="event_amount",
                        budget=(row.get("Бюджет") or "").strip(),
                        object_code=kcsr,
                        object_code_norm=normalize_code(kcsr),
                        object_name=organization or kcsr,
                        kfsr=display_code(row.get("КФСР")),
                        kvr=display_code(row.get("КВР")),
                        kosgu=display_code(row.get("КОСГУ")),
                        counterparty=organization,
                        document_number="",
                        description=(row.get("Орган, предоставляющий субсидии") or "").strip(),
                        limit=MONEY_ZERO,
                        obligation=MONEY_ZERO,
                        cash=MONEY_ZERO,
                        agreement=MONEY_ZERO,
                        contract=MONEY_ZERO,
                        payment=MONEY_ZERO,
                        buau=parse_money(row.get("Выплаты с учетом возврата")),
                    )
                )


def apply_filters(records: list[dict], params: dict[str, list[str]]) -> list[dict]:
    """Применяет текстовые, кодовые, бюджетные и шаблонные фильтры."""
    text = params.get("q", [""])[0].strip().lower()
    code = normalize_code(params.get("code", [""])[0])
    budget = params.get("budget", [""])[0].strip().lower()
    source = params.get("source", [""])[0].strip()
    start = params.get("start", [""])[0].strip()
    end = params.get("end", [""])[0].strip()
    template = params.get("template", ["all"])[0].strip() or "all"

    result = []
    for record in records:
        haystack = " ".join(
            str(record.get(key, ""))
            for key in ("object_name", "object_code", "budget", "counterparty", "document_number", "description")
        ).lower()
        if text and text not in haystack:
            continue
        if code and code not in record.get("object_code_norm", ""):
            continue
        if budget and budget not in record.get("budget", "").lower():
            continue
        if source and record.get("source") != source:
            continue
        if not matches_template(record, template):
            continue
        date_value = record.get("snapshot") or record.get("event_date") or ""
        if start and date_value and date_value < start:
            continue
        if end and date_value and date_value > end:
            continue
        result.append(record)
    return result


def reporting_dates_payload() -> list[dict]:
    return [{"date": date, "label": datetime.strptime(date, "%Y-%m-%d").strftime("%d.%m.%Y")} for date in STORE.meta.get("reporting_dates", [])]


def select_as_of(records: list[dict], date: str, params: dict[str, list[str]]) -> list[dict]:
    """Собирает состояние на дату.

    Срезовые источники берутся последним доступным снимком не позже даты,
    а событийные источники суммируются накопительно до этой же даты.
    """
    if not date:
        dates = STORE.meta.get("reporting_dates", []) if "STORE" in globals() else []
        date = dates[-1] if dates else ""
    latest_rcb = max(
        (record.get("snapshot") for record in records if record.get("record_kind") == "rcb_snapshot" and record.get("snapshot") and record["snapshot"] <= date),
        default="",
    )
    latest_agreement = max(
        (record.get("snapshot") for record in records if record.get("record_kind") == "agreement_snapshot" and record.get("snapshot") and record["snapshot"] <= date),
        default="",
    )
    selected = []
    for record in records:
        kind = record.get("record_kind")
        if kind == "rcb_snapshot":
            if record.get("snapshot") == latest_rcb:
                selected.append(record)
        elif kind == "agreement_snapshot":
            if record.get("snapshot") == latest_agreement:
                selected.append(record)
        elif kind in {"contract_document", "payment_event", "buau_event"}:
            event_date = record.get("document_date") or record.get("event_date") or record.get("snapshot") or ""
            if event_date and event_date <= date:
                selected.append(record)
    filter_params = {key: value for key, value in params.items() if key not in {"view", "date", "start", "end", "base", "target", "metrics", "post_filter"}}
    return apply_filters(selected, filter_params)


def percent_or_none(numerator: Decimal, denominator: Decimal) -> float | None:
    return float(numerator / denominator) if denominator else None


def row_pipeline(row: dict) -> dict:
    """Сводит метрики объекта в управленческую цепочку план-документы-оплаты-касса."""
    plan = money_sum([row.get("limit"), row.get("obligation")])
    documents = money_sum([row.get("agreement"), row.get("contract")])
    paid = money_sum([row.get("payment"), row.get("buau")])
    cash = parse_money(row.get("cash"))
    missing_steps = []
    if plan > 0 and documents == 0:
        missing_steps.append("documents")
    if documents > 0 and paid == 0:
        missing_steps.append("payment")
    if plan > 0 and cash == 0:
        missing_steps.append("cash")
    return {
        "plan": plan,
        "documents": documents,
        "paid": paid,
        "cash": cash,
        "plan_to_documents_percent": percent_or_none(documents, plan),
        "documents_to_paid_percent": percent_or_none(paid, documents),
        "cash_to_plan_percent": percent_or_none(cash, plan),
        "missing_steps": missing_steps,
    }


def problem_reasons(row: dict) -> list[str]:
    """Возвращает машинные причины, по которым объект попадает в проблемные."""
    pipeline = row.get("pipeline") or row_pipeline(row)
    reasons = []
    if "documents" in pipeline["missing_steps"]:
        reasons.append("no_documents")
    if "payment" in pipeline["missing_steps"]:
        reasons.append("no_payments")
    if "cash" in pipeline["missing_steps"]:
        reasons.append("no_cash")
    if pipeline["plan"] > 0 and pipeline["cash"] > 0 and pipeline["cash"] / pipeline["plan"] < 0.25:
        reasons.append("low_cash")
    if int(row.get("source_count") or 0) == 1:
        reasons.append("data_gap")
    return reasons


RISK_MODEL_VERSION = "2026-05-05"
RISK_RULES = [
    {"code": "no_cash", "label": "Нет кассового исполнения", "points": 30},
    {"code": "no_payments", "label": "Есть документы, но нет оплат", "points": 25},
    {"code": "no_documents", "label": "Есть план, но нет документов", "points": 20},
    {"code": "low_cash", "label": "Касса ниже 25% от плана", "points": 15},
    {"code": "data_gap", "label": "Данные есть не во всех источниках", "points": 15},
    {"code": "large_plan_1b", "label": "План от 1 млрд", "points": 10},
    {"code": "large_plan_100m", "label": "План от 100 млн", "points": 7},
    {"code": "documents_without_paid", "label": "Документы есть, оплат нет", "points": 5},
    {"code": "plan_without_cash", "label": "План есть, кассы нет", "points": 5},
]


def risk_breakdown(row: dict) -> dict:
    """Объясняет приоритет ручной проверки теми же правилами, что и численный риск."""
    pipeline = row.get("pipeline") or row_pipeline(row)
    reasons = row.get("problem_reasons") or problem_reasons(row)
    plan = float(pipeline.get("plan") or 0)
    documents = float(pipeline.get("documents") or 0)
    paid = float(pipeline.get("paid") or 0)
    cash = float(pipeline.get("cash") or 0)
    score = 0
    factors = []

    def add_factor(code: str) -> None:
        nonlocal score
        rule = next((item for item in RISK_RULES if item["code"] == code), None)
        if not rule:
            return
        points = int(rule["points"])
        score += points
        factors.append({"code": code, "label": rule["label"], "points": points})

    if "no_cash" in reasons:
        add_factor("no_cash")
    if "no_payments" in reasons:
        add_factor("no_payments")
    if "no_documents" in reasons:
        add_factor("no_documents")
    if "low_cash" in reasons:
        add_factor("low_cash")
    if "data_gap" in reasons:
        add_factor("data_gap")
    if plan >= 1_000_000_000:
        add_factor("large_plan_1b")
    elif plan >= 100_000_000:
        add_factor("large_plan_100m")
    if documents > 0 and paid == 0:
        add_factor("documents_without_paid")
    if cash == 0 and plan > 0:
        add_factor("plan_without_cash")
    score = min(score, 100)
    level = risk_level(score)
    return {
        "version": RISK_MODEL_VERSION,
        "score": score,
        "level": level,
        "label": risk_label(level),
        "factors": factors,
    }


def risk_score(row: dict) -> int:
    """Считает приоритет ручной проверки, а не юридическую оценку нарушения."""
    return risk_breakdown(row)["score"]


def risk_level(score: int) -> str:
    if score >= 75:
        return "critical"
    if score >= 50:
        return "high"
    if score >= 25:
        return "medium"
    return "low"


def risk_label(level: str) -> str:
    return {
        "critical": "Критичный",
        "high": "Высокий",
        "medium": "Средний",
        "low": "Низкий",
    }.get(level, "Низкий")


def risk_explanation(row: dict) -> list[str]:
    pipeline = row.get("pipeline") or row_pipeline(row)
    reasons = row.get("problem_reasons") or problem_reasons(row)
    explanation = []
    reason_labels = {
        "no_cash": "Нет кассового исполнения",
        "no_payments": "Есть документы, но нет оплат",
        "no_documents": "Есть план, но нет документов",
        "low_cash": "Касса ниже 25% от плана",
        "data_gap": "Данные есть не во всех источниках",
    }
    for reason in ("no_cash", "no_payments", "no_documents", "low_cash", "data_gap"):
        if reason in reasons:
            explanation.append(reason_labels[reason])
    if float(pipeline.get("plan") or 0) >= 100_000_000:
        explanation.append("Крупная сумма плана")
    return explanation[:4]


def short_object_name(row: dict) -> str:
    name = str(row.get("object_name") or row.get("object_code") or "Объект без названия")
    return name if len(name) <= 95 else f"{name[:92]}..."


def top_risk_payload(row: dict) -> dict:
    pipeline = row.get("pipeline") or row_pipeline(row)
    breakdown = row.get("risk_breakdown") or risk_breakdown(row)
    return {
        "object_key": row.get("object_key", ""),
        "object_code": row.get("object_code", ""),
        "object_name": row.get("object_name") or row.get("object_code") or "Объект без названия",
        "short_name": short_object_name(row),
        "budget": row.get("budget", ""),
        "risk_score": int(row.get("risk_score") or breakdown["score"]),
        "risk_label": row.get("risk_label") or breakdown["label"],
        "risk_breakdown": breakdown,
        "review": row.get("review") or review_for_object(row.get("object_key", "")),
        "plan": float(pipeline.get("plan") or 0),
        "cash": float(pipeline.get("cash") or 0),
        "documents": float(pipeline.get("documents") or 0),
        "paid": float(pipeline.get("paid") or 0),
        "reasons": row.get("risk_explanation") or risk_explanation(row),
    }


def next_actions_payload(mode: str, has_top_risk: bool = False, has_problems: bool = False) -> list[dict]:
    actions: list[dict] = []
    if has_top_risk:
        actions.append({"label": "Открыть главный риск", "action": {"open": "top_risk"}})
    if mode == "compare":
        if has_problems:
            actions.append({"label": "Показать новые проблемы", "action": {"open_view": "changes"}})
        actions.append({"label": "Скачать Excel", "action": {"download": "excel"}})
        actions.append({"label": "Скачать PDF", "action": {"download": "pdf"}})
        return actions[:5]
    if has_problems:
        actions.extend(
            [
                {"label": "Показать непроверенные", "action": {"post_filter": "unreviewed"}},
                {"label": "Показать без кассы", "action": {"post_filter": "no_cash"}},
                {"label": "Показать без оплат", "action": {"post_filter": "no_payments"}},
                {"label": "Показать без документов", "action": {"post_filter": "no_documents"}},
            ]
        )
    actions = actions[:3]
    actions.append({"label": "Скачать Excel", "action": {"download": "excel"}})
    actions.append({"label": "Скачать PDF", "action": {"download": "pdf"}})
    return actions[:5]


def object_diagnosis(row: dict) -> dict:
    reasons = row.get("problem_reasons") or problem_reasons(row)
    labels = {
        "no_cash": "Есть план, но кассового исполнения нет.",
        "no_documents": "Есть план, но документы не найдены.",
        "no_payments": "Есть документы, но оплат не найдено.",
        "low_cash": "Касса ниже 25% от плана.",
        "data_gap": "Данные есть не во всех источниках.",
    }
    bullets = [labels[reason] for reason in ("no_cash", "no_documents", "no_payments", "low_cash", "data_gap") if reason in reasons]
    if not bullets:
        bullets.append("Явных проблем по цепочке денег не видно.")
    bullets.append("Проверьте исходные строки и документы перед управленческим решением.")
    severity = "danger" if any(reason in reasons for reason in ("no_cash", "no_documents", "no_payments")) else "warning" if reasons else "normal"
    return {"title": "Что проверить", "bullets": bullets, "severity": severity}


def row_status_from_reasons(reasons: list[str]) -> str:
    if any(reason in reasons for reason in ("no_documents", "no_payments", "no_cash")):
        return "danger"
    if any(reason in reasons for reason in ("low_cash", "data_gap")):
        return "warning"
    return "ok"


def aggregate(records: list[dict], metrics: list[str] | None = None) -> dict:
    """Группирует нормализованные строки в таблицу объектов и totals для UI."""
    groups: dict[str, dict] = {}
    timeline: dict[str, dict[str, Decimal]] = defaultdict(lambda: defaultdict(lambda: MONEY_ZERO))
    metric_keys = metrics or list(METRIC_KEYS)

    totals = {key: MONEY_ZERO for key in metric_keys}
    for record in records:
        key = record.get("object_key") or object_group_key(record)
        row = groups.setdefault(
            key,
            {
                "object_key": key,
                "object_code": record.get("object_code") or "",
                "object_name": record.get("object_name") or "",
                "budget": record.get("budget") or "",
                "object_aliases": set(),
                "sources": set(),
                **{metric: MONEY_ZERO for metric in METRIC_KEYS},
            },
        )
        if not row["object_code"] and record.get("object_code"):
            row["object_code"] = record["object_code"]
        if (
            record.get("source") == "РЧБ"
            and record.get("object_name")
            and len(record["object_name"]) > len(row["object_name"])
        ):
            row["object_name"] = record["object_name"]
        if record.get("object_name"):
            row["object_aliases"].add(record["object_name"])
        row["sources"].add(record["source"])
        point = record.get("snapshot") or record.get("event_date") or "unknown"
        for metric in METRIC_KEYS:
            value = parse_money(record.get(metric))
            row[metric] += value
            if metric in metric_keys:
                totals[metric] += value
            if metric in metric_keys and point != "unknown":
                timeline[point][metric] += value

    rows = []
    for row in groups.values():
        row["source_count"] = len(row["sources"])
        row["sources"] = ", ".join(sorted(row["sources"]))
        row["object_aliases"] = sorted(row["object_aliases"])[:10]
        row["match_confidence"] = "high" if not str(row["object_key"]).startswith("name:") else "medium"
        row["pipeline"] = row_pipeline(row)
        row["problem_reasons"] = problem_reasons(row)
        row["status"] = row_status_from_reasons(row["problem_reasons"])
        row["risk_breakdown"] = risk_breakdown(row)
        row["risk_score"] = row["risk_breakdown"]["score"]
        row["risk_level"] = row["risk_breakdown"]["level"]
        row["risk_label"] = row["risk_breakdown"]["label"]
        row["risk_explanation"] = risk_explanation(row)
        row["review"] = review_for_object(row["object_key"])
        row["total"] = money_sum(row[metric] for metric in metric_keys)
        rows.append(row)
    rows.sort(key=lambda item: item["total"], reverse=True)

    timeline_rows = []
    for date in sorted(timeline):
        point = {"date": date}
        point.update({metric: timeline[date].get(metric, MONEY_ZERO) for metric in metric_keys})
        timeline_rows.append(point)

    return {
        "totals": totals,
        "rows": rows[:300],
        "details": records[:500],
        "timeline": timeline_rows,
        "count": len(records),
    }


def apply_aggregate_post_filter(result: dict, post_filter: str, metrics: list[str]) -> dict:
    """Фильтрует уже агрегированные объекты по проблемам исполнения."""
    aliases = {"execution_problems": "execution_problems"}
    selected = aliases.get(post_filter, post_filter)
    valid = {"no_documents", "no_payments", "no_cash", "low_cash", "low_execution", "data_gap", "execution_problems", "unreviewed"}
    if selected not in valid:
        return result
    rows = []
    for row in result.get("rows", []):
        reasons = row.get("problem_reasons") or problem_reasons(row)
        if selected == "unreviewed" and (row.get("review") or {}).get("status") in {"new", "in_progress"}:
            rows.append(row)
        elif selected == "execution_problems" and reasons:
            rows.append(row)
        elif selected == "low_execution":
            plan = money_sum([row.get("limit"), row.get("obligation")])
            execution = money_sum([row.get("cash"), row.get("payment"), row.get("buau")])
            if plan > 0 and (execution == 0 or execution / plan < 0.25):
                rows.append(row)
        elif selected in reasons:
            rows.append(row)
    filtered = dict(result)
    filtered["rows"] = sorted(
        rows,
        key=lambda item: (int(item.get("risk_score") or 0), float((item.get("pipeline") or row_pipeline(item)).get("plan") or 0)),
        reverse=True,
    )
    filtered["totals"] = {metric: money_sum(row.get(metric) for row in rows) for metric in metrics}
    filtered["count"] = len(rows)
    return filtered


def attention_summary(result: dict, template: str, date: str) -> dict:
    """Готовит короткий управленческий вывод и топ рисков для первого экрана."""
    rows = result.get("rows") or []
    if not rows:
        return {
            "title": "Что требует внимания",
            "severity": "empty",
            "bullets": ["По выбранным условиям нет объектов для проверки."],
            "top_risks": [],
            "next_actions": next_actions_payload("slice", False, False),
        }
    counts = {reason: 0 for reason in ("no_cash", "no_payments", "no_documents", "low_cash", "data_gap")}
    for row in rows:
        for reason in row.get("problem_reasons") or []:
            if reason in counts:
                counts[reason] += 1
    plan_total = money_sum((row.get("pipeline") or row_pipeline(row)).get("plan") for row in rows)
    cash_total = money_sum((row.get("pipeline") or row_pipeline(row)).get("cash") for row in rows)
    cash_percent = float(cash_total / plan_total * Decimal("100.0")) if plan_total > 0 else None
    has_critical = any(row.get("risk_level") == "critical" for row in rows)
    has_problems = any(row.get("problem_reasons") for row in rows)
    if has_critical or (cash_percent is not None and cash_percent < 10):
        severity = "danger"
    elif has_problems:
        severity = "warning"
    else:
        severity = "normal"

    candidates = [
        (counts["no_cash"], 50, f"{counts['no_cash']} объектов без кассового исполнения."),
        (counts["no_payments"], 45, f"{counts['no_payments']} объектов с документами, но без оплат."),
        (counts["no_documents"], 40, f"{counts['no_documents']} объектов с планом, но без документов."),
        (counts["data_gap"], 35, f"{counts['data_gap']} объектов имеют разрыв между источниками."),
        (counts["low_cash"], 30, f"{counts['low_cash']} объектов с кассой ниже 25% от плана."),
    ]
    bullets = [text for count, _, text in sorted(candidates, key=lambda item: (item[0] > 0, item[1], item[0]), reverse=True) if count > 0]
    if cash_percent is not None:
        bullets.append(f"Касса составляет {cash_percent:.1f}% от плана.".replace(".", ","))
    if not bullets:
        bullets.append("Явных проблем исполнения не найдено.")

    top_rows = sorted(
        [row for row in rows if int(row.get("risk_score") or 0) >= 25],
        key=lambda item: (int(item.get("risk_score") or 0), float((item.get("pipeline") or row_pipeline(item)).get("plan") or 0)),
        reverse=True,
    )[:5]
    top_risks = []
    for row in top_rows:
        top_risks.append(top_risk_payload(row))
    return {
        "title": "Что требует внимания",
        "severity": severity,
        "bullets": bullets[:5],
        "top_risks": top_risks,
        "next_actions": next_actions_payload("slice", bool(top_risks), has_problems),
    }


def catalog_objects(records: list[dict], params: dict[str, list[str]]) -> list[dict]:
    query = params.get("q", [""])[0].strip().lower()
    template = params.get("template", ["all"])[0].strip() or "all"
    seen: set[tuple[str, str, str]] = set()
    objects = []
    for record in records:
        if not matches_template(record, template):
            continue
        haystack = " ".join(
            str(record.get(key, ""))
            for key in ("object_code", "object_name", "budget", "counterparty", "document_number")
        ).lower()
        if query and query not in haystack:
            continue
        item = (
            record.get("object_code", ""),
            record.get("object_name", ""),
            record.get("budget", ""),
        )
        if item in seen:
            continue
        seen.add(item)
        objects.append({"code": item[0], "name": item[1], "budget": item[2]})
        if len(objects) >= 200:
            break
    return objects


def load_rag_documents() -> list[dict]:
    """Загружает небольшую markdown-базу знаний для помощника."""
    documents = []
    if not RAG_DIR.exists():
        return documents
    for path in sorted(RAG_DIR.glob("*.md")):
        text = path.read_text(encoding="utf-8").strip()
        if text:
            documents.append({"source_file": relative_source(path), "title": path.stem, "content": text})
    return documents


def retrieve_rag_context(message: str, limit: int = 4) -> str:
    """Выбирает релевантные RAG-документы простым лексическим скорингом."""
    query_words = {word for word in re.findall(r"[0-9A-Za-zА-Яа-яЁё/]+", message.lower()) if len(word) > 1}
    scored = []
    for document in load_rag_documents():
        content = document["content"].lower()
        score = sum(1 for word in query_words if word in content)
        if score:
            scored.append((score, document))
    scored.sort(key=lambda item: item[0], reverse=True)
    selected = [document for _, document in scored[:limit]]
    if not selected and load_rag_documents():
        selected = load_rag_documents()[:1]
    return "\n\n".join(f"[{item['source_file']}]\n{item['content']}" for item in selected)


def closest_reporting_date(date: str, dates: list[str] | None = None) -> str:
    """Возвращает доступную отчетную дату: точную или ближайшую не позже запроса."""
    available = dates or (STORE.meta.get("reporting_dates", []) if "STORE" in globals() else [])
    if not date or not available:
        return ""
    if date in available:
        return date
    earlier = [item for item in available if item <= date]
    return earlier[-1] if earlier else available[0]


def extract_assistant_dates(message: str, dates: list[str] | None = None) -> list[str]:
    """Достает даты из обычной русской фразы и приводит их к reporting dates."""
    available = dates or (STORE.meta.get("reporting_dates", []) if "STORE" in globals() else [])
    candidates: list[tuple[int, str]] = []

    for match in re.finditer(r"\b(\d{1,2})[.](\d{1,2})[.](20\d{2})\b", message):
        day, month, year = match.groups()
        normalized = f"{int(year):04d}-{int(month):02d}-{int(day):02d}"
        candidates.append((match.start(), normalized))

    for match in re.finditer(r"\b(20\d{2})-(\d{1,2})-(\d{1,2})\b", message):
        year, month, day = match.groups()
        normalized = f"{int(year):04d}-{int(month):02d}-{int(day):02d}"
        candidates.append((match.start(), normalized))

    month_pattern = "|".join(sorted((re.escape(name) for name in MONTHS), key=len, reverse=True))
    for match in re.finditer(rf"\b(\d{{1,2}})?\s*({month_pattern})\s+(20\d{{2}})\b", message.lower()):
        day, month_name, year = match.groups()
        normalized = f"{int(year):04d}-{MONTHS[month_name]:02d}-{int(day or 1):02d}"
        candidates.append((match.start(), normalized))

    result: list[str] = []
    seen = set()
    for _, raw_date in sorted(candidates, key=lambda item: item[0]):
        normalized = closest_reporting_date(raw_date, available)
        if normalized and normalized not in seen:
            result.append(normalized)
            seen.add(normalized)
    return result


def clean_search_text(message: str) -> str:
    month_pattern = "|".join(sorted((re.escape(name) for name in MONTHS), key=len, reverse=True))
    text = re.sub(r"\b\d{1,2}[.]\d{1,2}[.]20\d{2}\b", " ", message)
    text = re.sub(r"\b20\d{2}-\d{1,2}-\d{1,2}\b", " ", text)
    text = re.sub(rf"\b\d{{1,2}}\s*(?:{month_pattern})\s+20\d{{2}}\b", " ", text, flags=re.I)
    text = re.sub(rf"\b(?:{month_pattern})\s+20\d{{2}}\b", " ", text, flags=re.I)
    text = re.sub(r"\b(покажи|показать|найди|найти|сравни|сравнить|что|такое|где|есть|по|в|и|с|со|на|за|от|до|между|период|объект\w*|только|выведи|сделай|собери|но|или|либо)\b", " ", text, flags=re.I)
    text = re.sub(r"\b(скк|кик|окв|бо|бу/?ау|буау|отчет\w*|отчёт\w*|капитал\w*|капвлож\w*|вложени\w*|качеств\w*|лимит\w*|касс\w*|исполн\w*|платеж\w*|оплат\w*|динамик\w*|измен\w*|проблем\w*|разрыв\w*|данн\w*|документ\w*|договор\w*|соглаш\w*|непровер\w*|провер\w*|контроль|загрузк\w*|скач\w*|excel|эксель|pdf|пдф|таблиц\w*)\b", " ", text, flags=re.I)
    text = re.sub(r"\b(без|нет|низк\w*|нулев\w*)\b", " ", text, flags=re.I)
    text = re.sub(r"\b(6105|978|970|2/3)\b", " ", text)
    return " ".join(text.split())


def assistant_rule_based(message: str, context: dict | None = None) -> dict:
    """Детерминированный парсер пользовательского вопроса без внешнего LLM."""
    context = context or {}
    lower = message.lower()
    start, end = default_date_range()
    parsed_dates = extract_assistant_dates(message)
    action = {
        "mode": context.get("mode") or "slice",
        "template": context.get("template") or "all",
        "q": "",
        "code": "",
        "budget": "",
        "source": "",
        "date": parsed_dates[-1] if parsed_dates else context.get("date") or end,
        "metrics": context.get("selected_metrics") or ["limit", "obligation", "cash"],
        "open_view": "overview",
    }
    intent = "run_query"
    confidence = 0.62
    full_metrics = ["limit", "obligation", "cash", "agreement", "contract", "payment", "buau"]

    if "скк" in lower or "6105" in lower:
        action["template"] = "skk"
        confidence = 0.9
    elif "кик" in lower or "978" in lower:
        action["template"] = "kik"
        confidence = 0.88
    elif "2/3" in lower or "970" in lower:
        action["template"] = "two_thirds"
        confidence = 0.88
    elif "окв" in lower or "капитал" in lower or "капвлож" in lower:
        action["template"] = "okv"
        confidence = 0.86

    if any(word in lower for word in ("сравн", "измен", "динамик")):
        intent = "run_compare"
        action["mode"] = "compare"
        action["base"] = parsed_dates[0] if len(parsed_dates) >= 1 else start
        action["target"] = parsed_dates[1] if len(parsed_dates) >= 2 else parsed_dates[0] if len(parsed_dates) == 1 else end
        action["open_view"] = "changes"
        confidence = max(confidence, 0.82)

    if any(word in lower for word in ("касс", "исполн", "платеж", "оплат")):
        action["metrics"] = ["cash", "payment", "buau"]
    if any(word in lower for word in ("лимит", "план", "бо")):
        action["metrics"] = ["limit", "obligation"]
    if any(word in lower for word in ("проблем", "нет касс", "без касс", "низк", "нет оплат", "нет платеж", "без оплат", "без платеж", "нет документ", "без документ", "нет договор", "без договор", "нет соглаш", "без соглаш", "непровер", "разрыв")):
        intent = "show_execution_problems"
        if action["template"] == "all" and context.get("template") in TEMPLATES:
            action["template"] = context.get("template") or "all"
        action["metrics"] = full_metrics
        action["post_filter"] = "execution_problems"
        action["open_view"] = "problems"
    if "нет касс" in lower or "без касс" in lower or "нулев" in lower and "касс" in lower:
        action["post_filter"] = "no_cash"
    if "нет оплат" in lower or "нет платеж" in lower or "без оплат" in lower or "без платеж" in lower:
        action["post_filter"] = "no_payments"
    if "нет документ" in lower or "без документ" in lower or "нет договор" in lower or "без договор" in lower or "нет соглаш" in lower or "без соглаш" in lower:
        action["post_filter"] = "no_documents"
    if "низк" in lower and ("исполн" in lower or "касс" in lower):
        action["post_filter"] = "low_cash"
    if "разрыв" in lower or "одном источник" in lower or "одного источник" in lower:
        action["post_filter"] = "data_gap"
        action["open_view"] = "problems"
    if "непровер" in lower or "новые для провер" in lower:
        action["post_filter"] = "unreviewed"
        action["open_view"] = "problems"

    if any(word in lower for word in ("контроль загруз", "проверить загруз", "качество загруз", "качество данных")):
        intent = "help"
        action["q"] = ""
        action["open"] = "control"
        action["open_view"] = "overview"
        confidence = max(confidence, 0.9)

    if any(word in lower for word in ("скачать", "выгрузи", "выгрузить", "экспорт")):
        if "pdf" in lower or "пдф" in lower:
            action["download"] = "pdf"
        elif "csv" in lower or "таблиц" in lower:
            action["download"] = "csv"
        else:
            action["download"] = "excel"
        if action["download"] == "excel":
            intent = "export_excel"
        confidence = max(confidence, 0.84)

    if any(phrase in lower for phrase in ("что такое", "объясни", "расскажи")):
        intent = "explain_metric" if any(word in lower for word in ("бо", "касс", "лимит", "метрик")) else "explain_template"
        confidence = max(confidence, 0.75)

    search_text = clean_search_text(message)
    if intent in {"run_query", "run_compare", "find_object"} and search_text:
        action["q"] = search_text
        if re.fullmatch(r"\d{4,}", search_text) and search_text not in {"6105", "0970", "970", "0978", "978"}:
            action["code"] = search_text

    rag_context = retrieve_rag_context(message, limit=2)
    explanation = ""
    if intent.startswith("explain") and rag_context:
        explanation = " " + " ".join(rag_context.split())[:700]

    alternative_label = "Искать во всех данных" if search_text else "Показать все данные"
    followups = [
        {"label": "Открыть главный риск", "action": {"open": "top_risk"}},
        {"label": "Показать контроль загрузки", "action": {"open": "control"}},
        {"label": "Скачать Excel", "action": {"download": "excel"}},
    ]
    alternatives = [
        {
            "label": alternative_label,
            "action": {
                "mode": "slice",
                "template": "all",
                "q": search_text,
                "code": "",
                "budget": "",
                "source": "",
                "post_filter": "",
                "date": action.get("date", end),
                "reset_scope": True,
                "metrics": action["metrics"],
            },
        },
    ]
    return {
        "mode": "rule_based",
        "intent": intent,
        "confidence": confidence,
        "message": f"Я понял запрос как {'сравнение' if intent == 'run_compare' else 'выборку'}: {TEMPLATES.get(action['template'], TEMPLATES['all'])['label']}.{explanation}",
        "action": action,
        "followups": followups,
        "alternatives": alternatives,
        "rag_context": rag_context,
    }


def validate_assistant_action(action: dict, fallback: dict) -> dict:
    """Ограничивает действие LLM публичным набором фильтров, шаблонов и метрик."""
    dates = STORE.meta.get("reporting_dates", []) if "STORE" in globals() else []
    first_date = dates[0] if dates else ""
    last_date = dates[-1] if dates else ""
    result = {key: value for key, value in dict(fallback).items() if key in ASSISTANT_ACTION_FIELDS}
    if not isinstance(action, dict):
        if not result.get("date"):
            result["date"] = last_date
        return result
    if action.get("mode") in {"slice", "compare"}:
        result["mode"] = action["mode"]
    if action.get("template") in TEMPLATES:
        result["template"] = action["template"]
    for key in ("q", "code", "budget", "source"):
        if key in action:
            result[key] = str(action.get(key) or "")[:200]
    if action.get("post_filter") in ASSISTANT_POST_FILTERS:
        result["post_filter"] = action.get("post_filter") or ""
    if action.get("open_view") in ASSISTANT_OPEN_VIEWS:
        result["open_view"] = action["open_view"]
    if action.get("open") in ASSISTANT_OPEN_ACTIONS:
        result["open"] = action.get("open") or ""
    if action.get("download") in ASSISTANT_DOWNLOAD_ACTIONS:
        result["download"] = action.get("download") or ""
    if isinstance(action.get("reset_scope"), bool):
        result["reset_scope"] = action["reset_scope"]
    for key in ("date", "base", "target"):
        if action.get(key):
            normalized_date = closest_reporting_date(str(action.get(key)), dates)
            if normalized_date:
                result[key] = normalized_date
    metrics = action.get("metrics")
    if isinstance(metrics, list):
        filtered = [metric for metric in metrics if metric in METRIC_KEYS]
        if filtered:
            result["metrics"] = filtered
    if result.get("mode") == "compare":
        result["base"] = result.get("base") if result.get("base") in dates else first_date
        result["target"] = result.get("target") if result.get("target") in dates else last_date
        result["open_view"] = "changes"
    else:
        result["date"] = result.get("date") if result.get("date") in dates else last_date
        result.pop("base", None)
        result.pop("target", None)
        result["open_view"] = result.get("open_view") if result.get("open_view") in ASSISTANT_OPEN_VIEWS else "problems" if result.get("post_filter") else "overview"
    return result


def validate_assistant_followups(followups: object) -> list[dict]:
    result = []
    if not isinstance(followups, list):
        return result
    for item in followups[:5]:
        if not isinstance(item, dict):
            continue
        label = str(item.get("label") or "")[:80].strip()
        action = item.get("action") if isinstance(item.get("action"), dict) else {}
        clean_action = {}
        for key in ASSISTANT_FOLLOWUP_ACTION_FIELDS:
            if key in action:
                clean_action[key] = str(action.get(key) or "")[:80]
        if label and clean_action:
            result.append({"label": label, "action": clean_action})
    return result


def apply_message_overrides(message: str, action: dict) -> dict:
    """Исправляет типовые промахи LLM по датам, фильтрам и служебному поиску."""
    result = dict(action)
    lower = message.lower()
    dates = extract_assistant_dates(message)

    if "скк" in lower or "6105" in lower:
        result["template"] = "skk"
    elif "кик" in lower or "978" in lower:
        result["template"] = "kik"
    elif "2/3" in lower or "970" in lower:
        result["template"] = "two_thirds"
    elif "окв" in lower or "капитал" in lower or "капвлож" in lower:
        result["template"] = "okv"

    if any(word in lower for word in ("сравн", "измен", "динамик")):
        start, end = default_date_range()
        result["mode"] = "compare"
        result["base"] = dates[0] if len(dates) >= 1 else result.get("base") or start
        result["target"] = dates[1] if len(dates) >= 2 else result.get("target") or end
        result["open_view"] = "changes"
    elif dates:
        result["mode"] = "slice"
        result["date"] = dates[-1]
        if result.get("open_view") == "changes":
            result["open_view"] = "overview"

    if any(word in lower for word in ("проблем", "нет касс", "без касс", "низк", "нет оплат", "нет платеж", "без оплат", "без платеж", "нет документ", "без документ", "нет договор", "без договор", "нет соглаш", "без соглаш", "непровер", "разрыв")):
        result["open_view"] = "problems"
        result.setdefault("post_filter", "execution_problems")
    if "нет касс" in lower or "без касс" in lower or ("нулев" in lower and "касс" in lower):
        result["post_filter"] = "no_cash"
    if "нет оплат" in lower or "нет платеж" in lower or "без оплат" in lower or "без платеж" in lower:
        result["post_filter"] = "no_payments"
    if "нет документ" in lower or "без документ" in lower or "нет договор" in lower or "без договор" in lower or "нет соглаш" in lower or "без соглаш" in lower:
        result["post_filter"] = "no_documents"
    if "низк" in lower and ("исполн" in lower or "касс" in lower):
        result["post_filter"] = "low_cash"
    if "разрыв" in lower or "одном источник" in lower or "одного источник" in lower:
        result["post_filter"] = "data_gap"
    if "непровер" in lower:
        result["post_filter"] = "unreviewed"

    if any(word in lower for word in ("касс", "исполн", "платеж", "оплат", "буау", "бу/ау")):
        result["metrics"] = ["cash", "payment", "buau"]
    if any(word in lower for word in ("лимит", "план", "бо")):
        result["metrics"] = ["limit", "obligation"]
    if result.get("post_filter"):
        result["metrics"] = ["limit", "obligation", "cash", "agreement", "contract", "payment", "buau"]

    if any(word in lower for word in ("контроль загруз", "проверить загруз", "качество загруз", "качество данных")):
        result["open"] = "control"
        result["open_view"] = "overview"
    if any(word in lower for word in ("скачать", "выгрузи", "выгрузить", "экспорт")):
        result["download"] = "pdf" if "pdf" in lower or "пдф" in lower else "csv" if "csv" in lower or "таблиц" in lower else "excel"

    result["q"] = clean_search_text(message)
    if re.fullmatch(r"\d{4,}", result["q"]) and result["q"] not in {"6105", "0970", "970", "0978", "978"}:
        result["code"] = result["q"]
    return result


def normalize_assistant_intent(intent: str, message: str, action: dict, fallback_intent: str) -> str:
    """Keeps intent consistent with the final action applied by the UI."""
    lower = message.lower()
    if action.get("download") == "excel":
        return "export_excel"
    if action.get("open") == "control":
        return "help"
    if action.get("mode") == "compare":
        return "run_compare"
    if action.get("post_filter"):
        return "show_execution_problems"
    if any(phrase in lower for phrase in ("что такое", "объясни", "расскажи")):
        if intent in {"explain_metric", "explain_template", "explain_result"}:
            return intent
        return fallback_intent if fallback_intent in ASSISTANT_INTENTS else "explain_template"
    if action.get("q") or action.get("code") or action.get("budget") or action.get("source"):
        return "find_object"
    if intent in ASSISTANT_INTENTS:
        return intent
    return fallback_intent if fallback_intent in ASSISTANT_INTENTS else "run_query"


def assistant_json_schema() -> dict:
    action_schema = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "mode": {"type": "string", "enum": ["slice", "compare"]},
            "template": {"type": "string", "enum": list(TEMPLATES)},
            "date": {"type": "string"},
            "base": {"type": "string"},
            "target": {"type": "string"},
            "open": {"type": "string", "enum": sorted(ASSISTANT_OPEN_ACTIONS)},
            "download": {"type": "string", "enum": sorted(ASSISTANT_DOWNLOAD_ACTIONS)},
            "q": {"type": "string"},
            "code": {"type": "string"},
            "budget": {"type": "string"},
            "source": {"type": "string"},
            "post_filter": {"type": "string", "enum": sorted(ASSISTANT_POST_FILTERS)},
            "metrics": {"type": "array", "items": {"type": "string", "enum": METRIC_KEYS}},
            "open_view": {"type": "string", "enum": sorted(ASSISTANT_OPEN_VIEWS)},
            "reset_scope": {"type": "boolean"},
        },
        "required": sorted(ASSISTANT_ACTION_FIELDS),
    }
    followup_schema = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "label": {"type": "string"},
            "action": {
                "type": "object",
                "additionalProperties": False,
                "properties": {
                    "open": {"type": "string"},
                    "download": {"type": "string"},
                    "post_filter": {"type": "string"},
                    "open_view": {"type": "string"},
                },
                "required": sorted(ASSISTANT_FOLLOWUP_ACTION_FIELDS),
            },
        },
        "required": ["label", "action"],
    }
    return {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "intent": {"type": "string", "enum": sorted(ASSISTANT_INTENTS)},
            "confidence": {"type": "number"},
            "message": {"type": "string"},
            "action": action_schema,
            "followups": {"type": "array", "items": followup_schema},
        },
        "required": ["intent", "confidence", "message", "action", "followups"],
    }


def groq_chat_completion(model: str, api_key: str, messages: list[dict], response_format: dict) -> dict:
    payload_request = {
        "model": model,
        "messages": messages,
        "temperature": 0.1,
        "max_completion_tokens": 700,
        "response_format": response_format,
    }
    response = requests.post(
        "https://api.groq.com/openai/v1/chat/completions",
        json=payload_request,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "User-Agent": "expense-analytics/1.0",
        },
        timeout=30,
    )
    response.raise_for_status()
    payload = response.json()
    return json.loads(payload["choices"][0]["message"]["content"])


def assistant_llm(message: str, context: dict, rag_context: str) -> dict:
    """Опционально уточняет intent через Groq, не передавая исходные записи."""
    api_key = os.environ.get("GROQ_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError("GROQ_API_KEY is not set")
    model = os.environ.get("GROQ_MODEL", "llama-3.1-8b-instant").strip() or "llama-3.1-8b-instant"
    start, end = default_date_range()
    system_prompt = (
        "Return only compact JSON: intent, confidence, message, action, followups. "
        "Do not calculate money. Use only allowed codes from user JSON. "
        "Rules: SKK/6105=skk, KIK/978=kik, 970/2/3=two_thirds, OKV/capital=okv. "
        "Compare=>mode compare, base first date, target second date, open_view changes. "
        "No cash=>no_cash; no payments=>no_payments; no docs/contracts/agreements=>no_documents; unreviewed=>unreviewed; data gaps=>data_gap. "
        "Control load=>open control and q empty. Download=>download excel/pdf/csv and q empty. "
        "Do not put dates, months, service words, metric words, template words into q."
    )
    user_payload = {
        "message": message,
        "context": {
            "mode": context.get("mode") or "slice",
            "template": context.get("template") or "all",
            "date": context.get("date") or end,
            "base": context.get("base") or start,
            "target": context.get("target") or end,
        },
        "allowed": {
            "intents": sorted(ASSISTANT_INTENTS),
            "dates": STORE.meta.get("reporting_dates", []),
            "templates": list(TEMPLATES),
            "metrics": METRIC_KEYS,
            "post_filters": sorted(ASSISTANT_POST_FILTERS),
            "open_views": sorted(ASSISTANT_OPEN_VIEWS),
            "open": sorted(ASSISTANT_OPEN_ACTIONS),
            "download": sorted(ASSISTANT_DOWNLOAD_ACTIONS),
        },
        "required_action_fields": sorted(ASSISTANT_ACTION_FIELDS),
    }
    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": json.dumps(json_safe(user_payload), ensure_ascii=False)},
    ]
    if model.startswith("llama-") or model.startswith("meta-llama/"):
        parsed = groq_chat_completion(model, api_key, messages, {"type": "json_object"})
    else:
        schema_format = {
            "type": "json_schema",
            "json_schema": {
                "name": "assistant_action",
                "strict": True,
                "schema": assistant_json_schema(),
            },
        }
        try:
            parsed = groq_chat_completion(model, api_key, messages, schema_format)
        except requests.HTTPError as exc:
            if exc.response is not None and exc.response.status_code == 429:
                raise
            parsed = groq_chat_completion(model, api_key, messages, {"type": "json_object"})
        except Exception:
            parsed = groq_chat_completion(model, api_key, messages, {"type": "json_object"})
    fallback = assistant_rule_based(message, context)
    action = validate_assistant_action(parsed.get("action", {}), fallback["action"])
    action = validate_assistant_action(apply_message_overrides(message, action), fallback["action"])
    intent = normalize_assistant_intent(str(parsed.get("intent") or ""), message, action, fallback["intent"])
    followups = validate_assistant_followups(parsed.get("followups")) or fallback.get("followups", [])
    response_message = str(parsed.get("message") or fallback["message"])[:1000]
    if "Я понял запрос" not in response_message:
        response_message = fallback["message"]
    return {
        "mode": "llm",
        "intent": intent,
        "confidence": float(parsed.get("confidence") or fallback["confidence"]),
        "message": response_message,
        "action": action,
        "followups": followups,
        "alternatives": fallback.get("alternatives", []),
        "rag_context": rag_context,
    }


def assistant_response(message: str, context: dict | None = None) -> dict:
    """Возвращает ответ помощника с безопасным fallback на правила."""
    context = context or {}
    fallback = assistant_rule_based(message, context)
    if os.environ.get("ASSISTANT_ENABLED", "auto").lower() == "false":
        fallback["fallback_reason"] = "assistant_disabled"
        return fallback
    if not os.environ.get("GROQ_API_KEY", "").strip():
        fallback["fallback_reason"] = "missing_groq_key"
        return fallback
    try:
        return assistant_llm(message, context, fallback.get("rag_context", ""))
    except requests.HTTPError as exc:
        status = exc.response.status_code if exc.response is not None else "unknown"
        fallback["fallback_reason"] = f"groq_http_{status}"
        return fallback
    except Exception as exc:
        fallback["fallback_reason"] = exc.__class__.__name__
        return fallback


def explain_rule_based(kind: str, payload: dict) -> dict:
    summary = payload.get("attention_summary") if isinstance(payload.get("attention_summary"), dict) else {}
    compare = payload.get("compare_insights") if isinstance(payload.get("compare_insights"), dict) else {}
    readiness = payload.get("readiness") if isinstance(payload.get("readiness"), dict) else {}
    top_risks = payload.get("top_risks") if isinstance(payload.get("top_risks"), list) else []
    source = compare if kind == "compare" and compare else summary
    bullets = [str(item) for item in source.get("bullets", [])[:5] if item]
    if not bullets and readiness:
        checks = readiness.get("checks") if isinstance(readiness.get("checks"), list) else []
        bullets = [str(item.get("message") or item.get("label")) for item in checks[:4] if isinstance(item, dict)]
    if not bullets and top_risks:
        bullets = [f"Сначала проверьте {item.get('object_name') or 'объект'}: {item.get('risk_label') or 'высокий риск'}." for item in top_risks[:3]]
    if not bullets:
        bullets = ["Главное внимание стоит уделить объектам без кассы, оплат или документов."]
    return {
        "mode": "rule_based",
        "title": "Короткое объяснение",
        "bullets": bullets[:5],
        "next_actions": [
            {"label": "Открыть главный риск", "action": {"open": "top_risk"}},
            {"label": "Скачать Excel", "action": {"download": "excel"}},
        ],
    }


def explain_llm(kind: str, payload: dict) -> dict:
    api_key = os.environ.get("GROQ_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError("GROQ_API_KEY is not set")
    model = os.environ.get("GROQ_MODEL", "llama-3.1-8b-instant").strip() or "llama-3.1-8b-instant"
    allowed_payload = {
        "kind": kind,
        "attention_summary": payload.get("attention_summary") if isinstance(payload.get("attention_summary"), dict) else {},
        "compare_insights": payload.get("compare_insights") if isinstance(payload.get("compare_insights"), dict) else {},
        "readiness": payload.get("readiness") if isinstance(payload.get("readiness"), dict) else {},
        "top_risks": (payload.get("top_risks") if isinstance(payload.get("top_risks"), list) else [])[:5],
        "filters": payload.get("filters") if isinstance(payload.get("filters"), dict) else {},
    }
    schema = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "title": {"type": "string"},
            "bullets": {"type": "array", "items": {"type": "string"}},
            "next_actions": {
                "type": "array",
                "items": {
                    "type": "object",
                    "additionalProperties": False,
                    "properties": {
                        "label": {"type": "string"},
                        "action": {
                            "type": "object",
                            "additionalProperties": False,
                            "properties": {"open": {"type": "string"}, "download": {"type": "string"}},
                            "required": ["open", "download"],
                        },
                    },
                    "required": ["label", "action"],
                },
            },
        },
        "required": ["title", "bullets", "next_actions"],
    }
    messages = [
        {
            "role": "system",
            "content": "Объясни результат простыми словами. Не считай суммы, не проси raw records, верни только JSON по schema.",
        },
        {"role": "user", "content": json.dumps(json_safe(allowed_payload), ensure_ascii=False)},
    ]
    parsed = groq_chat_completion(
        model,
        api_key,
        messages,
        {"type": "json_schema", "json_schema": {"name": "explain_result", "strict": True, "schema": schema}},
    )
    fallback = explain_rule_based(kind, payload)
    bullets = [str(item)[:240] for item in parsed.get("bullets", []) if item][:5]
    return {
        "mode": "llm",
        "title": str(parsed.get("title") or fallback["title"])[:120],
        "bullets": bullets or fallback["bullets"],
        "next_actions": validate_assistant_followups(parsed.get("next_actions")) or fallback["next_actions"],
    }


def explain_response(kind: str, payload: dict | None = None) -> dict:
    payload = payload if isinstance(payload, dict) else {}
    kind = kind if kind in {"query", "compare", "object"} else "query"
    fallback = explain_rule_based(kind, payload)
    if os.environ.get("ASSISTANT_ENABLED", "auto").lower() == "false":
        return fallback
    if not os.environ.get("GROQ_API_KEY", "").strip():
        return fallback
    try:
        return explain_llm(kind, payload)
    except Exception:
        return fallback


def trace_record(record_id: str) -> dict | None:
    """Раскрывает нормализованную и исходную строку для кнопки 'Откуда цифра'."""
    for record in STORE.records:
        if record.get("id") == record_id:
            normalized = {key: value for key, value in record.items() if key not in {"raw"}}
            amount_fields = []
            for key, label in (
                ("limit", "Лимиты"),
                ("obligation", "БО"),
                ("cash", "Касса"),
                ("agreement", "Соглашения"),
                ("contract", "Договоры"),
                ("payment", "Оплаты"),
                ("buau", "БУ/АУ"),
            ):
                value = parse_money(record.get(key))
                if value:
                    amount_fields.append({"label": label, "field": key, "value": value})
            return {
                "id": record.get("id"),
                "source": record.get("source"),
                "source_file": record.get("source_file", ""),
                "source_row": record.get("source_row", ""),
                "human_summary": {
                    "title": "Сумма из исходной строки",
                    "date": record.get("event_date") or record.get("snapshot") or "",
                    "object": record.get("object_name") or record.get("object_code") or "",
                    "document": record.get("document_number") or record.get("description") or "",
                    "amount_fields": amount_fields,
                },
                "raw": record.get("raw", {}),
                "normalized": normalized,
            }
    return None


AS_OF_SEMANTICS = {
    "limit": "balance_as_of",
    "obligation": "balance_as_of",
    "cash": "balance_as_of",
    "agreement": "balance_as_of",
    "contract": "cumulative_to_date",
    "payment": "cumulative_to_date",
    "buau": "cumulative_to_date",
}


def readiness_check(status: str, code: str, label: str, message: str) -> dict:
    return {"code": code, "label": label, "status": status, "message": message}


def readiness_summary(records: list[dict], rows: list[dict], params: dict[str, list[str]]) -> dict:
    """Собирает проверки, которые предупреждают пользователя о неполной выборке."""
    date = params.get("date", [""])[0].strip()
    if not date:
        dates = STORE.meta.get("reporting_dates", [])
        date = dates[-1] if dates else ""
    template = params.get("template", ["all"])[0].strip() or "all"
    sources = {record.get("source") for record in records}
    checks = [
        readiness_check(
            "ok" if "РЧБ" in sources else "bad",
            "rcb_loaded",
            "Плановые данные найдены",
            "Есть строки РЧБ на выбранную дату" if "РЧБ" in sources else "Нет строк РЧБ на выбранную дату",
        ),
        readiness_check(
            "ok" if "Соглашения" in sources else "warn",
            "agreements_loaded",
            "Соглашения найдены",
            "Есть соглашения на выбранную дату" if "Соглашения" in sources else "Соглашения не попали в выборку",
        ),
        readiness_check(
            "ok" if "ГЗ: контракты" in sources else "warn",
            "contracts_loaded",
            "Контракты найдены",
            "Есть контракты на выбранную дату" if "ГЗ: контракты" in sources else "Контракты есть не по всем объектам",
        ),
        readiness_check(
            "ok" if {"ГЗ: платежи", "БУАУ"} & sources else "warn",
            "payments_loaded",
            "Платежи найдены",
            "Есть платежи до выбранной даты" if {"ГЗ: платежи", "БУАУ"} & sources else "Платежи не попали в выборку",
        ),
    ]
    has_gaps = any("data_gap" in (row.get("problem_reasons") or []) for row in rows)
    checks.append(
        readiness_check(
            "warn" if has_gaps else "ok",
            "data_gaps",
            "Разрывы данных",
            "Есть объекты только в одном источнике" if has_gaps else "Критичных разрывов по источникам не видно",
        )
    )
    if not rows:
        checks.append(readiness_check("bad", "empty_result", "Выборка не пустая", "По выбранным условиям нет объектов"))
    else:
        checks.append(readiness_check("ok", "empty_result", "Выборка не пустая", "Есть объекты для показа"))
    summary = {status: sum(1 for check in checks if check["status"] == status) for status in ("ok", "warn", "bad")}
    return {"date": date, "template": template, "summary": summary, "checks": checks}


def query_as_of(params: dict[str, list[str]]) -> dict:
    """Основной API-сценарий: состояние на дату с рисками, выводом и графиком."""
    date = params.get("date", [""])[0].strip()
    if not date:
        dates = STORE.meta.get("reporting_dates", [])
        date = dates[-1] if dates else ""
    metrics = selected_metrics(params)
    filtered = select_as_of(STORE.records, date, params)
    result = aggregate(filtered, metrics)
    result = apply_aggregate_post_filter(result, params.get("post_filter", [""])[0].strip(), metrics)
    result["timeline"] = as_of_timeline(date, params, metrics)
    summary = attention_summary(result, params.get("template", ["all"])[0].strip() or "all", date)
    return {
        "view": "as_of",
        "date": date,
        "semantics": AS_OF_SEMANTICS,
        "attention_summary": summary,
        **result,
    }


def as_of_timeline(date: str, params: dict[str, list[str]], metrics: list[str]) -> list[dict]:
    """Строит точки графика в той же as-of семантике, что и текущая выборка."""
    post_filter = params.get("post_filter", [""])[0].strip()
    dates = [item for item in STORE.meta.get("reporting_dates", []) if not date or item <= date]
    points = []
    for point_date in dates:
        point_params = {**params, "date": [point_date]}
        records = select_as_of(STORE.records, point_date, point_params)
        result = aggregate(records, metrics)
        result = apply_aggregate_post_filter(result, post_filter, metrics)
        point = {"date": point_date}
        point.update({metric: result["totals"].get(metric, MONEY_ZERO) for metric in metrics})
        points.append(point)
    return points


def readiness_response(params: dict[str, list[str]]) -> dict:
    date = params.get("date", [""])[0].strip()
    if not date:
        dates = STORE.meta.get("reporting_dates", [])
        date = dates[-1] if dates else ""
        params = {**params, "date": [date]}
    metrics = selected_metrics(params)
    records = select_as_of(STORE.records, date, params)
    result = aggregate(records, metrics)
    result = apply_aggregate_post_filter(result, params.get("post_filter", [""])[0].strip(), metrics)
    return readiness_summary(records, result["rows"], params)


def object_detail(params: dict[str, list[str]]) -> dict:
    """Возвращает карточку объекта с документами и исходными строками."""
    object_key = params.get("object_key", [""])[0].strip()
    if not object_key:
        return {"error": "object_key_required"}
    date = params.get("date", [""])[0].strip()
    if not date:
        dates = STORE.meta.get("reporting_dates", [])
        date = dates[-1] if dates else ""
    records = [record for record in select_as_of(STORE.records, date, params) if (record.get("object_key") or object_group_key(record)) == object_key]
    result = aggregate(records)
    if not result["rows"]:
        return {"error": "object_not_found"}
    row = result["rows"][0]
    documents = []
    for record in records:
        if record.get("source") in {"Соглашения", "ГЗ: контракты", "ГЗ: платежи", "БУАУ"}:
            amount = money_sum(record.get(metric) for metric in ("agreement", "contract", "payment", "buau"))
            documents.append(
                {
                    "source": record.get("source", ""),
                    "date": record.get("document_date") or record.get("event_date") or record.get("snapshot") or "",
                    "number": record.get("document_number", ""),
                    "counterparty": record.get("counterparty", ""),
                    "amount": amount,
                }
            )
    return {
        "object_key": object_key,
        "object_code": row.get("object_code", ""),
        "object_name": row.get("object_name", ""),
        "budget": row.get("budget", ""),
        "status": row.get("status", ""),
        "problem_reasons": row.get("problem_reasons", []),
        "risk_score": row.get("risk_score", 0),
        "risk_level": row.get("risk_level", "low"),
        "risk_label": row.get("risk_label", risk_label("low")),
        "risk_breakdown": row.get("risk_breakdown") or risk_breakdown(row),
        "risk_explanation": row.get("risk_explanation", []),
        "review": row.get("review") or review_for_object(object_key),
        "diagnosis": object_diagnosis(row),
        "pipeline": row.get("pipeline", {}),
        "sources": [source.strip() for source in str(row.get("sources", "")).split(",") if source.strip()],
        "documents": documents[:100],
        "records": records[:100],
    }


def export_excel(params: dict[str, list[str]]) -> tuple[bytes, str]:
    """Формирует рабочий Excel: выводы, итоги, объекты, проблемы и методику."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError("excel_dependency_missing") from exc

    mode = params.get("mode", ["slice"])[0].strip()
    template = params.get("template", ["all"])[0].strip() or "all"
    if mode == "compare":
        result = compare_periods(params)
        rows = result["rows"]
        details: list[dict] = []
        date_label = f"{result['base']} - {result['target']}"
        totals = {metric: money_sum(row.get("metrics", {}).get(metric, {}).get("target") for row in rows) for metric in selected_metrics(params)}
        summary = result.get("compare_insights") or {"title": "Что изменилось", "severity": "normal", "bullets": [], "top_risks": []}
    else:
        result = query_as_of(params)
        rows = result["rows"]
        details = result["details"]
        date_label = result["date"]
        totals = result["totals"]
        summary = result.get("attention_summary") or attention_summary(result, template, date_label)

    wb = Workbook()
    ws = wb.active
    ws.title = "Выводы"
    pipeline = row_pipeline(totals)
    problem_count = sum(1 for row in rows if row.get("problem_reasons"))
    report_created = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.append(["Аналитика расходов"])
    ws.append([f"Период: {date_label}", f"Шаблон: {TEMPLATES.get(template, TEMPLATES['all'])['label']}", f"Сформировано: {report_created}"])
    ws.append([])
    ws.append(["Что требует внимания"])
    for bullet in summary.get("bullets") or []:
        ws.append([bullet])
    top_risks = summary.get("top_risks") or summary.get("new_problem_objects") or summary.get("worsened_objects") or []
    ws.append([])
    ws.append(["Главные риски"])
    ws.append(["Объект", "Код", "Бюджет", "Риск", "План", "Касса", "Причины"])
    for item in top_risks[:5]:
        ws.append([
            item.get("short_name") or item.get("object_name", ""),
            item.get("object_code", ""),
            item.get("budget", ""),
            f"{item.get('risk_label', '')} {item.get('risk_score', '')}".strip(),
            item.get("plan", 0),
            item.get("cash", 0),
            ", ".join(item.get("reasons") or []),
        ])
    ws.append([])
    ws.append(["Следующие действия"])
    for item in ("Открыть карточку объекта в системе", "Проверить исходные строки", "Уточнить документы/оплаты"):
        ws.append([item])

    ws = wb.create_sheet("Итоги")
    summary_rows = [
        ("Лимиты", totals.get("limit", 0), "Доведенные бюджетные лимиты."),
        ("БО", totals.get("obligation", 0), "Принятые бюджетные обязательства."),
        ("Касса", totals.get("cash", 0), "Кассовое исполнение по РЧБ."),
        ("Соглашения", totals.get("agreement", 0), "Суммы соглашений."),
        ("Контракты", totals.get("contract", 0), "Суммы контрактов и договоров."),
        ("Платежи", totals.get("payment", 0), "Фактические платежи."),
        ("БУ/АУ", totals.get("buau", 0), "Выплаты учреждений."),
        ("Проблемных объектов", problem_count, "Строки с явными причинами риска."),
    ]
    ws.append(["Показатель", "Значение", "Пояснение"])
    for item in summary_rows:
        ws.append(list(item))

    objects_ws = wb.create_sheet("Объекты")
    objects_ws.append(["Объект", "Код", "Бюджет", "План", "Документы", "Оплачено", "Касса", "Статус", "Риск", "Балл риска", "Причины риска", "Факторы риска", "Статус проверки", "Ответственный", "Комментарий", "Источники"])
    for row in rows:
        pipeline = row.get("pipeline") or row_pipeline(row)
        factors = (row.get("risk_breakdown") or risk_breakdown(row)).get("factors") or []
        review = row.get("review") or default_review()
        objects_ws.append([
            row.get("object_name", ""),
            row.get("object_code", ""),
            row.get("budget", ""),
            pipeline.get("plan", 0),
            pipeline.get("documents", 0),
            pipeline.get("paid", 0),
            pipeline.get("cash", 0),
            row.get("status", ""),
            row.get("risk_label", ""),
            row.get("risk_score", 0),
            ", ".join(row.get("risk_explanation") or []),
            ", ".join(f"{item.get('label', '')} +{item.get('points', 0)}" for item in factors),
            review.get("label", ""),
            review.get("assignee", ""),
            review.get("comment", ""),
            row.get("sources", ""),
        ])

    problems_ws = wb.create_sheet("Проблемы")
    problems_ws.append(["Причина", "Причина текстом", "Риск", "Балл риска", "Объект", "Код", "План", "Документы", "Оплачено", "Касса", "Статус проверки", "Ответственный", "Комментарий", "Источники"])
    reason_labels = {
        "no_documents": "Есть план, но документы не найдены",
        "no_payments": "Есть документы, но оплат не найдено",
        "no_cash": "Есть план, но кассового исполнения нет",
        "low_cash": "Касса ниже 25% от плана",
        "data_gap": "Данные есть не во всех источниках",
    }
    problem_rows = sorted(
        [row for row in rows if row.get("problem_reasons")],
        key=lambda item: (int(item.get("risk_score") or 0), float((item.get("pipeline") or row_pipeline(item)).get("plan") or 0)),
        reverse=True,
    )
    for row in problem_rows:
        pipeline = row.get("pipeline") or row_pipeline(row)
        review = row.get("review") or default_review()
        for reason in row.get("problem_reasons") or []:
            problems_ws.append([
                reason,
                reason_labels.get(reason, "Проблема данных"),
                row.get("risk_label", ""),
                row.get("risk_score", 0),
                row.get("object_name", ""),
                row.get("object_code", ""),
                pipeline.get("plan", 0),
                pipeline.get("documents", 0),
                pipeline.get("paid", 0),
                pipeline.get("cash", 0),
                review.get("label", ""),
                review.get("assignee", ""),
                review.get("comment", ""),
                row.get("sources", ""),
            ])

    details_ws = wb.create_sheet("Исходные строки")
    details_ws.append(["Дата", "Источник", "Файл", "Строка", "Код", "Объект", "Документ", "Контрагент", "Сумма"])
    for record in details:
        amount = money_sum(record.get(metric) for metric in METRIC_KEYS)
        details_ws.append([
            record.get("event_date") or record.get("snapshot") or "",
            record.get("source", ""),
            record.get("source_file", ""),
            record.get("source_row", ""),
            record.get("object_code", ""),
            record.get("object_name", ""),
            record.get("document_number", ""),
            record.get("counterparty", ""),
            amount,
        ])

    control = control_summary(params)
    control_ws = wb.create_sheet("Контроль загрузки")
    control_ws.append(["Итоги загрузки", "", "", "", "", ""])
    control_ws.append(["Записей", "Источников", "Предупреждений", "Ошибок", "Связано по названию", "Объектов из одного источника"])
    control_ws.append([
        control["summary"].get("records", 0),
        control["summary"].get("sources", 0),
        control["summary"].get("warnings", 0),
        control["summary"].get("errors", 0),
        control["summary"].get("unmatched_name_keys", 0),
        control["summary"].get("single_source_objects", 0),
    ])
    control_ws.append([])
    control_ws.append(["Источники"])
    control_ws.append(["Источник", "Прочитано строк", "Записей", "Предупреждений", "Ошибок", "Сумма", "Файлы"])
    for item in control["sources"]:
        control_ws.append([
            item.get("source", ""),
            item.get("read_rows", 0),
            item.get("records", 0),
            item.get("warnings", 0),
            item.get("errors", 0),
            item.get("total_amount", 0),
            ", ".join(item.get("files") or []),
        ])
    control_ws.append([])
    control_ws.append(["Файлы"])
    control_ws.append(["Файл", "Прочитано строк", "Записей", "Предупреждений", "Ошибок"])
    for item in control["files"]:
        control_ws.append([
            item.get("source_file", ""),
            item.get("read_rows", 0),
            item.get("records", 0),
            item.get("warnings", 0),
            item.get("errors", 0),
        ])
    control_ws.append([])
    control_ws.append(["Связка объектов"])
    control_ws.append(["С кодом", "По названию", "Один источник", "Несколько источников"])
    linkage = control["object_linkage"]
    control_ws.append([
        linkage.get("with_code", 0),
        linkage.get("by_name", 0),
        linkage.get("single_source", 0),
        linkage.get("multi_source", 0),
    ])
    control_ws.append([])
    control_ws.append(["Предупреждения качества"])
    control_ws.append(["Уровень", "Код", "Файл", "Строка", "Поле", "Сообщение"])
    for issue in control["issues"]:
        control_ws.append([
            issue.get("severity", ""),
            issue.get("code", ""),
            issue.get("source_file", ""),
            issue.get("source_row", ""),
            issue.get("field", ""),
            issue.get("message", ""),
        ])

    method_ws = wb.create_sheet("Методика")
    method_ws.append(["Версия методики риска", RISK_MODEL_VERSION])
    method_ws.append([])
    for line in (
        "РЧБ и соглашения берутся как последний месячный срез не позже выбранной даты.",
        "Контракты, платежи и БУАУ учитываются накопительно до выбранной даты.",
        "План = лимиты + БО.",
        "Документы = соглашения + контракты.",
        "Оплачено = платежи + БУАУ.",
        "Касса = кассовые выплаты из РЧБ.",
        "Риск является управленческим индикатором для проверки и не является юридическим выводом.",
    ):
        method_ws.append([line])
    method_ws.append([])
    method_ws.append(["Код", "Фактор", "Баллы"])
    for rule in RISK_RULES:
        method_ws.append([rule["code"], rule["label"], rule["points"]])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=16, color="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    risk_fills = {
        "critical": PatternFill("solid", fgColor="F4CCCC"),
        "high": PatternFill("solid", fgColor="FCE4D6"),
        "medium": PatternFill("solid", fgColor="FFF2CC"),
    }
    thin_border = Border(
        left=Side(style="thin", color="D9E0E4"),
        right=Side(style="thin", color="D9E0E4"),
        top=Side(style="thin", color="D9E0E4"),
        bottom=Side(style="thin", color="D9E0E4"),
    )
    money_headers = {"Значение", "План", "Документы", "Оплачено", "Касса", "Сумма"}
    table_sheets = {"Итоги", "Объекты", "Проблемы", "Исходные строки", "Контроль загрузки"}

    ws = wb["Выводы"]
    ws.freeze_panes = "A4"
    ws["A1"].font = title_font
    for row_number in (4, 4 + len(summary.get("bullets") or []) + 2, ws.max_row - 3):
        for cell in ws[row_number]:
            cell.fill = section_fill
            cell.font = Font(bold=True, color="1F4E78")

    for sheet in wb.worksheets:
        if sheet.title in table_sheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
        if sheet.title != "Выводы":
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        if sheet.title == "Контроль загрузки":
            for row_number in (1, 5, control_ws.max_row):
                for cell in sheet[row_number]:
                    cell.fill = section_fill
                    cell.font = Font(bold=True, color="1F4E78")
        header_values = [cell.value for cell in sheet[1]]
        money_columns = {index + 1 for index, value in enumerate(header_values) if value in money_headers}
        for row in sheet.iter_rows():
            row_risk_level = ""
            if sheet.title == "Объекты":
                score = int(row[9].value or 0) if len(row) > 9 and isinstance(row[9].value, (int, float)) else 0
                row_risk_level = risk_level(score)
            elif sheet.title == "Проблемы":
                score = int(row[3].value or 0) if len(row) > 3 and isinstance(row[3].value, (int, float)) else 0
                row_risk_level = risk_level(score)
            fill = next((risk_fills[level] for level in risk_fills if level in row_risk_level), None)
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if cell.column in money_columns and isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"
                if fill and cell.row > 1:
                    cell.fill = fill
        for column in sheet.columns:
            width = min(55, max(12, *(len(str(cell.value or "")) + 2 for cell in column)))
            sheet.column_dimensions[column[0].column_letter].width = width

    buffer = BytesIO()
    wb.save(buffer)
    if mode == "compare":
        filename = f"analytics_compare_{template}_{params.get('base', [''])[0]}_{params.get('target', [''])[0]}.xlsx"
    else:
        filename = f"analytics_{template}_{params.get('date', ['as_of'])[0] or 'as_of'}.xlsx"
    return buffer.getvalue(), filename


def format_money_pdf(value: object) -> str:
    return f"{float(value or 0):,.0f}".replace(",", " ")


def register_pdf_font() -> tuple[str, str]:
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    # Кириллица надежнее отображается с системным TTF; Helvetica остается fallback.
    fonts = [
        ("ArialPdf", "ArialPdfBold", "C:/Windows/Fonts/arial.ttf", "C:/Windows/Fonts/arialbd.ttf"),
        ("CalibriPdf", "CalibriPdfBold", "C:/Windows/Fonts/calibri.ttf", "C:/Windows/Fonts/calibrib.ttf"),
        ("TahomaPdf", "TahomaPdfBold", "C:/Windows/Fonts/tahoma.ttf", "C:/Windows/Fonts/tahomabd.ttf"),
    ]
    for regular_name, bold_name, regular_path, bold_path in fonts:
        if not Path(regular_path).exists():
            continue
        try:
            pdfmetrics.registerFont(TTFont(regular_name, regular_path))
            if Path(bold_path).exists():
                pdfmetrics.registerFont(TTFont(bold_name, bold_path))
                return regular_name, bold_name
            return regular_name, regular_name
        except Exception:
            continue
    return "Helvetica", "Helvetica-Bold"


def export_pdf(params: dict[str, list[str]]) -> tuple[bytes, str]:
    """Формирует короткий презентационный PDF-отчет рядом с рабочим Excel."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    except ImportError as exc:
        raise RuntimeError("pdf_dependency_missing") from exc

    from html import escape

    mode = params.get("mode", ["slice"])[0].strip()
    template = params.get("template", ["all"])[0].strip() or "all"
    template_label = TEMPLATES.get(template, TEMPLATES["all"])["label"]
    if mode == "compare":
        result = compare_periods(params)
        rows = result["rows"]
        summary = result["compare_insights"]
        date_label = f"{result['base']} - {result['target']}"
        metrics = selected_metrics(params)
        totals = {metric: money_sum(row.get("metrics", {}).get(metric, {}).get("target") for row in rows) for metric in metrics}
        change_items = (
            list(summary.get("new_problem_objects") or [])
            + list(summary.get("worsened_objects") or [])
            + list(summary.get("stalled_cash_objects") or [])
        )
        filename = f"analytics_compare_{template}_{result['base']}_{result['target']}.pdf"
    else:
        result = query_as_of(params)
        rows = result["rows"]
        summary = result["attention_summary"]
        date_label = result["date"]
        totals = result["totals"]
        change_items = []
        filename = f"analytics_{template}_{params.get('date', ['as_of'])[0] or 'as_of'}.pdf"

    regular_font, bold_font = register_pdf_font()
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle("PdfTitle", parent=styles["Title"], fontName=bold_font, fontSize=22, leading=26, textColor=colors.HexColor("#1F4E78"), alignment=TA_CENTER, spaceAfter=10))
    styles.add(ParagraphStyle("PdfMeta", parent=styles["BodyText"], fontName=regular_font, fontSize=10, leading=14, alignment=TA_CENTER, textColor=colors.HexColor("#334155")))
    styles.add(ParagraphStyle("PdfSection", parent=styles["Heading2"], fontName=bold_font, fontSize=13, leading=16, textColor=colors.HexColor("#1F4E78"), backColor=colors.HexColor("#D9EAF7"), borderPadding=6, spaceBefore=10, spaceAfter=8))
    styles.add(ParagraphStyle("PdfBody", parent=styles["BodyText"], fontName=regular_font, fontSize=9, leading=12, textColor=colors.HexColor("#1f2937")))
    styles.add(ParagraphStyle("PdfSmall", parent=styles["BodyText"], fontName=regular_font, fontSize=8, leading=10, textColor=colors.HexColor("#475569")))
    styles.add(ParagraphStyle("PdfCell", parent=styles["BodyText"], fontName=regular_font, fontSize=7, leading=9, alignment=TA_LEFT))
    styles.add(ParagraphStyle("PdfHeaderCell", parent=styles["BodyText"], fontName=bold_font, fontSize=7, leading=9, textColor=colors.white, alignment=TA_LEFT))

    def p(text: object, style: str = "PdfBody") -> Paragraph:
        return Paragraph(escape(str(text or "")), styles[style])

    def short_cell(row: dict) -> str:
        name = str(row.get("short_name") or row.get("object_name") or row.get("object_code") or "Объект без названия")
        return name if len(name) <= 120 else f"{name[:117]}..."

    def risk_factor_text(item: dict) -> str:
        factors = ((item.get("risk_breakdown") or {}).get("factors") or [])[:3]
        if factors:
            text = ", ".join(str(factor.get("label") or "") for factor in factors if factor.get("label"))
        else:
            text = ", ".join(item.get("reasons") or [])[:160]
        review = item.get("review") or {}
        if review.get("label"):
            text = f"{text}; проверка: {review['label']}" if text else f"Проверка: {review['label']}"
        return text

    def table(data: list[list[object]], widths: list[float] | None = None, risk_column: int | None = None) -> Table:
        converted = [[p(cell, "PdfHeaderCell" if row_index == 0 else "PdfCell") for cell in row] for row_index, row in enumerate(data)]
        result_table = Table(converted, colWidths=widths, repeatRows=1, hAlign="LEFT")
        style = TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), bold_font),
                ("FONTNAME", (0, 1), (-1, -1), regular_font),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D9E0E4")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
        if risk_column is not None:
            fills = {"Критичный": "#F4CCCC", "Высокий": "#FCE4D6", "Средний": "#FFF2CC"}
            for row_index, raw_row in enumerate(data[1:], start=1):
                risk_text = str(raw_row[risk_column] if len(raw_row) > risk_column else "")
                for label, fill in fills.items():
                    if label in risk_text:
                        style.add("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor(fill))
                        break
        result_table.setStyle(style)
        return result_table

    pipeline = row_pipeline(totals)
    problem_count = sum(1 for row in rows if row.get("problem_reasons") or int(row.get("risk_score") or row.get("risk", {}).get("target") or 0) >= 25)
    story = [
        p("Аналитика расходов", "PdfTitle"),
        p(f"Шаблон: {template_label}", "PdfMeta"),
        p(f"Период/дата: {date_label}", "PdfMeta"),
        p(f"Сформировано: {datetime.now().strftime('%Y-%m-%d %H:%M')}", "PdfMeta"),
        Spacer(1, 5 * mm),
        p("Решение для руководителя", "PdfSection"),
    ]
    for bullet in (summary.get("bullets") or [])[:5]:
        story.append(p(f"• {bullet}"))
    story.append(p(f"• Проблемных объектов: {problem_count}"))

    executive_risks = []
    if mode == "compare":
        executive_risks = change_items[:3]
    else:
        executive_risks = (summary.get("top_risks") or [])[:3]
    for item in executive_risks:
        risk_text = f"{item.get('risk_label', '')} {item.get('risk_score', '')}".strip()
        factors_text = risk_factor_text(item)
        story.append(p(f"• {short_cell(item)}: {risk_text}; {factors_text}"))
    story.append(p("• Риск является приоритетом ручной проверки, не юридическим выводом."))
    story.extend([
        p("Что требует внимания", "PdfSection"),
    ])
    for bullet in (summary.get("bullets") or [])[:5]:
        story.append(p(f"• {bullet}"))
    story.extend(
        [
            p("Ключевые итоги", "PdfSection"),
            table(
                [
                    ["Показатель", "Значение"],
                    ["План", format_money_pdf(pipeline.get("plan"))],
                    ["Документы", format_money_pdf(pipeline.get("documents"))],
                    ["Оплачено", format_money_pdf(pipeline.get("paid"))],
                    ["Касса", format_money_pdf(pipeline.get("cash"))],
                    ["Проблемных объектов", problem_count],
                ],
                [55 * mm, 45 * mm],
            ),
        ]
    )

    risk_rows = [["Объект", "Код", "Бюджет", "Риск", "План", "Касса", "Причины"]]
    if mode == "compare":
        story.append(p("Главные изменения", "PdfSection"))
        for item in change_items[:7]:
            risk_rows.append(
                [
                    short_cell(item),
                    item.get("object_code", ""),
                    item.get("budget", ""),
                    f"{item.get('risk_label', '')} {item.get('risk_score', '')}".strip(),
                    format_money_pdf(item.get("plan")),
                    format_money_pdf(item.get("cash")),
                    risk_factor_text(item),
                ]
            )
    else:
        story.append(p("Главные риски", "PdfSection"))
        for item in (summary.get("top_risks") or [])[:7]:
            risk_rows.append(
                [
                    short_cell(item),
                    item.get("object_code", ""),
                    item.get("budget", ""),
                    f"{item.get('risk_label', '')} {item.get('risk_score', '')}".strip(),
                    format_money_pdf(item.get("plan")),
                    format_money_pdf(item.get("cash")),
                    risk_factor_text(item),
                ]
            )
    if len(risk_rows) == 1:
        risk_rows.append(["Нет данных", "", "", "", "", "", ""])
    story.append(table(risk_rows, [62 * mm, 24 * mm, 42 * mm, 27 * mm, 28 * mm, 28 * mm, 58 * mm], risk_column=3))

    actions = [item.get("label", "") for item in (summary.get("next_actions") or []) if item.get("label")]
    if not actions:
        actions = ["Открыть главный риск", "Проверить исходные строки", "Скачать Excel для детальной работы"]
    story.append(p("Что делать дальше", "PdfSection"))
    for action in actions[:5]:
        story.append(p(f"• {action}"))

    story.append(p("Методика", "PdfSection"))
    for line in (
        "План = лимиты + БО",
        "Документы = соглашения + контракты",
        "Оплачено = платежи + БУ/АУ",
        "Касса = кассовое исполнение из РЧБ",
        "Риск - управленческий индикатор для приоритизации проверки, не юридический вывод",
        f"Версия методики риска: {RISK_MODEL_VERSION}",
        "Сравнение использует те же правила состояния на дату." if mode == "compare" else "",
    ):
        if line:
            story.append(p(f"• {line}"))

    story.extend([PageBreak(), p("Приложение: Топ объектов", "PdfSection")])
    object_rows = [["Объект", "Код", "План", "Документы", "Оплачено", "Касса", "Риск"]]
    for row in rows[:20]:
        if mode == "compare":
            metrics = row.get("metrics", {})
            row_pipeline_data = {
                "plan": money_sum([metrics.get("limit", {}).get("target"), metrics.get("obligation", {}).get("target")]),
                "documents": money_sum([metrics.get("agreement", {}).get("target"), metrics.get("contract", {}).get("target")]),
                "paid": money_sum([metrics.get("payment", {}).get("target"), metrics.get("buau", {}).get("target")]),
                "cash": parse_money(metrics.get("cash", {}).get("target")),
            }
            risk_text = str(row.get("risk", {}).get("target") or 0)
        else:
            row_pipeline_data = row.get("pipeline") or row_pipeline(row)
            risk_text = f"{row.get('risk_label', '')} {row.get('risk_score', '')}".strip()
        object_rows.append(
            [
                short_cell(row),
                row.get("object_code", ""),
                format_money_pdf(row_pipeline_data.get("plan")),
                format_money_pdf(row_pipeline_data.get("documents")),
                format_money_pdf(row_pipeline_data.get("paid")),
                format_money_pdf(row_pipeline_data.get("cash")),
                risk_text,
            ]
        )
    story.append(table(object_rows, [76 * mm, 25 * mm, 28 * mm, 28 * mm, 28 * mm, 28 * mm, 32 * mm], risk_column=6))

    def on_page(canvas, doc):
        canvas.saveState()
        canvas.setFont(regular_font, 8)
        canvas.setFillColor(colors.HexColor("#64748B"))
        canvas.drawString(doc.leftMargin, 9 * mm, "Аналитика расходов")
        canvas.drawRightString(landscape(A4)[0] - doc.rightMargin, 9 * mm, f"стр. {doc.page}")
        canvas.restoreState()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=12 * mm, rightMargin=12 * mm, topMargin=12 * mm, bottomMargin=16 * mm)
    doc.build(story, onFirstPage=on_page, onLaterPages=on_page)
    return buffer.getvalue(), filename


def compare_object_summary(row: dict, risk_value: int | None = None, base_row: dict | None = None) -> dict:
    pipeline = row.get("pipeline") or row_pipeline(row)
    score = int(risk_value if risk_value is not None else row.get("risk_score") or 0)
    breakdown = row.get("risk_breakdown") or risk_breakdown(row)
    if risk_value is not None and int(breakdown.get("score") or 0) != score:
        level = risk_level(score)
        breakdown = {**breakdown, "score": score, "level": level, "label": risk_label(level)}
    payload = {
        "object_key": row.get("object_key", ""),
        "object_name": row.get("object_name") or row.get("object_code") or "Объект без названия",
        "object_code": row.get("object_code", ""),
        "budget": row.get("budget", ""),
        "risk_score": score,
        "risk_label": risk_label(risk_level(score)),
        "risk_breakdown": breakdown,
        "review": row.get("review") or review_for_object(row.get("object_key", "")),
        "plan": float(pipeline.get("plan") or 0),
        "cash": float(pipeline.get("cash") or 0),
        "documents": float(pipeline.get("documents") or 0),
        "paid": float(pipeline.get("paid") or 0),
        "reasons": row.get("risk_explanation") or risk_explanation(row),
    }
    if base_row is not None:
        base_pipeline = base_row.get("pipeline") or row_pipeline(base_row)
        payload["risk_delta"] = score - int(base_row.get("risk_score") or 0)
        payload["plan_delta"] = parse_money(pipeline.get("plan")) - parse_money(base_pipeline.get("plan"))
        payload["cash_delta"] = parse_money(pipeline.get("cash")) - parse_money(base_pipeline.get("cash"))
    return payload


def compare_insights(base_rows: list[dict], target_rows: list[dict]) -> dict:
    """Выделяет существенные изменения риска между двумя отчетными датами."""
    base_by_key = {row.get("object_key") or object_group_key(row): row for row in base_rows}
    target_by_key = {row.get("object_key") or object_group_key(row): row for row in target_rows}
    new_problem_objects = []
    resolved_problem_objects = []
    worsened_objects = []
    improved_objects = []
    stalled_cash_objects = []
    for key in sorted(set(base_by_key) | set(target_by_key)):
        base_row = base_by_key.get(key) or {"object_key": key, "risk_score": 0, "pipeline": row_pipeline({})}
        target_row = target_by_key.get(key) or {"object_key": key, "risk_score": 0, "pipeline": row_pipeline({})}
        base_score = int(base_row.get("risk_score") or 0)
        target_score = int(target_row.get("risk_score") or 0)
        base_pipeline = base_row.get("pipeline") or row_pipeline(base_row)
        target_pipeline = target_row.get("pipeline") or row_pipeline(target_row)
        display_row = target_row if key in target_by_key else base_row
        if base_score < 25 <= target_score:
            new_problem_objects.append(compare_object_summary(display_row, target_score, base_row if key in base_by_key else None))
        if base_score >= 25 > target_score:
            resolved_problem_objects.append(compare_object_summary(display_row, target_score, base_row if key in base_by_key else None))
        if target_score - base_score >= 20:
            worsened_objects.append(compare_object_summary(display_row, target_score, base_row if key in base_by_key else None))
        if base_score - target_score >= 20:
            improved_objects.append(compare_object_summary(display_row, target_score, base_row if key in base_by_key else None))
        if parse_money(target_pipeline.get("plan")) > parse_money(base_pipeline.get("plan")) and parse_money(target_pipeline.get("cash")) == parse_money(base_pipeline.get("cash")):
            stalled_cash_objects.append(compare_object_summary(display_row, target_score, base_row if key in base_by_key else None))

    sort_key = lambda item: (int(item.get("risk_score") or 0), float(item.get("plan") or 0))
    new_problem_objects = sorted(new_problem_objects, key=sort_key, reverse=True)[:10]
    resolved_problem_objects = sorted(resolved_problem_objects, key=sort_key, reverse=True)[:10]
    worsened_objects = sorted(worsened_objects, key=lambda item: (int(item.get("risk_delta") or 0), int(item.get("risk_score") or 0)), reverse=True)[:10]
    improved_objects = sorted(improved_objects, key=lambda item: (abs(int(item.get("risk_delta") or 0)), float(item.get("plan") or 0)), reverse=True)[:10]
    stalled_cash_objects = sorted(stalled_cash_objects, key=lambda item: float(item.get("plan_delta") or 0), reverse=True)[:10]

    base_cash = money_sum((row.get("pipeline") or row_pipeline(row)).get("cash") for row in base_rows)
    target_cash = money_sum((row.get("pipeline") or row_pipeline(row)).get("cash") for row in target_rows)
    cash_delta = target_cash - base_cash
    bullets = []
    if new_problem_objects:
        bullets.append(f"Появилось {len(new_problem_objects)} новых проблемных объектов.")
    if resolved_problem_objects:
        bullets.append(f"По {len(resolved_problem_objects)} объектам проблема ушла.")
    if improved_objects:
        bullets.append(f"У {len(improved_objects)} объектов риск снизился.")
    if worsened_objects:
        bullets.append(f"У {len(worsened_objects)} объектов риск вырос.")
    if cash_delta:
        direction = "выросла" if cash_delta > 0 else "снизилась"
        bullets.append(f"Касса {direction} на {abs(cash_delta):,.0f}.".replace(",", " "))
    if stalled_cash_objects:
        bullets.append(f"Есть {len(stalled_cash_objects)} объектов, где план вырос, а касса не изменилась.")
    if not bullets:
        bullets.append("Существенных ухудшений или улучшений по риску не видно.")
    severity = "danger" if new_problem_objects or worsened_objects else "warning" if stalled_cash_objects else "normal"
    return {
        "title": "Что изменилось",
        "severity": severity,
        "bullets": bullets[:5],
        "new_problem_objects": new_problem_objects,
        "resolved_problem_objects": resolved_problem_objects,
        "worsened_objects": worsened_objects,
        "improved_objects": improved_objects,
        "stalled_cash_objects": stalled_cash_objects,
        "next_actions": next_actions_payload("compare", bool(new_problem_objects or worsened_objects), bool(new_problem_objects)),
    }


def compare_periods(params: dict[str, list[str]]) -> dict:
    """Сравнивает две даты по тем же правилам as-of, что и основной режим."""
    base = params.get("base", [""])[0].strip()
    target = params.get("target", [""])[0].strip()
    if not base or not target:
        dates = STORE.meta.get("reporting_dates", [])
        base = base or (dates[0] if dates else "")
        target = target or (dates[-1] if dates else "")
    metrics = selected_metrics(params)
    base_rows = aggregate(select_as_of(STORE.records, base, params), metrics)["rows"]
    target_rows = aggregate(select_as_of(STORE.records, target, params), metrics)["rows"]
    insights = compare_insights(base_rows, target_rows)
    by_key: dict[tuple[str, str], dict] = {}
    for label, rows in (("base", base_rows), ("target", target_rows)):
        for row in rows:
            key = row.get("object_key") or object_group_key(row)
            item = by_key.setdefault(
                key,
                {
                    "object_key": key,
                    "object_code": row.get("object_code", ""),
                    "object_name": row.get("object_name", ""),
                    "budget": row.get("budget", ""),
                    "sources": row.get("sources", ""),
                    "risk": {"base": 0, "target": 0, "delta": 0},
                    "metrics": {metric: {"base": MONEY_ZERO, "target": MONEY_ZERO, "delta": MONEY_ZERO, "delta_percent": None} for metric in metrics},
                },
            )
            if row.get("sources"):
                item["sources"] = row["sources"]
            item["risk"][label] = int(row.get("risk_score") or 0)
            for metric in metrics:
                item["metrics"][metric][label] = parse_money(row.get(metric))
    rows = []
    for item in by_key.values():
        total_delta = MONEY_ZERO
        for metric in metrics:
            values = item["metrics"][metric]
            values["delta"] = values["target"] - values["base"]
            values["delta_percent"] = float(values["delta"] / values["base"] * Decimal("100.0")) if values["base"] else None
            total_delta += abs(values["delta"])
        item["risk"]["delta"] = item["risk"]["target"] - item["risk"]["base"]
        item["total_delta"] = total_delta
        item["review"] = review_for_object(item["object_key"])
        rows.append(item)
    rows.sort(key=lambda item: item["total_delta"], reverse=True)
    return {
        "base": base,
        "target": target,
        "view": "as_of",
        "semantics": AS_OF_SEMANTICS,
        "metrics": metrics,
        "available_dates": STORE.meta["reporting_dates"],
        "compare_insights": insights,
        "rows": rows[:300],
    }


STORE = load_data()


class Handler(SimpleHTTPRequestHandler):
    """Минимальный HTTP-слой: static files и JSON API без внешнего фреймворка."""

    def translate_path(self, path: str) -> str:
        parsed = urlparse(path)
        requested = parsed.path
        if requested == "/":
            return str(STATIC_DIR / "index.html")
        if requested.startswith("/static/"):
            return str(ROOT / requested.lstrip("/"))
        return str(STATIC_DIR / requested.lstrip("/"))

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/meta":
            self.write_json(STORE.meta)
            return
        if parsed.path == "/api/query":
            params = parse_qs(parsed.query)
            metrics = selected_metrics(params)
            if params.get("view", ["period"])[0].strip() == "as_of":
                self.write_json(query_as_of(params))
                return
            filtered = apply_filters(STORE.records, params)
            result = aggregate(filtered, metrics)
            post_filter = params.get("post_filter", [""])[0].strip()
            self.write_json(apply_aggregate_post_filter(result, post_filter, metrics))
            return
        if parsed.path == "/api/compare":
            self.write_json(compare_periods(parse_qs(parsed.query)))
            return
        if parsed.path == "/api/readiness":
            self.write_json(readiness_response(parse_qs(parsed.query)))
            return
        if parsed.path == "/api/control":
            self.write_json(control_summary(parse_qs(parsed.query)))
            return
        if parsed.path == "/api/reviews":
            self.write_json(reviews_payload())
            return
        if parsed.path == "/api/review":
            object_key = parse_qs(parsed.query).get("object_key", [""])[0].strip()
            if not object_key:
                self.write_json({"error": "object_key_required"}, status=400)
                return
            self.write_json(review_for_object(object_key))
            return
        if parsed.path == "/api/object":
            payload = object_detail(parse_qs(parsed.query))
            self.write_json(payload, status=404 if payload.get("error") == "object_not_found" else 400 if payload.get("error") else 200)
            return
        if parsed.path == "/api/export.xlsx":
            try:
                body, filename = export_excel(parse_qs(parsed.query))
            except RuntimeError as exc:
                if str(exc) == "excel_dependency_missing":
                    self.write_json({"error": "excel_dependency_missing"}, status=500)
                    return
                raise
            self.write_binary(
                body,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename,
            )
            return
        if parsed.path == "/api/export.pdf":
            try:
                body, filename = export_pdf(parse_qs(parsed.query))
            except RuntimeError as exc:
                if str(exc) == "pdf_dependency_missing":
                    self.write_json({"error": "pdf_dependency_missing"}, status=500)
                    return
                raise
            self.write_binary(body, "application/pdf", filename)
            return
        if parsed.path == "/api/quality":
            self.write_json({"issues": QUALITY_ISSUES, "summary": quality_summary(), "load_stats": LOAD_STATS})
            return
        if parsed.path == "/api/trace":
            record_id = parse_qs(parsed.query).get("id", [""])[0]
            payload = trace_record(record_id)
            if payload is None:
                self.write_json({"error": "record_not_found"}, status=404)
            else:
                self.write_json(payload)
            return
        if parsed.path == "/api/catalog/dates":
            self.write_json(STORE.meta["snapshots"])
            return
        if parsed.path == "/api/catalog/reporting-dates":
            self.write_json(reporting_dates_payload())
            return
        if parsed.path == "/api/catalog/sources":
            self.write_json(STORE.meta["sources"])
            return
        if parsed.path == "/api/catalog/budgets":
            self.write_json(STORE.meta["budgets"])
            return
        if parsed.path == "/api/catalog/templates":
            self.write_json(
                [{"code": code, "label": item["label"], "description": item["description"]} for code, item in TEMPLATES.items()]
            )
            return
        if parsed.path == "/api/catalog/metrics":
            self.write_json([{"code": code, "label": label} for code, label in METRICS.items()])
            return
        if parsed.path == "/api/catalog/quick-actions":
            self.write_json(quick_actions_payload())
            return
        if parsed.path == "/api/catalog/objects":
            self.write_json(catalog_objects(STORE.records, parse_qs(parsed.query)))
            return
        super().do_GET()

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/import":
            self.handle_import(parse_qs(parsed.query))
            return
        if parsed.path == "/api/review":
            try:
                length = int(self.headers.get("Content-Length", "0") or "0")
                payload = json.loads(self.rfile.read(length).decode("utf-8") or "{}")
                object_key = str(payload.get("object_key") or "").strip()
                result = update_review(object_key, payload)
                status = 400 if result.get("error") else 200
                self.write_json(result, status=status)
            except json.JSONDecodeError:
                self.write_json({"error": "invalid_json"}, status=400)
            return
        if parsed.path == "/api/assistant":
            try:
                length = int(self.headers.get("Content-Length", "0") or "0")
                payload = json.loads(self.rfile.read(length).decode("utf-8") or "{}")
                message = str(payload.get("message") or "").strip()
                context = payload.get("context") if isinstance(payload.get("context"), dict) else {}
                if not message:
                    self.write_json({"error": "message_required"}, status=400)
                    return
                self.write_json(assistant_response(message, context))
            except json.JSONDecodeError:
                self.write_json({"error": "invalid_json"}, status=400)
            return
        if parsed.path == "/api/explain":
            try:
                length = int(self.headers.get("Content-Length", "0") or "0")
                payload = json.loads(self.rfile.read(length).decode("utf-8") or "{}")
                kind = str(payload.get("kind") or "query")
                data = payload.get("payload") if isinstance(payload.get("payload"), dict) else {}
                self.write_json(explain_response(kind, data))
            except json.JSONDecodeError:
                self.write_json({"error": "invalid_json"}, status=400)
            return
        self.write_json({"error": "not_found"}, status=404)

    def handle_import(self, params: dict[str, list[str]] | None = None) -> None:
        global STORE
        try:
            length = int(self.headers.get("Content-Length", "0") or "0")
        except ValueError:
            self.write_json({"error": "invalid_request"}, status=400)
            return
        if length > IMPORT_MAX_BYTES:
            self.write_json({"error": "file_too_large"}, status=400)
            return
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": self.headers.get("Content-Type", ""),
                "CONTENT_LENGTH": str(length),
            },
        )
        source_type = str(form.getfirst("source_type") or "").strip()
        if source_type not in IMPORT_SOURCE_TYPES:
            self.write_json({"error": "source_type_required"}, status=400)
            return
        file_item = form["file"] if "file" in form else None
        if file_item is None or not getattr(file_item, "filename", ""):
            self.write_json({"error": "file_required"}, status=400)
            return
        filename = sanitize_upload_filename(file_item.filename)
        extension = Path(filename).suffix.lower()
        if extension not in {".csv", ".xlsx"}:
            self.write_json({"error": "invalid_extension"}, status=400)
            return
        content = file_item.file.read()
        if len(content) > IMPORT_MAX_BYTES:
            self.write_json({"error": "file_too_large"}, status=400)
            return
        target = save_import_upload(source_type, filename, content)
        try:
            STORE = load_data()
        except Exception:
            add_quality_issue(
                relative_source(target),
                "",
                "error",
                "import_parse_failed",
                "Не удалось разобрать импортированный файл",
                "file",
                filename,
            )
        self.write_json(import_payload(params or {}))

    def write_json(self, payload: object, status: int = 200) -> None:
        body = json.dumps(json_safe(payload), ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def write_binary(self, body: bytes, content_type: str, filename: str) -> None:
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)


async def fastapi_json_body(request: Request) -> dict | None:
    body = await request.body()
    if not body.strip():
        return {}
    payload = json.loads(body.decode("utf-8"))
    return payload if isinstance(payload, dict) else {}


@app.get("/")
def fastapi_index():
    return FileResponse(STATIC_DIR / "index.html", media_type="text/html")


@app.get("/api/meta")
def fastapi_meta():
    return fastapi_json(STORE.meta)


@app.get("/api/query")
def fastapi_query(request: Request):
    params = request_params(request)
    metrics = selected_metrics(params)
    if params.get("view", ["period"])[0].strip() == "as_of":
        return fastapi_json(query_as_of(params))
    filtered = apply_filters(STORE.records, params)
    result = aggregate(filtered, metrics)
    post_filter = params.get("post_filter", [""])[0].strip()
    return fastapi_json(apply_aggregate_post_filter(result, post_filter, metrics))


@app.get("/api/compare")
def fastapi_compare(request: Request):
    return fastapi_json(compare_periods(request_params(request)))


@app.get("/api/readiness")
def fastapi_readiness(request: Request):
    return fastapi_json(readiness_response(request_params(request)))


@app.get("/api/control")
def fastapi_control(request: Request):
    return fastapi_json(control_summary(request_params(request)))


@app.get("/api/reviews")
def fastapi_reviews():
    return fastapi_json(reviews_payload())


@app.get("/api/review")
def fastapi_review_get(request: Request):
    object_key = request_params(request).get("object_key", [""])[0].strip()
    if not object_key:
        return fastapi_json({"error": "object_key_required"}, status_code=400)
    return fastapi_json(review_for_object(object_key))


@app.get("/api/object")
def fastapi_object(request: Request):
    payload = object_detail(request_params(request))
    status_code = 404 if payload.get("error") == "object_not_found" else 400 if payload.get("error") else 200
    return fastapi_json(payload, status_code=status_code)


@app.get("/api/export.xlsx")
def fastapi_export_xlsx(request: Request):
    try:
        body, filename = export_excel(request_params(request))
    except RuntimeError as exc:
        if str(exc) == "excel_dependency_missing":
            return fastapi_json({"error": "excel_dependency_missing"}, status_code=500)
        raise
    return fastapi_binary(body, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename)


@app.get("/api/export.pdf")
def fastapi_export_pdf(request: Request):
    try:
        body, filename = export_pdf(request_params(request))
    except RuntimeError as exc:
        if str(exc) == "pdf_dependency_missing":
            return fastapi_json({"error": "pdf_dependency_missing"}, status_code=500)
        raise
    return fastapi_binary(body, "application/pdf", filename)


@app.get("/api/quality")
def fastapi_quality():
    return fastapi_json({"issues": QUALITY_ISSUES, "summary": quality_summary(), "load_stats": LOAD_STATS})


@app.get("/api/trace")
def fastapi_trace(request: Request):
    record_id = request_params(request).get("id", [""])[0]
    payload = trace_record(record_id)
    if payload is None:
        return fastapi_json({"error": "record_not_found"}, status_code=404)
    return fastapi_json(payload)


@app.get("/api/catalog/dates")
def fastapi_catalog_dates():
    return fastapi_json(STORE.meta["snapshots"])


@app.get("/api/catalog/reporting-dates")
def fastapi_catalog_reporting_dates():
    return fastapi_json(reporting_dates_payload())


@app.get("/api/catalog/sources")
def fastapi_catalog_sources():
    return fastapi_json(STORE.meta["sources"])


@app.get("/api/catalog/budgets")
def fastapi_catalog_budgets():
    return fastapi_json(STORE.meta["budgets"])


@app.get("/api/catalog/templates")
def fastapi_catalog_templates():
    return fastapi_json([{"code": code, "label": item["label"], "description": item["description"]} for code, item in TEMPLATES.items()])


@app.get("/api/catalog/metrics")
def fastapi_catalog_metrics():
    return fastapi_json([{"code": code, "label": label} for code, label in METRICS.items()])


@app.get("/api/catalog/quick-actions")
def fastapi_catalog_quick_actions():
    return fastapi_json(quick_actions_payload())


@app.get("/api/catalog/objects")
def fastapi_catalog_objects(request: Request):
    return fastapi_json(catalog_objects(STORE.records, request_params(request)))


@app.post("/api/review")
async def fastapi_review_post(request: Request):
    try:
        payload = await fastapi_json_body(request)
        object_key = str(payload.get("object_key") or "").strip()
        result = update_review(object_key, payload)
        return fastapi_json(result, status_code=400 if result.get("error") else 200)
    except (json.JSONDecodeError, UnicodeDecodeError):
        return fastapi_json({"error": "invalid_json"}, status_code=400)


@app.post("/api/assistant")
async def fastapi_assistant(request: Request):
    try:
        payload = await fastapi_json_body(request)
        message = str(payload.get("message") or "").strip()
        context = payload.get("context") if isinstance(payload.get("context"), dict) else {}
        if not message:
            return fastapi_json({"error": "message_required"}, status_code=400)
        return fastapi_json(assistant_response(message, context))
    except (json.JSONDecodeError, UnicodeDecodeError):
        return fastapi_json({"error": "invalid_json"}, status_code=400)


@app.post("/api/explain")
async def fastapi_explain(request: Request):
    try:
        payload = await fastapi_json_body(request)
        kind = str(payload.get("kind") or "query")
        data = payload.get("payload") if isinstance(payload.get("payload"), dict) else {}
        return fastapi_json(explain_response(kind, data))
    except (json.JSONDecodeError, UnicodeDecodeError):
        return fastapi_json({"error": "invalid_json"}, status_code=400)


@app.post("/api/import")
async def fastapi_import(
    request: Request,
    source_type: str = Form(""),
    file: UploadFile | None = File(None),
):
    global STORE
    source_type = str(source_type or "").strip()
    if source_type not in IMPORT_SOURCE_TYPES:
        return fastapi_json({"error": "source_type_required"}, status_code=400)
    if file is None or not file.filename:
        return fastapi_json({"error": "file_required"}, status_code=400)
    filename = sanitize_upload_filename(file.filename)
    extension = Path(filename).suffix.lower()
    if extension not in {".csv", ".xlsx"}:
        return fastapi_json({"error": "invalid_extension"}, status_code=400)
    content = await file.read()
    if len(content) > IMPORT_MAX_BYTES:
        return fastapi_json({"error": "file_too_large"}, status_code=400)
    target = save_import_upload(source_type, filename, content)
    try:
        STORE = load_data()
    except Exception:
        add_quality_issue(
            relative_source(target),
            "",
            "error",
            "import_parse_failed",
            "Не удалось разобрать импортированный файл",
            "file",
            filename,
        )
    return fastapi_json(import_payload(request_params(request)))


def main() -> None:
    port = int(os.environ.get("PORT") or (sys.argv[1] if len(sys.argv) > 1 else 8000))
    host = os.environ.get("HOST", "0.0.0.0")
    import uvicorn

    print(f"Loaded {STORE.meta['records']} records from {DATA_DIR}")
    print(f"Open http://{host}:{port}")
    uvicorn.run(app, host=host, port=port)


if __name__ == "__main__":
    main()
