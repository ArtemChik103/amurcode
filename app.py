from __future__ import annotations

import csv
import json
import os
import re
import sys
from io import BytesIO
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse
from urllib.request import Request, urlopen


ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "case"
STATIC_DIR = ROOT / "static"
RAG_DIR = ROOT / "docs" / "rag"
QUALITY_ISSUES: list[dict] = []
LOAD_STATS: dict[str, dict[str, int]] = {}
METRIC_KEYS = ["limit", "obligation", "cash", "agreement", "contract", "payment", "buau"]
METRICS = {
    "limit": "Лимиты",
    "obligation": "БО",
    "cash": "Касса",
    "agreement": "Соглашения",
    "contract": "Контракты",
    "payment": "Платежи",
    "buau": "БУ/АУ",
}
TEMPLATES = {
    "all": {"label": "Все данные", "description": "Без предметного шаблона"},
    "kik": {"label": "КИК", "description": "КЦСР содержит 975 или 978 с 6-й позиции"},
    "skk": {"label": "СКК", "description": "КЦСР содержит 6105 с 6-й позиции"},
    "two_thirds": {"label": "2/3", "description": "КЦСР содержит 970 с 6-й позиции"},
    "okv": {"label": "ОКВ", "description": "Капитальные вложения по КВР"},
}
QUICK_ACTIONS = {
    "show_skk": {
        "label": "Собрать отчет СКК",
        "description": "Готовая выборка по СКК на выбранную дату",
        "mode": "slice",
        "template": "skk",
        "metrics": ["limit", "obligation", "cash", "agreement", "contract", "payment", "buau"],
    },
    "show_kik": {
        "label": "Собрать отчет КИК",
        "description": "Готовая выборка по КИК на выбранную дату",
        "mode": "slice",
        "template": "kik",
        "metrics": ["limit", "obligation", "cash", "agreement", "contract", "payment"],
    },
    "show_two_thirds": {
        "label": "Собрать отчет 2/3",
        "description": "Готовая выборка по высвобождаемым средствам",
        "mode": "slice",
        "template": "two_thirds",
        "metrics": ["limit", "obligation", "cash", "agreement"],
    },
    "show_okv": {
        "label": "Собрать отчет ОКВ",
        "description": "Готовая выборка по капитальным вложениям",
        "mode": "slice",
        "template": "okv",
        "metrics": ["limit", "obligation", "cash", "contract", "payment", "buau"],
    },
    "compare_skk": {
        "label": "Сравнить две даты",
        "description": "Показать изменения между отчетными датами",
        "mode": "compare",
        "template": "skk",
        "metrics": ["limit", "obligation", "cash"],
    },
    "execution_problems": {
        "label": "Найти проблемные объекты",
        "description": "Нет документов, оплат, кассы или есть разрывы данных",
        "mode": "slice",
        "template": "all",
        "metrics": ["limit", "cash", "payment", "buau"],
        "post_filter": "execution_problems",
    },
    "demo_skk_problems": {
        "label": "Проблемные СКК",
        "description": "Показать проблемные объекты СКК и источник цифр",
        "mode": "slice",
        "template": "skk",
        "post_filter": "execution_problems",
        "metrics": ["limit", "obligation", "cash", "agreement", "contract", "payment", "buau"],
    },
    "find_object": {
        "label": "Найти объект",
        "description": "Поиск по названию, коду, бюджету или документу",
        "mode": "slice",
        "template": "all",
        "metrics": ["limit", "obligation", "cash", "agreement", "contract", "payment", "buau"],
    },
}
ASSISTANT_INTENTS = {
    "run_query",
    "run_compare",
    "explain_metric",
    "explain_template",
    "find_object",
    "show_execution_problems",
    "show_source",
    "help",
}
CAPITAL_KVR = {"400", "410", "411", "412", "413", "414", "415", "416", "417"}

SNAPSHOT_RE = re.compile(r"на\s+(\d{2}\.\d{2}\.\d{4})")
MONTHS = {
    "январь": 1,
    "февраль": 2,
    "март": 3,
    "апрель": 4,
    "май": 5,
    "июнь": 6,
    "июль": 7,
    "август": 8,
    "сентябрь": 9,
    "октябрь": 10,
    "ноябрь": 11,
    "декабрь": 12,
}


def relative_source(path: Path | str) -> str:
    try:
        return str(Path(path).resolve().relative_to(ROOT)).replace("\\", "/")
    except ValueError:
        return str(path).replace("\\", "/")


def add_quality_issue(
    source_file: str,
    source_row: int | str,
    severity: str,
    code: str,
    message: str,
    field: str = "",
    value: object = "",
) -> None:
    QUALITY_ISSUES.append(
        {
            "source_file": source_file,
            "source_row": source_row,
            "severity": severity,
            "code": code,
            "message": message,
            "field": field,
            "value": "" if value is None else str(value),
        }
    )


def parse_amount(value: object, source_file: str = "", source_row: int | str = "", field: str = "") -> float:
    if value is None:
        return 0.0
    text = str(value).strip().replace("\xa0", " ")
    if not text:
        return 0.0
    text = text.replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        if source_file or field:
            add_quality_issue(
                source_file,
                source_row,
                "warning",
                "amount_parse_failed",
                "Не удалось распарсить сумму",
                field,
                value,
            )
        return 0.0


def parse_date(value: object, source_file: str = "", source_row: int | str = "", field: str = "") -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    for fmt in ("%d.%m.%Y", "%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    if source_file or field:
        add_quality_issue(
            source_file,
            source_row,
            "warning",
            "date_parse_failed",
            "Не удалось распарсить дату",
            field,
            value,
        )
    return ""


def normalize_code(value: object) -> str:
    return re.sub(r"[^0-9A-Za-zА-Яа-я]", "", str(value or "")).upper()


def display_code(value: object) -> str:
    return str(value or "").strip()


def find_header_value(row: dict[str, str], *prefixes: str) -> str:
    for prefix in prefixes:
        for key, value in row.items():
            if str(key).strip().lower().startswith(prefix.lower()):
                return value
    return ""


def normalize_name(value: object) -> str:
    return " ".join(re.findall(r"[0-9A-Za-zА-Яа-яЁё]+", str(value or "").lower()))


def object_group_key(record: dict) -> str:
    budget_norm = normalize_name(record.get("budget"))
    code_norm = normalize_code(record.get("object_code_norm") or record.get("object_code"))
    if code_norm:
        return f"{code_norm}|{budget_norm}"
    return f"name:{normalize_name(record.get('object_name'))}|{budget_norm}"


def make_record(records: list[dict], source_file: str, source_row: int, raw: dict, **fields: object) -> dict:
    record = {
        "id": f"r{len(records) + 1}",
        "source_file": source_file,
        "source_row": source_row,
        "raw": dict(raw),
    }
    record.update(fields)
    return record


def selected_metrics(params: dict[str, list[str]] | None = None) -> list[str]:
    if not params:
        return list(METRIC_KEYS)
    raw = params.get("metrics", [""])[0].strip()
    if not raw:
        return list(METRIC_KEYS)
    result = [metric for metric in raw.split(",") if metric in METRIC_KEYS]
    return result or list(METRIC_KEYS)


def default_date_range() -> tuple[str, str]:
    snapshots = STORE.meta.get("reporting_dates", []) if "STORE" in globals() else []
    if not snapshots:
        return "", ""
    return snapshots[0], snapshots[-1]


def quick_actions_payload() -> list[dict]:
    return [{"code": code, **action} for code, action in QUICK_ACTIONS.items()]


def matches_template(record: dict, template: str) -> bool:
    if not template or template == "all":
        return True
    code = record.get("object_code_norm", "")
    if template == "kik":
        return code[5:8] in {"975", "978"}
    if template == "skk":
        return code[5:9] == "6105"
    if template == "two_thirds":
        return code[5:8] == "970"
    if template == "okv":
        return normalize_code(record.get("kvr")) in CAPITAL_KVR
    return True


def rcb_snapshot_from_rows(rows: list[list[str]], fallback_name: str) -> str:
    for row in rows[:8]:
        joined = " ".join(row)
        match = SNAPSHOT_RE.search(joined)
        if match:
            return parse_date(match.group(1))
    lower = fallback_name.lower()
    for name, month in MONTHS.items():
        if name in lower:
            year_match = re.search(r"(20\d{2})", lower)
            year = int(year_match.group(1)) if year_match else 2025
            next_month = month + 1
            next_year = year
            if next_month == 13:
                next_month = 1
                next_year += 1
            return f"{next_year:04d}-{next_month:02d}-01"
    return ""


def agreement_snapshot(row: dict[str, str], filename: str) -> str:
    period = row.get("period_of_date", "")
    if " - " in period:
        return parse_date(period.split(" - ")[-1])
    digits = re.search(r"(\d{2})(\d{2})(20\d{2})", filename)
    if digits:
        day, month, year = digits.groups()
        return f"{year}-{month}-{day}"
    return ""


def buau_snapshot(filename: str) -> str:
    lower = filename.lower()
    year_match = re.search(r"(20\d{2})", lower)
    year = int(year_match.group(1)) if year_match else 2025
    for name, month in MONTHS.items():
        if name in lower:
            return f"{year:04d}-{month:02d}-01"
    return ""


@dataclass
class DataStore:
    records: list[dict]
    meta: dict


def load_data() -> DataStore:
    QUALITY_ISSUES.clear()
    LOAD_STATS.clear()
    records: list[dict] = []
    load_rcb(records)
    load_agreements(records)
    load_state_task(records)
    load_buau(records)
    enrich_records(records)

    budgets = sorted({r["budget"] for r in records if r.get("budget")})
    sources = sorted({r["source"] for r in records})
    snapshots = sorted({r["snapshot"] for r in records if r.get("snapshot")})
    reporting_dates = sorted(
        {
            r["snapshot"]
            for r in records
            if r.get("snapshot") and r.get("record_kind") in {"rcb_snapshot", "agreement_snapshot"}
        }
    )
    objects = sorted(
        {r["object_name"] for r in records if r.get("object_name")},
        key=lambda x: x.lower(),
    )
    meta = {
        "records": len(records),
        "budgets": budgets,
        "sources": sources,
        "snapshots": snapshots,
        "reporting_dates": reporting_dates,
        "objects": objects[:500],
        "load_stats": LOAD_STATS,
        "quality": quality_summary(),
    }
    return DataStore(records=records, meta=meta)


def enrich_records(records: list[dict]) -> None:
    source_counts: dict[str, int] = defaultdict(int)
    source_folders = {
        "РЧБ": "case/1_RCB",
        "Соглашения": "case/2_Agreements",
        "ГЗ: контракты": "case/3_StateTask",
        "ГЗ: платежи": "case/3_StateTask",
        "БУАУ": "case/4_BUAU_Export",
    }
    for index, record in enumerate(records, start=1):
        source = record.get("source", "")
        source_counts[source] += 1
        record.setdefault("id", f"r{index}")
        record.setdefault("source_file", source_folders.get(source, ""))
        record.setdefault("source_row", source_counts[source])
        record.setdefault("record_kind", "event")
        record.setdefault("document_id", record.get("document_number", ""))
        record.setdefault("document_date", record.get("event_date", ""))
        record.setdefault("snapshot_source", "event")
        record.setdefault("amount_semantics", "event_amount")
        record.setdefault("raw", {key: value for key, value in record.items() if key not in {"raw"}})
        record["object_key"] = object_group_key(record)
    for source, count in source_counts.items():
        LOAD_STATS[source or "unknown"] = {"read_rows": count, "records": count, "warnings": 0, "errors": 0}


def quality_summary() -> dict[str, int]:
    return {
        "warnings": sum(1 for issue in QUALITY_ISSUES if issue["severity"] == "warning"),
        "errors": sum(1 for issue in QUALITY_ISSUES if issue["severity"] == "error"),
    }


def start_load_stats(path: Path) -> dict[str, int]:
    key = relative_source(path)
    stats = {"read_rows": 0, "records": 0, "warnings": 0, "errors": 0}
    LOAD_STATS[key] = stats
    return stats


def finish_load_stats(stats: dict[str, int], issue_start: int) -> None:
    issues = QUALITY_ISSUES[issue_start:]
    stats["warnings"] = sum(1 for issue in issues if issue["severity"] == "warning")
    stats["errors"] = sum(1 for issue in issues if issue["severity"] == "error")


def load_rcb(records: list[dict]) -> None:
    folder = DATA_DIR / "1_RCB"
    for path in sorted(folder.glob("*.csv")):
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.reader(handle, delimiter=";"))
        header_index = next(
            (i for i, row in enumerate(rows) if row and row[0].strip() == "Бюджет"),
            None,
        )
        if header_index is None:
            continue
        snapshot = rcb_snapshot_from_rows(rows, path.name)
        header = rows[header_index]
        source_file = relative_source(path)
        for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            if not any(cell.strip() for cell in row):
                continue
            item = dict(zip(header, row))
            kcsr = display_code(item.get("КЦСР"))
            object_name = item.get("Наименование КЦСР", "").strip() or kcsr
            budget = item.get("Бюджет", "").strip()
            records.append(
                make_record(
                    records,
                    source_file,
                    row_number,
                    item,
                    source="РЧБ",
                    snapshot=snapshot,
                    event_date=parse_date(item.get("Дата проводки")),
                    record_kind="rcb_snapshot",
                    document_id="",
                    document_date="",
                    snapshot_source="monthly_snapshot",
                    amount_semantics="balance_as_of",
                    budget=budget,
                    object_code=kcsr,
                    object_code_norm=normalize_code(kcsr),
                    object_name=object_name,
                    kfsr=display_code(item.get("КФСР")),
                    kvr=display_code(item.get("КВР")),
                    kosgu=display_code(item.get("КОСГУ")),
                    counterparty=item.get("Наименование КВСР", "").strip(),
                    document_number="",
                    description=item.get("Наименование КВР", "").strip(),
                    limit=parse_amount(find_header_value(item, "Лимиты ПБС"), source_file, row_number, "Лимиты ПБС"),
                    obligation=parse_amount(find_header_value(item, "Подтв. лимитов по БО"), source_file, row_number, "Подтв. лимитов по БО"),
                    cash=parse_amount(find_header_value(item, "Всего выбытий"), source_file, row_number, "Всего выбытий"),
                    agreement=0.0,
                    contract=0.0,
                    payment=0.0,
                    buau=0.0,
                )
            )


def load_agreements(records: list[dict]) -> None:
    folder = DATA_DIR / "2_Agreements"
    class_names = {
        "273": "МБТ",
        "278": "Иные цели БУ/АУ",
        "272": "Госзадание",
        "313": "ЮЛ/ИП/ФЛ",
    }
    for path in sorted(folder.glob("*.csv")):
        source_file = relative_source(path)
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row_number, row in enumerate(csv.DictReader(handle), start=2):
                kcsr = display_code(row.get("kcsr_code"))
                recipient = (row.get("dd_recipient_caption") or row.get("dd_estimate_caption") or "").strip()
                document_id = (row.get("document_id") or row.get("id") or row.get("reg_number") or "").strip()
                document_date = parse_date(row.get("document_date") or row.get("date") or row.get("agreement_date") or row.get("close_date"))
                records.append(
                    make_record(
                        records,
                        source_file,
                        row_number,
                        row,
                        source="Соглашения",
                        snapshot=agreement_snapshot(row, path.name),
                        event_date=parse_date(row.get("close_date")),
                        record_kind="agreement_snapshot",
                        document_id=document_id,
                        document_date=document_date,
                        snapshot_source="monthly_snapshot",
                        amount_semantics="balance_as_of",
                        budget=(row.get("caption") or "").replace("!!! НЕ РАБОТАТЬ !!!", "").strip(),
                        object_code=kcsr,
                        object_code_norm=normalize_code(kcsr),
                        object_name=recipient or kcsr,
                        kfsr=display_code(row.get("kfsr_code")),
                        kvr=display_code(row.get("kvr_code")),
                        kosgu=display_code(row.get("kesr_code")),
                        counterparty=recipient,
                        document_number=(row.get("reg_number") or "").strip(),
                        description=class_names.get(str(row.get("documentclass_id")), "Соглашение"),
                        limit=0.0,
                        obligation=0.0,
                        cash=0.0,
                        agreement=parse_amount(row.get("amount_1year")),
                        contract=0.0,
                        payment=0.0,
                        buau=0.0,
                    )
                )


def load_state_task(records: list[dict]) -> None:
    folder = DATA_DIR / "3_StateTask"
    budget_lines: dict[str, list[dict[str, str]]] = defaultdict(list)
    lines_path = folder / "Бюджетные строки.csv"
    contracts_path = folder / "Контракты и договора.csv"
    payments_path = folder / "Платежки.csv"

    if lines_path.exists():
        with lines_path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                budget_lines[row["con_document_id"]].append(row)

    contracts: dict[str, dict[str, str]] = {}
    if contracts_path.exists():
        source_file = relative_source(contracts_path)
        with contracts_path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row_number, row in enumerate(csv.DictReader(handle), start=2):
                contracts[row["con_document_id"]] = row
                document_date = parse_date(row.get("con_date"))
                for line in budget_lines.get(row["con_document_id"], [{}]):
                    kcsr = display_code(line.get("kcsr_code"))
                    records.append(
                        make_record(
                            records,
                            source_file,
                            row_number,
                            row,
                            source="ГЗ: контракты",
                            snapshot=document_date,
                            event_date=document_date,
                            record_kind="contract_document",
                            document_id=(row.get("con_document_id") or row.get("con_number") or "").strip(),
                            document_date=document_date,
                            snapshot_source="event",
                            amount_semantics="event_amount",
                            budget="",
                            object_code=kcsr,
                            object_code_norm=normalize_code(kcsr),
                            object_name=kcsr or row.get("con_number", "").strip(),
                            kfsr=display_code(line.get("kfsr_code")),
                            kvr=display_code(line.get("kvr_code")),
                            kosgu=display_code(line.get("kesr_code")),
                            counterparty=(row.get("zakazchik_key") or "").strip(),
                            document_number=(row.get("con_number") or "").strip(),
                            description="Контракт/договор",
                            limit=0.0,
                            obligation=0.0,
                            cash=0.0,
                            agreement=0.0,
                            contract=parse_amount(row.get("con_amount")),
                            payment=0.0,
                            buau=0.0,
                        )
                    )

    if payments_path.exists():
        source_file = relative_source(payments_path)
        with payments_path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row_number, row in enumerate(csv.DictReader(handle), start=2):
                contract = contracts.get(row["con_document_id"], {})
                related_lines = budget_lines.get(row["con_document_id"], [{}])
                document_date = parse_date(row.get("platezhka_paydate"))
                for line in related_lines:
                    kcsr = display_code(line.get("kcsr_code"))
                    records.append(
                        make_record(
                            records,
                            source_file,
                            row_number,
                            row,
                            source="ГЗ: платежи",
                            snapshot=document_date,
                            event_date=document_date,
                            record_kind="payment_event",
                            document_id=(row.get("platezhka_id") or row.get("platezhka_num") or "").strip(),
                            document_date=document_date,
                            snapshot_source="event",
                            amount_semantics="event_amount",
                            budget="",
                            object_code=kcsr,
                            object_code_norm=normalize_code(kcsr),
                            object_name=kcsr or contract.get("con_number", "").strip(),
                            kfsr=display_code(line.get("kfsr_code")),
                            kvr=display_code(line.get("kvr_code")),
                            kosgu=display_code(line.get("kesr_code")),
                            counterparty=(contract.get("zakazchik_key") or "").strip(),
                            document_number=(row.get("platezhka_num") or "").strip(),
                            description="Оплата по контракту",
                            limit=0.0,
                            obligation=0.0,
                            cash=0.0,
                            agreement=0.0,
                            contract=0.0,
                            payment=parse_amount(row.get("platezhka_amount")),
                            buau=0.0,
                        )
                    )


def load_buau(records: list[dict]) -> None:
    folder = DATA_DIR / "4_BUAU_Export"
    for path in sorted(folder.glob("*.csv")):
        snapshot = buau_snapshot(path.name)
        source_file = relative_source(path)
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row_number, row in enumerate(csv.DictReader(handle, delimiter=";"), start=2):
                kcsr = display_code(row.get("КЦСР"))
                organization = (row.get("Организация") or "").strip()
                event_date = parse_date(row.get("Дата проводки")) or snapshot
                records.append(
                    make_record(
                        records,
                        source_file,
                        row_number,
                        row,
                        source="БУАУ",
                        snapshot=snapshot,
                        event_date=event_date,
                        record_kind="buau_event",
                        document_id="",
                        document_date=event_date,
                        snapshot_source="event",
                        amount_semantics="event_amount",
                        budget=(row.get("Бюджет") or "").strip(),
                        object_code=kcsr,
                        object_code_norm=normalize_code(kcsr),
                        object_name=organization or kcsr,
                        kfsr=display_code(row.get("КФСР")),
                        kvr=display_code(row.get("КВР")),
                        kosgu=display_code(row.get("КОСГУ")),
                        counterparty=organization,
                        document_number="",
                        description=(row.get("Орган, предоставляющий субсидии") or "").strip(),
                        limit=0.0,
                        obligation=0.0,
                        cash=0.0,
                        agreement=0.0,
                        contract=0.0,
                        payment=0.0,
                        buau=parse_amount(row.get("Выплаты с учетом возврата")),
                    )
                )


def apply_filters(records: list[dict], params: dict[str, list[str]]) -> list[dict]:
    text = params.get("q", [""])[0].strip().lower()
    code = normalize_code(params.get("code", [""])[0])
    budget = params.get("budget", [""])[0].strip().lower()
    source = params.get("source", [""])[0].strip()
    start = params.get("start", [""])[0].strip()
    end = params.get("end", [""])[0].strip()
    template = params.get("template", ["all"])[0].strip() or "all"

    result = []
    for record in records:
        haystack = " ".join(
            str(record.get(key, ""))
            for key in ("object_name", "object_code", "budget", "counterparty", "document_number", "description")
        ).lower()
        if text and text not in haystack:
            continue
        if code and code not in record.get("object_code_norm", ""):
            continue
        if budget and budget not in record.get("budget", "").lower():
            continue
        if source and record.get("source") != source:
            continue
        if not matches_template(record, template):
            continue
        date_value = record.get("snapshot") or record.get("event_date") or ""
        if start and date_value and date_value < start:
            continue
        if end and date_value and date_value > end:
            continue
        result.append(record)
    return result


def reporting_dates_payload() -> list[dict]:
    return [{"date": date, "label": datetime.strptime(date, "%Y-%m-%d").strftime("%d.%m.%Y")} for date in STORE.meta.get("reporting_dates", [])]


def select_as_of(records: list[dict], date: str, params: dict[str, list[str]]) -> list[dict]:
    if not date:
        dates = STORE.meta.get("reporting_dates", []) if "STORE" in globals() else []
        date = dates[-1] if dates else ""
    latest_rcb = max(
        (record.get("snapshot") for record in records if record.get("record_kind") == "rcb_snapshot" and record.get("snapshot") and record["snapshot"] <= date),
        default="",
    )
    latest_agreement = max(
        (record.get("snapshot") for record in records if record.get("record_kind") == "agreement_snapshot" and record.get("snapshot") and record["snapshot"] <= date),
        default="",
    )
    selected = []
    for record in records:
        kind = record.get("record_kind")
        if kind == "rcb_snapshot":
            if record.get("snapshot") == latest_rcb:
                selected.append(record)
        elif kind == "agreement_snapshot":
            if record.get("snapshot") == latest_agreement:
                selected.append(record)
        elif kind in {"contract_document", "payment_event", "buau_event"}:
            event_date = record.get("document_date") or record.get("event_date") or record.get("snapshot") or ""
            if event_date and event_date <= date:
                selected.append(record)
    filter_params = {key: value for key, value in params.items() if key not in {"view", "date", "start", "end", "base", "target", "metrics", "post_filter"}}
    return apply_filters(selected, filter_params)


def percent_or_none(numerator: float, denominator: float) -> float | None:
    return numerator / denominator if denominator else None


def row_pipeline(row: dict) -> dict:
    plan = float(row.get("limit") or 0) + float(row.get("obligation") or 0)
    documents = float(row.get("agreement") or 0) + float(row.get("contract") or 0)
    paid = float(row.get("payment") or 0) + float(row.get("buau") or 0)
    cash = float(row.get("cash") or 0)
    missing_steps = []
    if plan > 0 and documents == 0:
        missing_steps.append("documents")
    if documents > 0 and paid == 0:
        missing_steps.append("payment")
    if plan > 0 and cash == 0:
        missing_steps.append("cash")
    return {
        "plan": plan,
        "documents": documents,
        "paid": paid,
        "cash": cash,
        "plan_to_documents_percent": percent_or_none(documents, plan),
        "documents_to_paid_percent": percent_or_none(paid, documents),
        "cash_to_plan_percent": percent_or_none(cash, plan),
        "missing_steps": missing_steps,
    }


def problem_reasons(row: dict) -> list[str]:
    pipeline = row.get("pipeline") or row_pipeline(row)
    reasons = []
    if "documents" in pipeline["missing_steps"]:
        reasons.append("no_documents")
    if "payment" in pipeline["missing_steps"]:
        reasons.append("no_payments")
    if "cash" in pipeline["missing_steps"]:
        reasons.append("no_cash")
    if pipeline["plan"] > 0 and pipeline["cash"] > 0 and pipeline["cash"] / pipeline["plan"] < 0.25:
        reasons.append("low_cash")
    if int(row.get("source_count") or 0) == 1:
        reasons.append("data_gap")
    return reasons


def row_status_from_reasons(reasons: list[str]) -> str:
    if any(reason in reasons for reason in ("no_documents", "no_payments", "no_cash")):
        return "danger"
    if any(reason in reasons for reason in ("low_cash", "data_gap")):
        return "warning"
    return "ok"


def aggregate(records: list[dict], metrics: list[str] | None = None) -> dict:
    groups: dict[str, dict] = {}
    timeline: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    metric_keys = metrics or list(METRIC_KEYS)

    totals = {key: 0.0 for key in metric_keys}
    for record in records:
        key = record.get("object_key") or object_group_key(record)
        row = groups.setdefault(
            key,
            {
                "object_key": key,
                "object_code": record.get("object_code") or "",
                "object_name": record.get("object_name") or "",
                "budget": record.get("budget") or "",
                "object_aliases": set(),
                "sources": set(),
                **{metric: 0.0 for metric in METRIC_KEYS},
            },
        )
        if not row["object_code"] and record.get("object_code"):
            row["object_code"] = record["object_code"]
        if (
            record.get("source") == "РЧБ"
            and record.get("object_name")
            and len(record["object_name"]) > len(row["object_name"])
        ):
            row["object_name"] = record["object_name"]
        if record.get("object_name"):
            row["object_aliases"].add(record["object_name"])
        row["sources"].add(record["source"])
        point = record.get("snapshot") or record.get("event_date") or "unknown"
        for metric in METRIC_KEYS:
            value = float(record.get(metric) or 0.0)
            row[metric] += value
            if metric in metric_keys:
                totals[metric] += value
            if metric in metric_keys and point != "unknown":
                timeline[point][metric] += value

    rows = []
    for row in groups.values():
        row["source_count"] = len(row["sources"])
        row["sources"] = ", ".join(sorted(row["sources"]))
        row["object_aliases"] = sorted(row["object_aliases"])[:10]
        row["match_confidence"] = "high" if not str(row["object_key"]).startswith("name:") else "medium"
        row["pipeline"] = row_pipeline(row)
        row["problem_reasons"] = problem_reasons(row)
        row["status"] = row_status_from_reasons(row["problem_reasons"])
        row["total"] = sum(row[metric] for metric in metric_keys)
        rows.append(row)
    rows.sort(key=lambda item: item["total"], reverse=True)

    timeline_rows = []
    for date in sorted(timeline):
        point = {"date": date}
        point.update({metric: timeline[date].get(metric, 0.0) for metric in metric_keys})
        timeline_rows.append(point)

    return {
        "totals": totals,
        "rows": rows[:300],
        "details": records[:500],
        "timeline": timeline_rows,
        "count": len(records),
    }


def apply_aggregate_post_filter(result: dict, post_filter: str, metrics: list[str]) -> dict:
    aliases = {"execution_problems": "execution_problems"}
    selected = aliases.get(post_filter, post_filter)
    valid = {"no_documents", "no_payments", "no_cash", "low_cash", "low_execution", "data_gap", "execution_problems"}
    if selected not in valid:
        return result
    rows = []
    for row in result.get("rows", []):
        reasons = row.get("problem_reasons") or problem_reasons(row)
        if selected == "execution_problems" and reasons:
            rows.append(row)
        elif selected == "low_execution":
            plan = float(row.get("limit") or 0) + float(row.get("obligation") or 0)
            execution = float(row.get("cash") or 0) + float(row.get("payment") or 0) + float(row.get("buau") or 0)
            if plan > 0 and (execution == 0 or execution / plan < 0.25):
                rows.append(row)
        elif selected in reasons:
            rows.append(row)
    filtered = dict(result)
    filtered["rows"] = rows
    filtered["totals"] = {metric: sum(float(row.get(metric) or 0) for row in rows) for metric in metrics}
    filtered["count"] = len(rows)
    return filtered


def catalog_objects(records: list[dict], params: dict[str, list[str]]) -> list[dict]:
    query = params.get("q", [""])[0].strip().lower()
    template = params.get("template", ["all"])[0].strip() or "all"
    seen: set[tuple[str, str, str]] = set()
    objects = []
    for record in records:
        if not matches_template(record, template):
            continue
        haystack = " ".join(
            str(record.get(key, ""))
            for key in ("object_code", "object_name", "budget", "counterparty", "document_number")
        ).lower()
        if query and query not in haystack:
            continue
        item = (
            record.get("object_code", ""),
            record.get("object_name", ""),
            record.get("budget", ""),
        )
        if item in seen:
            continue
        seen.add(item)
        objects.append({"code": item[0], "name": item[1], "budget": item[2]})
        if len(objects) >= 200:
            break
    return objects


def load_rag_documents() -> list[dict]:
    documents = []
    if not RAG_DIR.exists():
        return documents
    for path in sorted(RAG_DIR.glob("*.md")):
        text = path.read_text(encoding="utf-8").strip()
        if text:
            documents.append({"source_file": relative_source(path), "title": path.stem, "content": text})
    return documents


def retrieve_rag_context(message: str, limit: int = 4) -> str:
    query_words = {word for word in re.findall(r"[0-9A-Za-zА-Яа-яЁё/]+", message.lower()) if len(word) > 1}
    scored = []
    for document in load_rag_documents():
        content = document["content"].lower()
        score = sum(1 for word in query_words if word in content)
        if score:
            scored.append((score, document))
    scored.sort(key=lambda item: item[0], reverse=True)
    selected = [document for _, document in scored[:limit]]
    if not selected and load_rag_documents():
        selected = load_rag_documents()[:1]
    return "\n\n".join(f"[{item['source_file']}]\n{item['content']}" for item in selected)


def clean_search_text(message: str) -> str:
    text = re.sub(r"\b(покажи|показать|найди|найти|сравни|сравнить|что|такое|где|есть|по|в|и|с|со)\b", " ", message, flags=re.I)
    text = re.sub(r"\b(скк|кик|окв|лимит\w*|касс\w*|исполн\w*|платеж\w*|оплат\w*|динамик\w*|измен\w*)\b", " ", text, flags=re.I)
    text = re.sub(r"\b(6105|978|970|2/3)\b", " ", text)
    return " ".join(text.split())


def assistant_rule_based(message: str, context: dict | None = None) -> dict:
    context = context or {}
    lower = message.lower()
    start, end = default_date_range()
    action = {
        "mode": context.get("mode") or "slice",
        "template": context.get("template") or "all",
        "q": "",
        "code": "",
        "budget": "",
        "source": "",
        "date": context.get("date") or end,
        "start": start,
        "end": end,
        "metrics": context.get("selected_metrics") or ["limit", "obligation", "cash"],
    }
    intent = "run_query"
    confidence = 0.62

    if "скк" in lower or "6105" in lower:
        action["template"] = "skk"
        confidence = 0.9
    elif "кик" in lower or "978" in lower:
        action["template"] = "kik"
        confidence = 0.88
    elif "2/3" in lower or "970" in lower:
        action["template"] = "two_thirds"
        confidence = 0.88
    elif "окв" in lower or "капитал" in lower or "капвлож" in lower:
        action["template"] = "okv"
        confidence = 0.86

    if any(word in lower for word in ("сравн", "измен", "динамик")):
        intent = "run_compare"
        action["mode"] = "compare"
        action["base"] = start
        action["target"] = end
        confidence = max(confidence, 0.82)

    if any(word in lower for word in ("касс", "исполн", "платеж", "оплат")):
        action["metrics"] = ["cash", "payment", "buau"]
    if any(word in lower for word in ("лимит", "план", "бо")):
        action["metrics"] = ["limit", "obligation"]
    if any(word in lower for word in ("проблем", "нет касс", "низк")):
        intent = "show_execution_problems"
        action["template"] = "all"
        action["metrics"] = ["limit", "cash", "payment", "buau"]
        action["post_filter"] = "execution_problems"
    if "нет касс" in lower:
        action["post_filter"] = "no_cash"
    if "нет оплат" in lower or "нет платеж" in lower:
        action["post_filter"] = "no_payments"
    if "нет документ" in lower or "нет договор" in lower or "нет соглаш" in lower:
        action["post_filter"] = "no_documents"
    if "низк" in lower and ("исполн" in lower or "касс" in lower):
        action["post_filter"] = "low_cash"

    date_match = re.search(r"\b(\d{2})[.](\d{2})[.](20\d{2})\b", message)
    if date_match:
        day, month, year = date_match.groups()
        action["date"] = f"{year}-{month}-{day}"

    if any(phrase in lower for phrase in ("что такое", "объясни", "расскажи")):
        intent = "explain_metric" if any(word in lower for word in ("бо", "касс", "лимит", "метрик")) else "explain_template"
        confidence = max(confidence, 0.75)

    search_text = clean_search_text(message)
    if intent in {"run_query", "run_compare", "find_object"} and search_text:
        action["q"] = search_text

    rag_context = retrieve_rag_context(message, limit=2)
    explanation = ""
    if intent.startswith("explain") and rag_context:
        explanation = " " + " ".join(rag_context.split())[:700]

    alternative_label = "Искать во всех данных" if search_text else "Показать все данные"
    return {
        "mode": "rule_based",
        "intent": intent,
        "confidence": confidence,
        "message": f"Я понял запрос как {'сравнение' if intent == 'run_compare' else 'выборку'}: {TEMPLATES.get(action['template'], TEMPLATES['all'])['label']}.{explanation}",
        "action": action,
        "alternatives": [
            {
                "label": alternative_label,
                "action": {
                    "mode": "slice",
                    "template": "all",
                    "q": search_text,
                    "code": "",
                    "budget": "",
                    "source": "",
                    "post_filter": "",
                    "date": action.get("date", end),
                    "reset_scope": True,
                    "metrics": action["metrics"],
                },
            },
        ],
        "rag_context": rag_context,
    }


def validate_assistant_action(action: dict, fallback: dict) -> dict:
    result = dict(fallback)
    if not isinstance(action, dict):
        return result
    if action.get("mode") in {"slice", "compare"}:
        result["mode"] = action["mode"]
    if action.get("template") in TEMPLATES:
        result["template"] = action["template"]
    for key in ("q", "code", "budget", "source", "start", "end", "base", "target", "date", "post_filter"):
        if key in action:
            result[key] = str(action.get(key) or "")[:200]
    metrics = action.get("metrics")
    if isinstance(metrics, list):
        filtered = [metric for metric in metrics if metric in METRIC_KEYS]
        if filtered:
            result["metrics"] = filtered
    return result


def assistant_llm(message: str, context: dict, rag_context: str) -> dict:
    api_key = os.environ.get("GROQ_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError("GROQ_API_KEY is not set")
    model = os.environ.get("GROQ_MODEL", "openai/gpt-oss-120b").strip() or "openai/gpt-oss-120b"
    start, end = default_date_range()
    system_prompt = (
        "Ты помощник конструктора бюджетных выборок. Ты не считаешь суммы сам. "
        "Верни только JSON без markdown. Допустимые intent: "
        + ", ".join(sorted(ASSISTANT_INTENTS))
        + ". Допустимые шаблоны: "
        + ", ".join(TEMPLATES)
        + ". Допустимые метрики: "
        + ", ".join(METRIC_KEYS)
        + ". JSON: {\"intent\":\"run_query\",\"confidence\":0.8,\"message\":\"...\",\"action\":{...},\"alternatives\":[]}."
    )
    user_payload = {
        "message": message,
        "context": context,
        "available_dates": {"start": start, "end": end, "all": STORE.meta.get("reporting_dates", [])},
        "templates": {code: item["description"] for code, item in TEMPLATES.items()},
        "metrics": METRICS,
        "rag_context": rag_context,
    }
    body = json.dumps(
        {
            "model": model,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": json.dumps(user_payload, ensure_ascii=False)},
            ],
            "temperature": 0.1,
            "response_format": {"type": "json_object"},
        },
        ensure_ascii=False,
    ).encode("utf-8")
    request = Request(
        "https://api.groq.com/openai/v1/chat/completions",
        data=body,
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        method="POST",
    )
    with urlopen(request, timeout=15) as response:
        payload = json.loads(response.read().decode("utf-8"))
    content = payload["choices"][0]["message"]["content"]
    parsed = json.loads(content)
    fallback = assistant_rule_based(message, context)
    intent = parsed.get("intent") if parsed.get("intent") in ASSISTANT_INTENTS else fallback["intent"]
    action = validate_assistant_action(parsed.get("action", {}), fallback["action"])
    return {
        "mode": "llm",
        "intent": intent,
        "confidence": float(parsed.get("confidence") or fallback["confidence"]),
        "message": str(parsed.get("message") or fallback["message"])[:1000],
        "action": action,
        "alternatives": parsed.get("alternatives") if isinstance(parsed.get("alternatives"), list) else fallback["alternatives"],
        "rag_context": rag_context,
    }


def assistant_response(message: str, context: dict | None = None) -> dict:
    context = context or {}
    fallback = assistant_rule_based(message, context)
    if os.environ.get("ASSISTANT_ENABLED", "auto").lower() == "false":
        return fallback
    if not os.environ.get("GROQ_API_KEY", "").strip():
        return fallback
    try:
        return assistant_llm(message, context, fallback.get("rag_context", ""))
    except Exception:
        return fallback


def trace_record(record_id: str) -> dict | None:
    for record in STORE.records:
        if record.get("id") == record_id:
            normalized = {key: value for key, value in record.items() if key not in {"raw"}}
            amount_fields = []
            for key, label in (
                ("limit", "Лимиты"),
                ("obligation", "БО"),
                ("cash", "Касса"),
                ("agreement", "Соглашения"),
                ("contract", "Договоры"),
                ("payment", "Оплаты"),
                ("buau", "БУ/АУ"),
            ):
                value = float(record.get(key) or 0)
                if value:
                    amount_fields.append({"label": label, "field": key, "value": value})
            return {
                "id": record.get("id"),
                "source": record.get("source"),
                "source_file": record.get("source_file", ""),
                "source_row": record.get("source_row", ""),
                "human_summary": {
                    "title": "Сумма из исходной строки",
                    "date": record.get("event_date") or record.get("snapshot") or "",
                    "object": record.get("object_name") or record.get("object_code") or "",
                    "document": record.get("document_number") or record.get("description") or "",
                    "amount_fields": amount_fields,
                },
                "raw": record.get("raw", {}),
                "normalized": normalized,
            }
    return None


AS_OF_SEMANTICS = {
    "limit": "balance_as_of",
    "obligation": "balance_as_of",
    "cash": "balance_as_of",
    "agreement": "balance_as_of",
    "contract": "cumulative_to_date",
    "payment": "cumulative_to_date",
    "buau": "cumulative_to_date",
}


def readiness_check(status: str, code: str, label: str, message: str) -> dict:
    return {"code": code, "label": label, "status": status, "message": message}


def readiness_summary(records: list[dict], rows: list[dict], params: dict[str, list[str]]) -> dict:
    date = params.get("date", [""])[0].strip()
    if not date:
        dates = STORE.meta.get("reporting_dates", [])
        date = dates[-1] if dates else ""
    template = params.get("template", ["all"])[0].strip() or "all"
    sources = {record.get("source") for record in records}
    checks = [
        readiness_check(
            "ok" if "РЧБ" in sources else "bad",
            "rcb_loaded",
            "Плановые данные найдены",
            "Есть строки РЧБ на выбранную дату" if "РЧБ" in sources else "Нет строк РЧБ на выбранную дату",
        ),
        readiness_check(
            "ok" if "Соглашения" in sources else "warn",
            "agreements_loaded",
            "Соглашения найдены",
            "Есть соглашения на выбранную дату" if "Соглашения" in sources else "Соглашения не попали в выборку",
        ),
        readiness_check(
            "ok" if "ГЗ: контракты" in sources else "warn",
            "contracts_loaded",
            "Контракты найдены",
            "Есть контракты на выбранную дату" if "ГЗ: контракты" in sources else "Контракты есть не по всем объектам",
        ),
        readiness_check(
            "ok" if {"ГЗ: платежи", "БУАУ"} & sources else "warn",
            "payments_loaded",
            "Платежи найдены",
            "Есть платежи до выбранной даты" if {"ГЗ: платежи", "БУАУ"} & sources else "Платежи не попали в выборку",
        ),
    ]
    has_gaps = any("data_gap" in (row.get("problem_reasons") or []) for row in rows)
    checks.append(
        readiness_check(
            "warn" if has_gaps else "ok",
            "data_gaps",
            "Разрывы данных",
            "Есть объекты только в одном источнике" if has_gaps else "Критичных разрывов по источникам не видно",
        )
    )
    if not rows:
        checks.append(readiness_check("bad", "empty_result", "Выборка не пустая", "По выбранным условиям нет объектов"))
    else:
        checks.append(readiness_check("ok", "empty_result", "Выборка не пустая", "Есть объекты для показа"))
    summary = {status: sum(1 for check in checks if check["status"] == status) for status in ("ok", "warn", "bad")}
    return {"date": date, "template": template, "summary": summary, "checks": checks}


def query_as_of(params: dict[str, list[str]]) -> dict:
    date = params.get("date", [""])[0].strip()
    if not date:
        dates = STORE.meta.get("reporting_dates", [])
        date = dates[-1] if dates else ""
    metrics = selected_metrics(params)
    filtered = select_as_of(STORE.records, date, params)
    result = aggregate(filtered, metrics)
    result = apply_aggregate_post_filter(result, params.get("post_filter", [""])[0].strip(), metrics)
    result["timeline"] = as_of_timeline(date, params, metrics)
    return {
        "view": "as_of",
        "date": date,
        "semantics": AS_OF_SEMANTICS,
        **result,
    }


def as_of_timeline(date: str, params: dict[str, list[str]], metrics: list[str]) -> list[dict]:
    post_filter = params.get("post_filter", [""])[0].strip()
    dates = [item for item in STORE.meta.get("reporting_dates", []) if not date or item <= date]
    points = []
    for point_date in dates:
        point_params = {**params, "date": [point_date]}
        records = select_as_of(STORE.records, point_date, point_params)
        result = aggregate(records, metrics)
        result = apply_aggregate_post_filter(result, post_filter, metrics)
        point = {"date": point_date}
        point.update({metric: result["totals"].get(metric, 0.0) for metric in metrics})
        points.append(point)
    return points


def readiness_response(params: dict[str, list[str]]) -> dict:
    date = params.get("date", [""])[0].strip()
    if not date:
        dates = STORE.meta.get("reporting_dates", [])
        date = dates[-1] if dates else ""
        params = {**params, "date": [date]}
    metrics = selected_metrics(params)
    records = select_as_of(STORE.records, date, params)
    result = aggregate(records, metrics)
    result = apply_aggregate_post_filter(result, params.get("post_filter", [""])[0].strip(), metrics)
    return readiness_summary(records, result["rows"], params)


def object_detail(params: dict[str, list[str]]) -> dict:
    object_key = params.get("object_key", [""])[0].strip()
    if not object_key:
        return {"error": "object_key_required"}
    date = params.get("date", [""])[0].strip()
    if not date:
        dates = STORE.meta.get("reporting_dates", [])
        date = dates[-1] if dates else ""
    records = [record for record in select_as_of(STORE.records, date, params) if (record.get("object_key") or object_group_key(record)) == object_key]
    result = aggregate(records)
    if not result["rows"]:
        return {"error": "object_not_found"}
    row = result["rows"][0]
    documents = []
    for record in records:
        if record.get("source") in {"Соглашения", "ГЗ: контракты", "ГЗ: платежи", "БУАУ"}:
            amount = sum(float(record.get(metric) or 0) for metric in ("agreement", "contract", "payment", "buau"))
            documents.append(
                {
                    "source": record.get("source", ""),
                    "date": record.get("document_date") or record.get("event_date") or record.get("snapshot") or "",
                    "number": record.get("document_number", ""),
                    "counterparty": record.get("counterparty", ""),
                    "amount": amount,
                }
            )
    return {
        "object_key": object_key,
        "object_code": row.get("object_code", ""),
        "object_name": row.get("object_name", ""),
        "budget": row.get("budget", ""),
        "status": row.get("status", ""),
        "problem_reasons": row.get("problem_reasons", []),
        "pipeline": row.get("pipeline", {}),
        "sources": [source.strip() for source in str(row.get("sources", "")).split(",") if source.strip()],
        "documents": documents[:100],
        "records": records[:100],
    }


def export_excel(params: dict[str, list[str]]) -> tuple[bytes, str]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise RuntimeError("excel_dependency_missing") from exc

    mode = params.get("mode", ["slice"])[0].strip()
    template = params.get("template", ["all"])[0].strip() or "all"
    if mode == "compare":
        result = compare_periods(params)
        rows = result["rows"]
        details: list[dict] = []
        date_label = f"{result['base']} - {result['target']}"
        totals = {metric: sum(float(row.get("metrics", {}).get(metric, {}).get("target") or 0) for row in rows) for metric in selected_metrics(params)}
    else:
        result = query_as_of(params)
        rows = result["rows"]
        details = result["details"]
        date_label = result["date"]
        totals = result["totals"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Итоги"
    pipeline = row_pipeline(totals)
    problem_count = sum(1 for row in rows if row.get("problem_reasons"))
    summary_rows = [
        ("Дата отчета", date_label),
        ("Готовый отчет", TEMPLATES.get(template, TEMPLATES["all"])["label"]),
        ("Найдено объектов", len(rows)),
        ("План", pipeline["plan"]),
        ("Документы", pipeline["documents"]),
        ("Оплачено", pipeline["paid"]),
        ("Касса", pipeline["cash"]),
        ("Проблемных объектов", problem_count),
    ]
    ws.append(["Показатель", "Значение"])
    for item in summary_rows:
        ws.append(list(item))

    objects_ws = wb.create_sheet("Объекты")
    objects_ws.append(["Объект", "Код", "Бюджет", "План", "Документы", "Оплачено", "Касса", "Статус", "Причины", "Источники"])
    for row in rows:
        pipeline = row.get("pipeline") or row_pipeline(row)
        objects_ws.append([
            row.get("object_name", ""),
            row.get("object_code", ""),
            row.get("budget", ""),
            pipeline.get("plan", 0),
            pipeline.get("documents", 0),
            pipeline.get("paid", 0),
            pipeline.get("cash", 0),
            row.get("status", ""),
            ", ".join(row.get("problem_reasons") or []),
            row.get("sources", ""),
        ])

    problems_ws = wb.create_sheet("Проблемы")
    problems_ws.append(["Причина", "Объект", "Код", "План", "Документы", "Оплачено", "Касса", "Источники"])
    for row in rows:
        pipeline = row.get("pipeline") or row_pipeline(row)
        for reason in row.get("problem_reasons") or []:
            problems_ws.append([
                reason,
                row.get("object_name", ""),
                row.get("object_code", ""),
                pipeline.get("plan", 0),
                pipeline.get("documents", 0),
                pipeline.get("paid", 0),
                pipeline.get("cash", 0),
                row.get("sources", ""),
            ])

    details_ws = wb.create_sheet("Исходные строки")
    details_ws.append(["Дата", "Источник", "Файл", "Строка", "Код", "Объект", "Документ", "Контрагент", "Сумма"])
    for record in details:
        amount = sum(float(record.get(metric) or 0) for metric in METRIC_KEYS)
        details_ws.append([
            record.get("event_date") or record.get("snapshot") or "",
            record.get("source", ""),
            record.get("source_file", ""),
            record.get("source_row", ""),
            record.get("object_code", ""),
            record.get("object_name", ""),
            record.get("document_number", ""),
            record.get("counterparty", ""),
            amount,
        ])

    method_ws = wb.create_sheet("Методика")
    for line in (
        "РЧБ и соглашения берутся как последний месячный срез не позже выбранной даты.",
        "Контракты, платежи и БУАУ учитываются накопительно до выбранной даты.",
        "План = лимиты + БО.",
        "Документы = соглашения + контракты.",
        "Оплачено = платежи + БУАУ.",
        "Касса = кассовые выплаты из РЧБ.",
    ):
        method_ws.append([line])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for column in sheet.columns:
            width = min(55, max(12, *(len(str(cell.value or "")) + 2 for cell in column)))
            sheet.column_dimensions[column[0].column_letter].width = width

    buffer = BytesIO()
    wb.save(buffer)
    if mode == "compare":
        filename = f"analytics_compare_{template}_{params.get('base', [''])[0]}_{params.get('target', [''])[0]}.xlsx"
    else:
        filename = f"analytics_{template}_{params.get('date', ['as_of'])[0] or 'as_of'}.xlsx"
    return buffer.getvalue(), filename


def compare_periods(params: dict[str, list[str]]) -> dict:
    base = params.get("base", [""])[0].strip()
    target = params.get("target", [""])[0].strip()
    if not base or not target:
        dates = STORE.meta.get("reporting_dates", [])
        base = base or (dates[0] if dates else "")
        target = target or (dates[-1] if dates else "")
    metrics = selected_metrics(params)
    base_rows = aggregate(select_as_of(STORE.records, base, params), metrics)["rows"]
    target_rows = aggregate(select_as_of(STORE.records, target, params), metrics)["rows"]
    by_key: dict[tuple[str, str], dict] = {}
    for label, rows in (("base", base_rows), ("target", target_rows)):
        for row in rows:
            key = row.get("object_key") or object_group_key(row)
            item = by_key.setdefault(
                key,
                {
                    "object_key": key,
                    "object_code": row.get("object_code", ""),
                    "object_name": row.get("object_name", ""),
                    "budget": row.get("budget", ""),
                    "sources": row.get("sources", ""),
                    "metrics": {metric: {"base": 0.0, "target": 0.0, "delta": 0.0, "delta_percent": None} for metric in metrics},
                },
            )
            if row.get("sources"):
                item["sources"] = row["sources"]
            for metric in metrics:
                item["metrics"][metric][label] = row.get(metric, 0.0)
    rows = []
    for item in by_key.values():
        total_delta = 0.0
        for metric in metrics:
            values = item["metrics"][metric]
            values["delta"] = values["target"] - values["base"]
            values["delta_percent"] = (values["delta"] / values["base"] * 100.0) if values["base"] else None
            total_delta += abs(values["delta"])
        item["total_delta"] = total_delta
        rows.append(item)
    rows.sort(key=lambda item: item["total_delta"], reverse=True)
    return {
        "base": base,
        "target": target,
        "view": "as_of",
        "semantics": AS_OF_SEMANTICS,
        "metrics": metrics,
        "available_dates": STORE.meta["reporting_dates"],
        "rows": rows[:300],
    }


STORE = load_data()


class Handler(SimpleHTTPRequestHandler):
    def translate_path(self, path: str) -> str:
        parsed = urlparse(path)
        requested = parsed.path
        if requested == "/":
            return str(STATIC_DIR / "index.html")
        if requested.startswith("/static/"):
            return str(ROOT / requested.lstrip("/"))
        return str(STATIC_DIR / requested.lstrip("/"))

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/meta":
            self.write_json(STORE.meta)
            return
        if parsed.path == "/api/query":
            params = parse_qs(parsed.query)
            metrics = selected_metrics(params)
            if params.get("view", ["period"])[0].strip() == "as_of":
                self.write_json(query_as_of(params))
                return
            filtered = apply_filters(STORE.records, params)
            result = aggregate(filtered, metrics)
            post_filter = params.get("post_filter", [""])[0].strip()
            self.write_json(apply_aggregate_post_filter(result, post_filter, metrics))
            return
        if parsed.path == "/api/compare":
            self.write_json(compare_periods(parse_qs(parsed.query)))
            return
        if parsed.path == "/api/readiness":
            self.write_json(readiness_response(parse_qs(parsed.query)))
            return
        if parsed.path == "/api/object":
            payload = object_detail(parse_qs(parsed.query))
            self.write_json(payload, status=404 if payload.get("error") == "object_not_found" else 400 if payload.get("error") else 200)
            return
        if parsed.path == "/api/export.xlsx":
            try:
                body, filename = export_excel(parse_qs(parsed.query))
            except RuntimeError as exc:
                if str(exc) == "excel_dependency_missing":
                    self.write_json({"error": "excel_dependency_missing"}, status=500)
                    return
                raise
            self.write_binary(
                body,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename,
            )
            return
        if parsed.path == "/api/quality":
            self.write_json({"issues": QUALITY_ISSUES, "summary": quality_summary(), "load_stats": LOAD_STATS})
            return
        if parsed.path == "/api/trace":
            record_id = parse_qs(parsed.query).get("id", [""])[0]
            payload = trace_record(record_id)
            if payload is None:
                self.write_json({"error": "record_not_found"}, status=404)
            else:
                self.write_json(payload)
            return
        if parsed.path == "/api/catalog/dates":
            self.write_json(STORE.meta["snapshots"])
            return
        if parsed.path == "/api/catalog/reporting-dates":
            self.write_json(reporting_dates_payload())
            return
        if parsed.path == "/api/catalog/sources":
            self.write_json(STORE.meta["sources"])
            return
        if parsed.path == "/api/catalog/budgets":
            self.write_json(STORE.meta["budgets"])
            return
        if parsed.path == "/api/catalog/templates":
            self.write_json(
                [{"code": code, "label": item["label"], "description": item["description"]} for code, item in TEMPLATES.items()]
            )
            return
        if parsed.path == "/api/catalog/metrics":
            self.write_json([{"code": code, "label": label} for code, label in METRICS.items()])
            return
        if parsed.path == "/api/catalog/quick-actions":
            self.write_json(quick_actions_payload())
            return
        if parsed.path == "/api/catalog/objects":
            self.write_json(catalog_objects(STORE.records, parse_qs(parsed.query)))
            return
        super().do_GET()

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/assistant":
            try:
                length = int(self.headers.get("Content-Length", "0") or "0")
                payload = json.loads(self.rfile.read(length).decode("utf-8") or "{}")
                message = str(payload.get("message") or "").strip()
                context = payload.get("context") if isinstance(payload.get("context"), dict) else {}
                if not message:
                    self.write_json({"error": "message_required"}, status=400)
                    return
                self.write_json(assistant_response(message, context))
            except json.JSONDecodeError:
                self.write_json({"error": "invalid_json"}, status=400)
            return
        self.write_json({"error": "not_found"}, status=404)

    def write_json(self, payload: object, status: int = 200) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def write_binary(self, body: bytes, content_type: str, filename: str) -> None:
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)


def main() -> None:
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 8000
    server = ThreadingHTTPServer(("127.0.0.1", port), Handler)
    print(f"Loaded {STORE.meta['records']} records from {DATA_DIR}")
    print(f"Open http://127.0.0.1:{port}")
    server.serve_forever()


if __name__ == "__main__":
    main()
